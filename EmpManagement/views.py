from django.shortcuts import render
from django.conf import settings
from datetime import date
import logging
from django.utils import timezone
from openpyxl import load_workbook
from .models import (emp_family,Emp_Documents,EmpJobHistory,EmpLeaveRequest,EmpQualification,GeneralRequest,RequestType,
                     emp_master,notification,EmpFamily_CustomField,EmpJobHistory_CustomField,
                     EmpQualification_CustomField,EmpDocuments_CustomField,LanguageSkill,MarketingSkill,ProgrammingLanguageSkill,Emp_CustomField,Report,Doc_Report,GeneralRequest,RequestType,GeneralRequestReport,EmployeeLangSkill,EmployeeProgramSkill,
                     EmployeeMarketingSkill,Approval,ApprovalLevel,RequestNotification,Emp_CustomFieldValue,
                     EmailTemplate,EmailConfiguration,SelectedEmpNotify,NotificationSettings,DocExpEmailTemplate,CommonWorkflow,Doc_CustomFieldValue,EmployeeBankDetail,Fam_CustomFieldValue,Qualification_CustomFieldValue,JobHistory_CustomFieldValue,
                     DocumentApprovalLevel,DocumentApproval,DocumentRequest,ResignationApprovalLevel,ResignationApproval,DocRequestEmailTemplate,DocRequestNotification,EndOfService,EmployeeResignation,DocRequestType,ResignationEmailTemplate,ResignationRequestNotification,
                     ApprovalWorkflow,DocumentApprovalWorkflow,ResignationApprovalWorkflow,document_type
                     )
from .serializer import (Emp_qf_Serializer,EmpFamSerializer,EmpSerializer,NotificationSerializer,RequestTypeSerializer,
                         EmpJobHistorySerializer,EmpLeaveRequestSerializer,DocumentSerializer,GeneralRequestSerializer,
                         GeneralReportSerializer,EmpMarketSkillSerializer,EmployeeReportSerializer,EmpBulkUploadSerializer,CustomFieldSerializer,
                         EmpFam_CustomFieldSerializer,EmpJobHistory_Udf_Serializer,Emp_qf_udf_Serializer,EmpDocuments_Udf_Serializer,
                         DocBulkuploadSerializer,DocumentReportSerializer,EmpPrgrmSkillSerializer,EmpLangSkillSerializer,ApprovalSerializer,ApprovalLevelSerializer,
                         ReqNotifySerializer,Emp_CustomFieldValueSerializer,EmailTemplateSerializer,EmployeeFilterSerializer,EmailConfigurationSerializer,SelectedEmpNotifySerializer,
                         NotificationSettingsSerializer,DocExpEmailTemplateSerializer,CommonWorkflowSerializer,DOC_CustomFieldValueSerializer,EmpBankDetailsSerializer,EmpBankBulkuploadSerializer,EmplistSerializer,Fam_CustomFieldValueSerializer,
                         Qualification_CustomFieldValueSerializer,JobHistory_CustomFieldValueSerializer,DocApprovalLevelSerializer,DocApprovalSerializer,DocRequestSerializer,ResignationApprovalLevelSerializer,ResignationApprovalSerializer,
                         DocRequestEmailTemplateSerializer,DocRequestNotificationSerializer,EndOfServiceSerializer,EmployeeResignationSerializer,DocRequestTypeSerializer,EscalationRuleSerializer,ResignationTemplateSerializer,ResignationRequestNotificationSerializer,
                         ApprovalWorkflowSerializer,DocumentApprovalWorkflowSerializer,ResignationApprovalWorkflowSerializer,Document_typeSerializer)

from .resource import EmployeeResource,DocumentResource,EmpCustomFieldValueResource,EmpDocumentCustomFieldValueResource,EmpBankDetailsResource, MarketingSkillResource,ProLangSkillResource
from .permissions import (IsSuperUserOrHasGeneralRequestPermission,IsSuperUserOrInSameBranch,EmpCustomFieldPermission,EmpCustomFieldValuePermission,
                        EmpFamilyCustomFieldPermission,EmpJobHistoryCustomFieldPermission,EmpQualificationCustomFieldPermission,ReportPermission,DocReportPermission,GeneralRequestReportPermission,
                        EmployeeMarketingSkillPermission,EmployeeProgramSkillPermission,EmployeeLangSkillPermission,NotificationPermission,ApprovalLevelPermission,EmployeeMarketingSkillPermission,RequestTypePermission,
                        CanViewApprovedResignations,CanCreateEOS,EmployeeResignationPermission,DocTypePermission)
from django.core.exceptions import ValidationError
from rest_framework.decorators import action
from phonenumber_field.modelfields import PhoneNumberField
from rest_framework import viewsets,filters
from tablib import Dataset
from io import BytesIO
import os,json
from django.db.models import Field
from rest_framework.response import Response
from rest_framework.views import APIView
from django.http import HttpResponse,JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Alignment,Font,NamedStyle,Border, Side
from rest_framework import status,generics,viewsets,permissions
from .permissions import EmployeePermission
from datetime import datetime, timedelta
from OrganisationManager.models import DocumentNumbering
from OrganisationManager.serializer import AnnouncementSerializer,AssetAllocationSerializer
from rest_framework.parsers import MultiPartParser, FormParser
import pandas as pd,openpyxl
from django.views.decorators.csrf import csrf_exempt
from rest_framework.authentication import BaseAuthentication
from rest_framework.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404, redirect
from django.core.cache import cache
import redis
import json
from calendars .serializer import AttendanceSerializer,LateinEarlyoutRequestSerializer
from rest_framework.exceptions import NotFound
from calendars .serializer import EmployeeLeaveBalanceSerializer,LeaveTypeSerializer
from calendars .models import leave_type, employee_leave_request
from django.db.models import Q
from PayrollManagement .serializer import PayslipSerializer,LoanApplicationSerializer,AdvanceSalaryRequestSerializer,AirTicketRequestSerializer,EmployeeSalaryStructureSerializer
from PayrollManagement .models import EmployeeSalaryStructure
from .utils import calculate_settlement,send_notification_email
import csv
import io
from django.db import models
from Core .mixins import BranchAccessMixin
from django.core.mail import send_mail
from .utils import send_notification_email,get_employee_context
from UserManagement.models import CustomUser



r = redis.StrictRedis(host='localhost', port=6379, db=0)

class CustomAuthentication(BaseAuthentication):
    def authenticate(self, request):
        user = request.user
        if user.is_authenticated:
            companies = user.companies.all()
            if companies:
                # If user has associated companies, return the user and None
                return user, None
            else:
                # If user has no associated companies, raise PermissionDenied
                raise PermissionDenied("You do not have access to any schemas.")
        return None

#EMPLOYEE CRUD
class EmpViewSet(viewsets.ModelViewSet):
    queryset = emp_master.objects.all()
    serializer_class = EmpSerializer
    # permission_classes = [EmployeePermission]
    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()

        # ESS user → show only own employee profile
        if user.is_authenticated and getattr(user, 'is_ess', False):
            return queryset.filter(users=user)

        # HR / Admin → show all employees
        return queryset
    # def get_queryset(self):
    #     user = self.request.user
    #     if user.is_authenticated:
    #         if hasattr(user, 'is_ess') and user.is_ess:  # If user is an ESS, they can only access their own employee information
    #             return emp_master.objects.filter(users=user)
    #         else:
    #             return emp_master.objects.all()  # Other users can access all employee information
    #     return emp_master.objects.none()
    @action(detail=True, methods=['POST', 'GET'])
    def emp_family(self, request, pk=None):
        employee = self.get_object()

        if request.method == 'POST':
    # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpFamSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_family.all()
            serializer = EmpFamSerializer(family_members, many=True)
            return Response(serializer.data)
    
    @action(detail=True, methods=['POST', 'GET'])
    def emp_qualification(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
        # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = Emp_qf_Serializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_qualification.all()
            serializer = Emp_qf_Serializer(family_members, many=True)
            return Response(serializer.data)

    @action(detail=True, methods=['POST', 'GET'])
    def emp_job_history(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
    # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpJobHistorySerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_job_history.all()
            serializer = EmpJobHistorySerializer(family_members, many=True)
            return Response(serializer.data)
    



    @action(detail=True, methods=['POST', 'GET','DELETE'])
    def emp_documents(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
    # Add the employee.pk to the request data
            data = request.data.copy()
            data['emp_id'] = employee.pk

            serializer = DocumentSerializer(data=data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_documents.all()
            serializer = DocumentSerializer(family_members, many=True)
            return Response(serializer.data)
    
    # @action(detail=True, methods=['POST', 'GET'])
    @action(detail=True, methods=['POST', 'GET', 'DELETE'])
    def emp_market_skills(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
    # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpMarketSkillSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_market_skills.all()
            serializer = EmpMarketSkillSerializer(family_members, many=True)
            return Response(serializer.data)
            
        elif request.method == 'DELETE':
            # Delete all emp_market_skills related to the employee
            skills = employee.emp_market_skills.all()
            skills_count = skills.count()
            skills.delete()
            return Response({"detail": f"{skills_count} skills deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


    @action(detail=True, methods=['POST', 'GET','DELETE'])
    def emp_programlangskill(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
        # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpPrgrmSkillSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_prgrm_skills.all()
            serializer = EmpPrgrmSkillSerializer(family_members, many=True)
            return Response(serializer.data)

        elif request.method == 'DELETE':
            # Delete all emp_market_skills related to the employee
            skills = employee.emp_prgrm_skills.all()
            skills_count = skills.count()
            skills.delete()
            return Response({"detail": f"{skills_count} skills deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


    @action(detail=True, methods=['POST', 'GET','DELETE'])
    def emp_languageskill(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'POST':
        # Add the employee.pk to the request data
            request.data['emp_id'] = employee.pk

            serializer = EmpLangSkillSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            family_members = employee.emp_lang_skills.all()
            serializer = EmpLangSkillSerializer(family_members, many=True)
            return Response(serializer.data)
        
        elif request.method == 'DELETE':
            # Delete all emp_market_skills related to the employee
            skills = employee.emp_lang_skills.all()
            skills_count = skills.count()
            skills.delete()
            return Response({"detail": f"{skills_count} skills deleted successfully."}, status=status.HTTP_204_NO_CONTENT)
    
    
    @action(detail=True, methods=['get'])
    def attendance(self, request, pk=None):
        user = self.get_object()
        attendance = user.get_attendance()
        serializer = AttendanceSerializer(attendance, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def leave_balance(self, request, pk=None):
        user = self.get_object()

        # Get leave balance
        leave_balance = user.get_leave_balance()

        # Get pending leave requests for the employee
        pending_leave_types = employee_leave_request.objects.filter(
            employee=user, status="pending"
        ).values_list("leave_type", flat=True)  # Extract leave_type IDs

        # Exclude leave types that are in pending requests
        available_leave_types = leave_type.objects.exclude(id__in=pending_leave_types)

        # Serialize data
        leave_balance_serializer = EmployeeLeaveBalanceSerializer(leave_balance, many=True)
        leave_type_serializer = LeaveTypeSerializer(available_leave_types, many=True)

        return Response({
            "leave_balance": leave_balance_serializer.data,
            "available_leave_types": leave_type_serializer.data,
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['GET'])
    def emp_payslip(self, request, pk=None):
        employee = self.get_object()
        # Only fetch payslips where confirm_status is True
        payslips = employee.payslips.filter(status="Approved")
        serializer = PayslipSerializer(payslips, many=True)
        return Response(serializer.data)
    @action(detail=True, methods=['GET'])
    def emp_asset(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'GET':
            payslip = employee.allocations.all()
            serializer = AssetAllocationSerializer(payslip, many=True)
            return Response(serializer.data)
    @action(detail=True, methods=['GET'])
    def emp_announcement(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'GET':
            payslip = employee.employee_announcements.all()
            serializer = AnnouncementSerializer(payslip, many=True)
            return Response(serializer.data)
    @action(detail=True, methods=['GET'])
    def emp_loan(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'GET':
            payslip = employee.loan.all()
            serializer = LoanApplicationSerializer(payslip, many=True)
            return Response(serializer.data) 
    @action(detail=True, methods=['GET'])
    def emp_adv_salary(self, request, pk=None):
        employee = self.get_object()
        if request.method == 'GET':
            payslip = employee.advance_salary_requests.all()
            serializer = AdvanceSalaryRequestSerializer(payslip, many=True)
            return Response(serializer.data)
    @action(detail=True, methods=['POST', 'GET'])
    def emp_bank_details(self, request, pk=None):
        employee = self.get_object()

        if request.method == 'POST':
            request.data['employee'] = employee.pk

            serializer = EmpBankDetailsSerializer(data=request.data, context={'request': request})
            serializer.is_valid(raise_exception=True)  # Raise exception for invalid data
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        elif request.method == 'GET':
            emp_banks = employee.bank_details.all()
            serializer = EmpBankDetailsSerializer(emp_banks, many=True)
            return Response(serializer.data)
    @action(detail=True, methods=['GET'])
    def emp_projects(self, request, pk=None):
        employee = self.get_object()
        projects = Project.objects.filter(
            models.Q(managers=employee) | models.Q(members=employee)
        ).distinct()

        serializer = ProjectSerializer(projects, many=True)
        return Response(serializer.data)
        return Response(serializer.data)
        return Response(serializer.data)
        return Response(serializer.data)
    @action(detail=True, methods=['GET'])
    def emp_resignation(self, request, pk=None):
        employee = self.get_object()
        resignation= employee.resignation_requests.filter(status='APPROVED'  ) # ✅ MUST match model choices
        serializer = EmployeeResignationSerializer(resignation, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['GET'])
    def lateinearlyout_requests(self, request, pk=None):
        employee = self.get_object()
        requests = employee.lateinearlyout_requests.filter(status='APPROVED')
        serializer = LateinEarlyoutRequestSerializer(requests, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['GET'])
    def my_profile(self, request):
        user = request.user

        try:
            employee = emp_master.objects.get(users=user)
        except emp_master.DoesNotExist:
            return Response({"detail": "No employee record linked to this user."}, status=404)

        serializer = self.get_serializer(employee)
        return Response(serializer.data)
    @action(detail=False, methods=['get'])
    def filter_empty_user_non_ess(self, request):
        filtered_employees = self.queryset.filter(users__isnull=True, is_ess=False)
        serializer = EmployeeFilterSerializer(filtered_employees, many=True)
        return Response(serializer.data)
    @action(detail=True, methods=['GET'])
    def emp_airticket(self, request, pk=None):
        employee = self.get_object()
        airtickets = employee.airticket_requests.filter(status='APPROVED'  ) # ✅ MUST match model choices
        serializer = AirTicketRequestSerializer(airtickets, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    @action(detail=True, methods=['GET'])
    def emp_document_requests(self, request, pk=None):
        employee = self.get_object()
        document_requests = employee.document_requests.all()
        serializer = DocRequestSerializer(document_requests, many=True)
        return Response(serializer.data)
    @action(detail=True, methods=['GET'])
    def salary(self, request, pk=None):
        from django.db.models import Sum
        from decimal import Decimal
        employee = self.get_object()

        salary = EmployeeSalaryStructure.objects.filter(
            employee=employee,
            is_active=True,
            component__component_value_type="fixed"   # <-- only fixed components now
        ).select_related("component")

        serializer = EmployeeSalaryStructureSerializer(salary, many=True)

        gross = salary.filter(
            component__component_type="addition"
        ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

        deductions = salary.filter(
            component__component_type="deduction"
        ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

        return Response({
            "employee": employee.emp_code,
            "employee_name": f"{employee.emp_first_name} {employee.emp_last_name}",
            "gross_salary": gross,
            "total_deductions": deductions,
            "net_salary": gross - deductions,
            "components": serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def export_employee_data(self, request):
        excluded_fields = {'id', 'is_ess', 'created_at', 'created_by', 'updated_at', 'updated_by', 'emp_profile_pic'}
        display_names = {
            "emp_code": "Employee Code",
            "emp_first_name": "First Name",
            "emp_last_name": "Last Name",
            "emp_gender": "Gender",
            "emp_date_of_birth": "Date of Birth",
            "emp_personal_email": "Email",
            "emp_mobile_number_1": "Mobile Number",
            "emp_mobile_number_2": "Mobile Number2",
            "emp_country_id": "Country",
            "emp_state_id": "State",
            "emp_city": "City",
            "emp_permenent_address": "Permanent Address",
            "emp_present_address": "Present Address",
            "emp_status": "Status",
            "emp_hired_date": "Hired Date",
            "emp_active_date": "Active Date",
            "emp_relegion": "Religion",
            "emp_blood_group": "Blood Group",
            "emp_nationality_id": "Nationality",
            "emp_marital_status": "Marital Status",
            "emp_father_name": "Father Name",
            "emp_mother_name": "Mother Name",
            "emp_posting_location": "Posting Location",
            "is_active": "Active",
            "epm_ot_applicable": "OT Applicable",
            "emp_company_id": "Company",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category"
        }

        # Fetch all employees
        employees = emp_master.objects.all()

        # Fetch all distinct field names from Emp_CustomField
        custom_fields = Emp_CustomField.objects.values_list('emp_custom_field', flat=True).distinct()

        # Prepare headers combining emp_master fields and custom_fields
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        headers = emp_master_fields + list(custom_fields)

        # Create an Excel workbook
        wb = Workbook()
        ws = wb.active

        # Define default cell formats
        default_fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')  # pink background color
        default_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write headers with display names
        for col_num, header in enumerate(headers, start=1):
            display_name = display_names.get(header, header.capitalize())
            ws.cell(row=1, column=col_num, value=display_name)  # Capitalize header names
            ws.cell(row=1, column=col_num).fill = default_fill
            ws.cell(row=1, column=col_num).alignment = default_alignment

        # Write employee data and custom fields
        for row_num, employee in enumerate(employees, start=2):
            for col_num, header in enumerate(headers, start=1):
                if header in custom_fields:
                    # Fetch custom field value
                    custom_field_value = Emp_CustomFieldValue.objects.filter(emp_master=employee, emp_custom_field=header).first()
                    value = custom_field_value.field_value if custom_field_value else ''
                else:
                    value = getattr(employee, header, '')

                ws.cell(row=row_num, column=col_num, value=str(value))
                ws.cell(row=row_num, column=col_num).alignment = default_alignment

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Save the workbook to an in-memory buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Prepare response
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Empployee_data.xlsx"'
        return response

    @action(detail=False, methods=['get'])
    def export_employee_data_csv(self, request):
        excluded_fields = {
            'id', 'is_ess', 'created_at', 'created_by',
            'updated_at', 'updated_by', 'emp_profile_pic'
        }
        display_names = {
            "emp_code": "Employee Code",
            "emp_first_name": "First Name",
            "emp_last_name": "Last Name",
            "emp_gender": "Gender",
            "emp_date_of_birth": "Date of Birth",
            "emp_personal_email": "Email",
            "emp_mobile_number_1": "Mobile Number",
            "emp_mobile_number_2": "Mobile Number2",
            "emp_country_id": "Country",
            "emp_state_id": "State",
            "emp_city": "City",
            "emp_permenent_address": "Permanent Address",
            "emp_present_address": "Present Address",
            "emp_status": "Status",
            "emp_hired_date": "Hired Date",
            "emp_active_date": "Active Date",
            "emp_relegion": "Religion",
            "emp_blood_group": "Blood Group",
            "emp_nationality_id": "Nationality",
            "emp_marital_status": "Marital Status",
            "emp_father_name": "Father Name",
            "emp_mother_name": "Mother Name",
            "emp_posting_location": "Posting Location",
            "is_active": "Active",
            "epm_ot_applicable": "OT Applicable",
            "emp_company_id": "Company",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category"
        }

        # Fetch all employees
        employees = emp_master.objects.all()

        # Fetch all distinct field names from Emp_CustomField
        custom_fields = Emp_CustomField.objects.values_list('emp_custom_field', flat=True).distinct()

        # Prepare headers
        emp_master_fields = [
            field.name for field in emp_master._meta.get_fields()
            if isinstance(field, Field) and field.name not in excluded_fields
        ]
        headers = emp_master_fields + list(custom_fields)

        # Prepare HTTP response with CSV content type
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Employee_data.csv"'

        writer = csv.writer(response)

        # Write header row
        header_row = [display_names.get(h, h.capitalize()) for h in headers]
        writer.writerow(header_row)

        # Write employee rows
        for employee in employees:
            row = []
            for header in headers:
                if header in custom_fields:
                    custom_field_value = Emp_CustomFieldValue.objects.filter(
                        emp_master=employee, emp_custom_field=header
                    ).first()
                    value = custom_field_value.field_value if custom_field_value else ''
                else:
                    value = getattr(employee, header, '')
                row.append(str(value))
            writer.writerow(row)

        return response

class ReportViewset(viewsets.ModelViewSet):
    queryset = Report.objects.all()
    serializer_class = EmployeeReportSerializer
    permission_classes = [ReportPermission]
    # permission_classes = [IsSuperUserOrInSameBranch]

     
    # def get_queryset(self):
    #     user = self.request.user
    #     print ("user",user)
    #     # If superuser, return all reports
    #     if user.is_superuser:
    #         return Report.objects.all()
       
    #    # Filter reports based on user's branch
    #     if user.is_authenticated:
    #         print("authenticated")
    #         if hasattr(user, 'branches'):
            
    #             print("all")  
    #             user_branch_id = user.branches
    #             print("branch",user_branch_id)
    #             return Report.objects.filter(branch_id=user_branch_id)
            
    #     # If user does not have branch_id (should not happen with proper user model setup), return empty queryset or handle as needed
    #     return Report.objects.none()
   

    def __init__(self, *args, **kwargs):
        super(ReportViewset, self).__init__(*args, **kwargs)
        self.ensure_standard_report_exists()

    def get_available_fields(self):
        excluded_fields = {'id', 'is_ess','created_at', 'created_by', 'updated_at', 'updated_by', 'emp_profile_pic'}
        display_names = {
            "emp_code": "Employee Code",
            "emp_first_name": "First Name",
            "emp_last_name": "Last Name",
            "emp_gender": "Gender",
            "emp_date_of_birth": "Date of Birth",
            "emp_personal_email": "Email",
            "emp_mobile_number_1": "Mobile Number",
            "emp_mobile_number_2": "Mobile Number2",
            "emp_country_id": "Country",
            "emp_state_id": "State",
            "emp_city": "City",
            "emp_permenent_address": "Permanent Address",
            "emp_present_address": "Present Address",
            "emp_status": "Status",
            "emp_hired_date": "Hired Date",
            "emp_active_date": "Active Date",
            "emp_relegion": "Religion",
            "emp_blood_group": "Blood Group",
            "emp_nationality_id": "Nationality",
            "emp_marital_status": "Marital Status",
            "emp_father_name": "Father Name",
            "emp_mother_name": "Mother Name",
            "emp_posting_location": "Posting Location",
            "is_active": "Active",
            "epm_ot_applicable": "OT Applicable",
            "emp_company_id": "Company",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category"
        }
        
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        emp_custom_fields = list(Emp_CustomField.objects.values_list('emp_custom_field', flat=True))        
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + emp_custom_fields} 
        return available_fields

    @action(detail=False, methods=['get'])
    def select_employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
        

    @csrf_exempt
    @action(detail=False, methods=['post'])
    def emp_select_report(self, request, *args, **kwargs):
        # if not request.user.is_superuser:
        #     return Response({"error": "You do not have permission to access this resource."}, status=status.HTTP_403_FORBIDDEN)
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'reports')  # Default to 'report' if 'file_name' is not provided
                fields_to_include = request.POST.getlist('fields', [])
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())

            employees = emp_master.objects.all()

            report_data = self.generate_report_data(fields_to_include, employees)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')  # Use 'file_name' provided by the user

            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format

            Report.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({
                'status': 'success',
                'file_path': file_path,
                'selected_fields_data': fields_to_include,
                
            })

        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
    def ensure_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if Report.objects.filter(file_name='std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'std_report'
            fields_to_include = self.get_available_fields().keys()
            employees = emp_master.objects.all()

            report_data = self.generate_report_data(fields_to_include, employees)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            Report.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )

            print("Standard report generated successfully.")

        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = Report.objects.get(file_name='std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except Report.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    

    def generate_report_data(self, fields_to_include, employees):
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        custom_fields = list(Emp_CustomFieldValue.objects.filter(emp_master__in=employees).values_list('emp_custom_field', flat=True).distinct())

        report_data = []
        for employee in employees:
            employee_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(employee, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()  # Convert date to ISO format string
                elif field in custom_fields:
                    # Fetch the custom field value directly from Emp_CustomFieldValue
                    custom_field_value = Emp_CustomFieldValue.objects.filter(
                        emp_master=employee, 
                        emp_custom_field=field
                    ).first()
                    value = custom_field_value.field_value if custom_field_value else 'N/A'
                else:
                    value = 'N/A'
                employee_data[field] = value
            report_data.append(employee_data)

        return report_data

    
    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])  # Get selected fields from session
        print("selected fields:",selected_fields)
        report_id = request.GET.get('report_id')  # Get report_id from query parameters

        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        })   

        

    @csrf_exempt
    @action(detail=False, methods=['post'])
    def generate_employee_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.POST.get('report_id')
        available_fields = self.get_available_fields()
       
        # Save selected fields to session
        request.session['selected_fields'] = selected_fields
        print("select fields",selected_fields)
        # Fetch report data based on report_id
        try:
            report = Report.objects.get(id=report_id)
            report_file_path = os.path.join(settings.MEDIA_ROOT, report.report_data.name)  # Assuming report_data is a FileField
            with open(report_file_path, 'r') as file:
                report_content = json.load(file)  # Load content of the report file as JSON
        except Report.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'})
        print("reportcontnt",report_content)
        # If no fields are selected for filtration, default to all existing fields in the report
        if not selected_fields:
            if report_content:
                selected_fields = list(report_content[0].keys())  # Default to all keys in the first record
            else:
                selected_fields = []  # No data in the report

        # Fetch employees data from emp_master
        employees = emp_master.objects.all()

        # Get unique values for selected_fields
        unique_values = self.get_unique_values_for_fields(employees, selected_fields, report_content)
        processed_unique_values = {}
        for field, values in unique_values.items():
            processed_unique_values[field] = {
                'values': values,
            }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': report_content,  # Pass report_content to the frontend
            'unique_values': processed_unique_values,
        })

       

    def get_unique_values_for_fields(self, employees, selected_fields, report_content):
        unique_values = {field: set() for field in selected_fields}

        # Extract data from the JSON content
        for record in report_content:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        # Fetch additional data from Emp_CustomField if necessary
        for field in selected_fields:
            if field not in unique_values:
                continue
            for employee in employees:
                if not hasattr(employee, field):
                    custom_field_value = Emp_CustomField.objects.filter(emp_master=employee, field_name=field).first()
                    if custom_field_value:
                        unique_values[field].add(custom_field_value.field_value)

        # Convert sets to lists
        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values
    
    
    @csrf_exempt
    @action(detail=False, methods=['post'])
    def filter_existing_report(self, request, *args, **kwargs):
        report_id = request.data.get('report_id')
        if not report_id:
            return HttpResponse('Report ID is missing', status=400)

        try:
            report_instance = Report.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except (Report.DoesNotExist, json.JSONDecodeError) as e:
            return HttpResponse(f'Report not found or invalid JSON format: {str(e)}', status=404)

        selected_fields = [key for key in request.data.keys() if key != 'report_id']
        filter_criteria = {}

        for field in selected_fields:
            values = [val.strip() for val in request.data.getlist(field) if val.strip()]
            if values:
                filter_criteria[field] = values

        field_names = {
            "Employee Code": "emp_code",
            "First Name": "emp_first_name",
            "Last Name": "emp_last_name",
            "Gender": "emp_gender",
            "Date of Birth": "emp_date_of_birth",
            "Email": "emp_personal_email",
            "Mobile Number": "emp_mobile_number_1",
            "Mobile Number2": "emp_mobile_number_2",
            "Country": "emp_country_id",
            "State": "emp_state_id",
            "City": "emp_city",
            "Permanent Address": "emp_permenent_address",
            "Present Address": "emp_present_address",
            "Status": "emp_status",
            "Hired Date": "emp_hired_date",
            "Active Date": "emp_active_date",
            "Religion": "emp_relegion",
            "Blood Group": "emp_blood_group",
            "Nationality": "emp_nationality_id",
            "Marital Status": "emp_marital_status",
            "Father Name": "emp_father_name",
            "Mother Name": "emp_mother_name",
            "Posting Location": "emp_posting_location",
            "Active": "is_active",
            "OT Applicable": "epm_ot_applicable",
            "Company": "emp_company_id",
            "Branch": "emp_branch_id",
            "Department": "emp_dept_id",
            "Designation": "emp_desgntn_id",
            "Category": "emp_ctgry_id",
            # Add other field mappings as per your needs
        }

        filtered_data = [row for row in report_data if self.match_filter_criteria(row, filter_criteria, field_names)]
        print("filtered data",filtered_data)
        # Save filtered data to session for Excel generation
        request.session['filtered_data'] = filtered_data
        request.session.modified = True
        display_named = self.get_available_fields()

        return JsonResponse({
        'filtered_data': filtered_data,
        'report_id': report_id,
    })
        

    def match_filter_criteria(self, row_data, filter_criteria, field_names):
        for column_heading, field_name in field_names.items():
            if field_name in filter_criteria:
                values = filter_criteria[field_name]
                row_value = row_data.get(field_name)
                if row_value is None or row_value.strip() not in values:
                    return False
        for custom_field_name in filter_criteria.keys():
            if custom_field_name not in field_names.values():
                custom_field_values = filter_criteria[custom_field_name]
                custom_field_value = row_data.get(custom_field_name, '').strip().lower()
                if custom_field_value and custom_field_value not in [val.lower() for val in custom_field_values]:
                    return False
        return True

    

    
    @action(detail=False, methods=['get'])
    def generate_excel(self, request, *args, **kwargs):
        report_id = request.GET.get('report_id')
        if not report_id:
            return HttpResponse('Report ID is missing', status=400)

        filtered_data = request.session.get('filtered_data')
        if not filtered_data:
            return HttpResponse('No filtered data available', status=400)

        # Mapping of internal field names to display names
        field_names_mapping = {
            "emp_code": "Employee Code",
            "emp_first_name": "First Name",
            "emp_last_name": "Last Name",
            "emp_gender": "Gender",
            "emp_date_of_birth": "Date of Birth",
            "emp_personal_email": "Email",
            "emp_mobile_number_1": "Mobile Number",
            "emp_mobile_number_2": "Mobile Number2",
            "emp_country_id": "Country",
            "emp_state_id": "State",
            "emp_city": "City",
            "emp_permenent_address": "Permanent Address",
            "emp_present_address": "Present Address",
            "emp_status": "Status",
            "emp_hired_date": "Hired Date",
            "emp_active_date": "Active Date",
            "emp_relegion": "Religion",
            "emp_blood_group": "Blood Group",
            "emp_nationality_id": "Nationality",
            "emp_marital_status": "Marital Status",
            "emp_father_name": "Father Name",
            "emp_mother_name": "Mother Name",
            "emp_posting_location": "Posting Location",
            "is_active": "Active",
            "epm_ot_applicable": "OT Applicable",
            "emp_company_id": "Company",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
        }

        try:
            report_instance = Report.objects.get(id=int(report_id))
        except (Report.DoesNotExist, ValueError):
            return HttpResponse('Invalid or missing Report ID', status=404)

        # Create an Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Filtered Report'
        
        # Define style for header row
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

        # Add header row to Excel using display names and apply style
        if filtered_data:
            headers = [field_names_mapping.get(field_name, field_name) for field_name in filtered_data[0].keys()]
            sheet.append(headers)
            for cell in sheet[1]:
                cell.style = header_style

        # Add data rows to Excel using values from filtered_data
        for row in filtered_data:
            row_values = [row.get(field_name, '') for field_name in filtered_data[0].keys()]
            sheet.append(row_values)

        # Autofit column widths
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save the workbook to a BytesIO stream
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Prepare the response with Excel file as attachment
        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=filtered_report_{report_id}.xlsx'

        return response

class CustomFieldViewset(viewsets.ModelViewSet):
    queryset = Emp_CustomField.objects.all()
    serializer_class = CustomFieldSerializer
    permission_classes = [EmpCustomFieldPermission]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in emp_master._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
    
    # def get_available_fields(self):
    #     emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) ]                
    #     return emp_master_fields

    # @action(detail=False, methods=['get'])
    # def employee_fields(self, request, *args, **kwargs):
    #     available_fields = self.get_available_fields()
    #     return Response({'available_fields': available_fields})


class Emp_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = Emp_CustomFieldValue.objects.all()
    serializer_class = Emp_CustomFieldValueSerializer
    permission_classes = [EmpCustomFieldValuePermission]

    
      
class EmpFam_CustomFieldViewset(viewsets.ModelViewSet):
    queryset = EmpFamily_CustomField.objects.all()
    serializer_class = EmpFam_CustomFieldSerializer
    # permission_classes = [EmpFamilyCustomFieldPermission]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in emp_family._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
class Fam_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = Fam_CustomFieldValue.objects.all()
    serializer_class = Fam_CustomFieldValueSerializer


class EmpJobHistory_UdfViewset(viewsets.ModelViewSet):
    queryset = EmpJobHistory_CustomField.objects.all()
    serializer_class = EmpJobHistory_Udf_Serializer
    # permission_classes = [EmpJobHistoryCustomFieldPermission]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in EmpJobHistory._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})

class JobHistory_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = JobHistory_CustomFieldValue.objects.all()
    serializer_class = JobHistory_CustomFieldValueSerializer

class EmpQf_UdfViewset(viewsets.ModelViewSet):
    queryset = EmpQualification_CustomField.objects.all()
    serializer_class = Emp_qf_udf_Serializer
    # permission_classes = [EmpQualificationCustomFieldPermission]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in EmpQualification._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})

class Qf_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = Qualification_CustomFieldValue.objects.all()
    serializer_class = Qualification_CustomFieldValueSerializer



class EmpDoc_UdfViewset(viewsets.ModelViewSet):
    queryset = EmpDocuments_CustomField.objects.all()
    serializer_class = EmpDocuments_Udf_Serializer
    # permission_classes = [IsAuthenticated]

    def handle_exception(self, exc):
        if isinstance(exc, ValidationError):
            error_messages = [str(error) for error in exc]
            error_message = ', '.join(error_messages)
            return Response({'error': error_message}, status=status.HTTP_400_BAD_REQUEST)
        
        return super().handle_exception(exc)
    
    def get_available_fields(self):
        # Get the field names along with their data types
        # emp_master_fields = [
        #     {'name': field.name, 'type': field.get_internal_type()}
        #     for field in emp_master._meta.get_fields()
        #     if isinstance(field, Field)
        # ]
        # return emp_master_fields
        emp_master_fields = [
        {'name': field.name, 'type': field.__class__.__name__}
        for field in Emp_Documents._meta.get_fields()
        if isinstance(field, Field)
        ]
        return emp_master_fields
    
    @action(detail=False, methods=['get'])
    def employee_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
class Doc_CustomFieldValueViewSet(viewsets.ModelViewSet):
    queryset = Doc_CustomFieldValue.objects.all()
    serializer_class = DOC_CustomFieldValueSerializer
    # permission_classes = [EmpCustomFieldValuePermission]



class EmpbulkuploadViewSet(viewsets.ModelViewSet):
    queryset = emp_master.objects.all()
    serializer_class = EmpBulkUploadSerializer
    parser_classes = (MultiPartParser, FormParser)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if 'file' not in request.FILES:
            return Response({"error": "Please provide a file."}, status=400)

        upload_file = request.FILES['file']
        filename = upload_file.name.lower()
        all_errors = {"sheet1_errors": [], "sheet2_errors": []}

        try:
            dataset_sheet1 = Dataset()
            dataset_sheet2 = Dataset()
            dataset_sheet1.headers = []
            dataset_sheet2.headers = ['Employee Code', 'Field Name', 'Field Value']

            # ---------------- XLS/XLSX ----------------
            if filename.endswith(('.xlsx', '.xls')):
                workbook = load_workbook(upload_file, data_only=True)

                # Sheet 1: EmployeeMaster
                if "EmployeeMaster" not in workbook.sheetnames:
                    return Response({"error": "Sheet1 (EmployeeMaster) missing"}, status=400)
                sheet1 = workbook["EmployeeMaster"]
                dataset_sheet1.headers = [cell.value for cell in sheet1[1]]
                for row in sheet1.iter_rows(min_row=2):
                    dataset_sheet1.append([cell.value for cell in row])

                # Sheet 2: UDF
                if "UDF" in workbook.sheetnames:
                    sheet2 = workbook["UDF"]
                    if sheet2.max_row > 1:
                        headers = [cell.value for cell in sheet2[1]]

                        # Wide format parsing: all columns after Employee Code are UDF
                        emp_code_index = headers.index("Employee Code")
                        for row in sheet2.iter_rows(min_row=2):
                            emp_code = str(row[emp_code_index].value or "")
                            for col_idx, field_name in enumerate(headers):
                                if field_name == "Employee Code":
                                    continue
                                field_value = row[col_idx].value
                                if field_value not in (None, ''):
                                    dataset_sheet2.append([emp_code, field_name, str(field_value)])

            # ---------------- CSV ----------------
            elif filename.endswith('.csv'):
                file_data = upload_file.read().decode("utf-8")
                csv_reader = csv.DictReader(io.StringIO(file_data))
                headers = csv_reader.fieldnames

                employee_columns = [f.column_name for f in EmployeeResource().fields.values()]
                dataset_sheet1.headers = employee_columns

                for row in csv_reader:
                    emp_row = [row.get(col, '') for col in employee_columns]
                    dataset_sheet1.append(emp_row)

                    emp_code = row.get('Employee Code', '')

                    # Wide format UDF columns
                    udf_columns = [h for h in headers if h not in employee_columns]
                    for udf_col in udf_columns:
                        field_value = row.get(udf_col, '')
                        if field_value not in (None, ''):
                            dataset_sheet2.append([emp_code, udf_col, str(field_value)])

            else:
                return Response({"error": "Invalid file format. Only .xlsx, .xls, .csv supported."}, status=400)

            # ---------------- Validation ----------------
            employee_resource = EmployeeResource()
            custom_field_resource = EmpCustomFieldValueResource()

            # EmployeeMaster validation
            for row_idx, row in enumerate(dataset_sheet1.dict, start=2):
                try:
                    employee_resource.before_import_row(row, row_idx=row_idx)
                except ValidationError as e:
                    all_errors["sheet1_errors"].append({"row": row_idx, "error": str(e)})

            # UDF validation
            for row_idx, row in enumerate(dataset_sheet2.dict, start=2):
                try:
                    field_name = (row.get('Field Name') or '').strip()
                    field_value = row.get('Field Value', '')
                    emp_code = row.get('Employee Code', '')

                    if not field_name:
                        raise ValidationError("Field Name cannot be empty")
                    if not emp_code:
                        raise ValidationError("Employee Code cannot be empty")

                    # Check if custom field exists
                    custom_field = Emp_CustomField.objects.filter(emp_custom_field=field_name).first()
                    if not custom_field:
                        raise ValidationError(f"Custom field '{field_name}' does not exist.")

                    # Handle date fields
                    if custom_field.data_type == 'date' and field_value:
                        if isinstance(field_value, (datetime, date)):
                            field_value = field_value.strftime('%d-%m-%Y')
                        elif isinstance(field_value, str):
                            field_value = field_value.strip()
                            if ' ' in field_value:
                                field_value = field_value.split(' ')[0]
                            for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
                                try:
                                    parsed_date = datetime.strptime(field_value, fmt)
                                    field_value = parsed_date.strftime('%d-%m-%Y')
                                    break
                                except ValueError:
                                    continue
                            else:
                                raise ValidationError(
                                    f"Invalid date format for field '{field_name}'. Expected DD-MM-YYYY or DD/MM/YYYY."
                                )

                    # Update row with normalized value
                    row['Field Value'] = field_value

                except ValidationError as e:
                    all_errors["sheet2_errors"].append({"row": row_idx, "error": str(e)})

            if all_errors["sheet1_errors"] or all_errors["sheet2_errors"]:
                return Response({"errors": all_errors}, status=400)

            # ---------------- Import ----------------
            with transaction.atomic():
                employee_result = employee_resource.import_data(dataset_sheet1, dry_run=False, raise_errors=True)

            if dataset_sheet2:
                with transaction.atomic():
                    custom_field_result = custom_field_resource.import_data(dataset_sheet2, dry_run=False, raise_errors=True)
                return Response({
                    "message": f"{employee_result.total_rows} Employee records created, "
                               f"{custom_field_result.total_rows} UDF records created successfully"
                })

            return Response({"message": f"{employee_result.total_rows} Employee records created successfully."})

        except Exception as e:
            return Response({"error": str(e)}, status=400)

    
    @action(detail=False, methods=['get'])
    def download_default_excel_file(self, request):
        resource = EmployeeResource()
        headers = [field.column_name for field in resource.fields.values()]

        wb = Workbook()

        # ======== Common Styles ========
        black_font = Font(color="000000", bold=True)
        blue_fill = PatternFill(start_color="FF87CEEB", end_color="FF87CEEB", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")  # light cream/yellow
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Helper function to style header row
        def style_header_row(ws, max_cols=10):
            """Style header row with blue fill and black bold text across full width."""
            for col in range(1, max_cols + 1):
                cell = ws.cell(row=1, column=col)
                if not cell.value:
                    cell.value = ""
                cell.fill = blue_fill
                cell.font = black_font
                cell.border = border_style
                ws.column_dimensions[cell.column_letter].width = 25
            ws.freeze_panes = "A2"  # freeze header

        # ======================================================
        # Sheet 1: EmployeeMaster
        # ======================================================
        ws1 = wb.active
        ws1.title = "EmployeeMaster"

        for col_num, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col_num, value=header)
        style_header_row(ws1, max_cols=len(headers) + 5)

        # ======================================================
        # Sheet 2: UDF
        # ======================================================
        ws2 = wb.create_sheet(title="UDF")
        udf_headers = ["Employee Code"]
        for col_num, header in enumerate(udf_headers, 1):
            ws2.cell(row=1, column=col_num, value=header)
        style_header_row(ws2, max_cols=10)

        # ======================================================
        # Sheet 3: Instructions
        # ======================================================
        ws3 = wb.create_sheet(title="Instructions")
        ws3.sheet_view.showGridLines = False

        # Header Row
        ws3.cell(row=1, column=1, value="Instructions")
        ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
        style_header_row(ws3, max_cols=1)

        # Detailed instruction content
        instruction_text = (
            "📘 INSTRUCTIONS FOR EMPLOYEE BULK UPLOAD\n\n"
            "➡ SHEET 1: EmployeeMaster\n"
            "   • Employee Code - Must be unique for each employee (no duplicates).\n"
            "   • Employee First Name, Gender, DOB, Joining Date, and Confirmation Date are mandatory.\n"
            "   • DOB, Joining Date, and Confirmation Date must be in format DD/MM/YYYY.\n"
            "   • Branch, Department, Designation, and Category must match existing master data.\n"
            "   • Work Location, Visa Location must match existing branch names.\n"
            "   • Country and State names must match the database (case-insensitive).\n"
            "   • Nationality and Religion must exist in their respective master tables.\n"
            "   • Marital Status allowed values: Married, Single, Divorced, Widow.\n"
            "   • Gender allowed values: Male, Female, Other (or M, F, O).\n"
            "   • Email must follow valid format (e.g., name@domain.com).\n"
            "   • Person ID must be exactly 14 digits (numbers only, no scientific notation).\n"
            "   • Boolean fields (Iss ESS, Employee Status, Active, OT Applicable) accept values: True/False, Yes/No, 1/0.\n"
            "   • Trim unnecessary spaces — extra spaces are auto-normalized but discouraged.\n\n"
            "➡ SHEET 2: UDF (User Defined Fields)\n"
            "   • Used for uploading additional (custom) fields linked to employees.\n"
            "   • Only 'Employee Code' column is predefined; users can add extra columns for their custom fields.\n"
            "   • Employee Code must exist in EmployeeMaster.\n"
            "   • Date fields must be in DD-MM-YYYY or DD/MM/YYYY format.\n"
            "   • Dropdown, Radio, or Checkbox values must match allowed options in their setup.\n"
            "   • If a UDF field already exists for an employee, the system updates the value instead of duplicating.\n\n"
            "⚠️ COMMON VALIDATION NOTES:\n"
            "   • Ensure mandatory columns are filled; missing mandatory data will stop import.\n"
            "   • Ensure column headers are not renamed.\n"
            "   • Do not change sheet names.\n"
            "   • Do not delete or reorder system-generated columns.\n"
            "   • For optional fields, leave them blank (do not delete the column).\n\n"
                    )

        # Create bordered instruction box
        ws3.merge_cells("A2:A45")
        cell = ws3["A2"]
        cell.value = instruction_text
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        cell.font = Font(color="000000", bold=False)
        cell.fill = yellow_fill
        cell.border = border_style

        # Set readable layout
        ws3.column_dimensions["A"].width = 120  # wide enough for content
        for row in range(2, 46):
            ws3.row_dimensions[row].height = 25  # comfortable text height

        # ======================================================
        # Save to response
        # ======================================================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=\"Employee_BulkUpload_Template.xlsx\"'
        return response
    @action(detail=False, methods=['get'])
    def download_default_csv_file(self, request):
        # Use EmployeeResource column names
        resource = EmployeeResource()
        headers = [field.column_name for field in resource.fields.values()]

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)  # only headers, no data

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Employee_Sample_Template.csv"'
        return response
#EMP_FAMILY
class EmpFamViewSet(viewsets.ModelViewSet):
    queryset = emp_family.objects.all()  # Retrieve all instances of emp_family model
    serializer_class = EmpFamSerializer  # Use EmpFamSerializer for serialization
    # permission_classes = [IsAuthenticated]  # Require authentication for access

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:  # Check if user is authenticated
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return emp_family.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # If not a superuser or staff, filter based on emp_id
                return emp_family.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return emp_family.objects.filter(created_by=user)
        return emp_family.objects.none()  # Return an empty queryset if user is not authenticated or does not meet any condition

#EMP_JOB HISTORY
class EmpJobHistoryvSet(viewsets.ModelViewSet):
    queryset = EmpJobHistory.objects.all()
    serializer_class = EmpJobHistorySerializer
    # permission_classes = [IsAuthenticated]
    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return EmpJobHistory.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # Assuming 'emp_id' is the attribute that stores employee ID
                return EmpJobHistory.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return EmpJobHistory.objects.filter(created_by=user)
        return EmpJobHistory.objects.none()
    def get_serializer_context(self):
        return {'request': self.request}
    
#EMP_QUALIFICATION HISTORY
class Emp_QualificationViewSet(viewsets.ModelViewSet):
    queryset = EmpQualification.objects.all()
    serializer_class = Emp_qf_Serializer
    permission_classes = [EmpQualificationCustomFieldPermission]
    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return EmpQualification.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # Assuming 'emp_id' is the attribute that stores employee ID
                return EmpQualification.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return EmpQualification.objects.filter(created_by=user)
        return EmpQualification.objects.none()
    def get_serializer_context(self):
        return {'request': self.request}
    

    
class DocumentViewSet(viewsets.ModelViewSet):
    queryset = document_type.objects.all()
    serializer_class = Document_typeSerializer
    # permission_classes = [DocTypePermission,] 
    def get_queryset(self):
        """Return only active document types by default."""
        return document_type.objects.filter(is_active=True)
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.is_active = False  # Soft delete instead of actual deletion
        instance.save()
        return Response({"message": "Document type deactivated successfully"}, status=status.HTTP_204_NO_CONTENT)
    

#EMP_DOCUMENT 
class Emp_DocumentViewSet(viewsets.ModelViewSet):
    queryset = Emp_Documents.objects.all()
    serializer_class = DocumentSerializer
    # permission_classes = [IsAuthenticated]
    def get_serializer_context(self):
        return {'request': self.request}
    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated:
            if user.is_superuser or user.is_staff:  # Check if user is a superuser or staff
                return Emp_Documents.objects.all()  # Return all instances of emp_family
            elif hasattr(user, 'emp_id'):  # Assuming 'emp_id' is the attribute that stores employee ID
                return Emp_Documents.objects.filter(emp_id=user.emp_id)
            elif user.is_ess:  # If user is an ESS, filter based on created_by
                return Emp_Documents.objects.filter(created_by=user)
        return Emp_Documents.objects.none()
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Check if there are any existing documents with the same emp_id and document_type
        existing_documents = Emp_Documents.objects.filter(
            emp_id=serializer.validated_data['emp_id'],
            document_type=serializer.validated_data['document_type']
        )

        if existing_documents.exists():
            # Deactivate existing documents
            for doc in existing_documents:
                doc.is_active = False
                doc.save()

        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

class Doc_ReportViewset(viewsets.ModelViewSet):
    queryset = Doc_Report.objects.all()
    serializer_class = DocumentReportSerializer
    permission_classes = [DocReportPermission]
    def __init__(self, *args, **kwargs):
        super(Doc_ReportViewset, self).__init__(*args, **kwargs)
        self.ensure_standard_report_exists()

    def get_available_fields(self):
        # Define your available fields logic specific to documents
        excluded_fields = {'id', 'created_at', 'created_by', 'updated_at', 'updated_by', 'emp_sl_no', 'emp_doc_document'}
        included_emp_master_fields = {'emp_first_name', 'emp_active_date', 'emp_branch_id', 'emp_dept_id', 'emp_desgntn_id', 'emp_ctgry_id'}
        
        display_names = {
            "emp_id": "Employee ID",
            "emp_first_name": "First Name",
            "emp_active_date": "Active Date",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "emp_doc_type": "Document Type",
            "emp_doc_number": "Document Number",
            "emp_doc_issued_date": "Issued Date",
            "emp_doc_expiry_date": "Expiry Date",
            "is_active": "Active",
        }
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name in included_emp_master_fields]
        emp_document_fields = [field.name for field in Emp_Documents._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + emp_document_fields}
        return available_fields
    def ensure_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if Doc_Report.objects.filter(file_name='doc_std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()   
    def generate_standard_report(self):
        try:
            file_name = 'doc_std_report'
            fields_to_include = self.get_available_fields().keys()
            documents = Emp_Documents.objects.all()

            report_data = self.doc_report_data(fields_to_include, documents)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            Doc_Report.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )

            print("Standard report generated successfully.")

        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = Doc_Report.objects.get(file_name='doc_std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except Doc_Report.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['get'])
    def select_document_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return JsonResponse({'available_fields': available_fields})
        
    @action(detail=False, methods=['post'])
    def generate_document_report(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                file_name = request.data.get('file_name', 'report')
                fields_to_include = request.data.getlist('fields', [])
                # from_date = request.data.get('from_date')
                # to_date = request.data.get('to_date')
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})

            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())

            documents = Emp_Documents.objects.all()
            # documents = self.filter_documents_by_date_range(documents, from_date, to_date)

            report_data = self.doc_report_data(fields_to_include, documents)

            if not report_data:
                return JsonResponse({'status': 'error', 'message': 'No data to write into report'})

            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')
            try:
                with open(file_path, 'w') as file:
                    json.dump(report_data, file, default=str)
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Failed to write file: {str(e)}'})

            try:
                Doc_Report.objects.create(file_name=file_name, report_data=file_name + '.json')
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Failed to save report: {str(e)}'})

            return JsonResponse({'status': 'success', 'file_path': file_path, 'selected_fields_data': fields_to_include})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

    def doc_report_data(self, fields_to_include, documents):
        column_headings = {
            "emp_id": "Employee ID",
            "emp_first_name": "First Name",
            "emp_active_date": "Active Date",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "emp_doc_type": "Document Type",
            "emp_doc_number": "Document Number",
            "emp_doc_issued_date": "Issued Date",
            "emp_doc_expiry_date": "Expiry Date",
            "is_active": "Active",
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        emp_document_fields = [field.name for field in Emp_Documents._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        report_data = []
        for document in documents:
            document_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(document.emp_id, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()
                elif field in emp_document_fields:
                    value = getattr(document, field, 'N/A')
                else:
                    value = 'N/A'
                document_data[field] = value
            report_data.append(document_data)

        # print(f"Final report data: {report_data}")
        
        return report_data

    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])
        report_id = request.GET.get('report_id')
        
        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        }) 
    @action(detail=False, methods=['post'])
    def filter_by_date(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        # Replace slashes with hyphens
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')

        # Parse and validate the date range
        try:
            start_date = datetime.fromisoformat(start_date)
            end_date = datetime.fromisoformat(end_date)
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid date format: {str(e)}'}, status=400)

        # Fetch report data from your database
        try:
            report_instance = Doc_Report.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except Doc_Report.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)

        # Filter data by date range
        date_filtered_data = [
            row for row in report_data
            if 'emp_doc_expiry_date' in row and row['emp_doc_expiry_date'] and
            start_date <= datetime.fromisoformat(row['emp_doc_expiry_date']) <= end_date
        ]

        # Save filtered data to Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        cache.set(cache_key, date_filtered_data, timeout=None)  # Set timeout as needed

        return JsonResponse({
            'date_filtered_data': date_filtered_data,
            'report_id': report_id,
        })
      
    @action(detail=False, methods=['post'])
    def generate_doc_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.data.get('report_id')

        # Save selected fields to session
        request.session['selected_fields'] = selected_fields

        # Fetch date-filtered report data from session
        date_filtered_data = request.session.get('date_filtered_data', [])
        print("previosly date filtered ",date_filtered_data)
        
        # If no date-filtered data, attempt to fetch full report
        if not date_filtered_data:
            try:
                report = Doc_Report.objects.get(id=report_id)
                report_file_path = os.path.join(settings.MEDIA_ROOT, report.report_data.name)
                with open(report_file_path, 'r') as file:
                    report_content = json.load(file)
            except Doc_Report.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Report not found'})

            date_filtered_data = report_content

        # If no fields are selected for filtration, default to all existing fields in the report
        if not selected_fields:
            if date_filtered_data:
                selected_fields = list(date_filtered_data[0].keys())  # Default to all keys in the first record
            else:
                selected_fields = []  # No data in the report
        # Get unique values for selected_fields from date-filtered data
        unique_values = self.get_unique_values_for_fields(date_filtered_data, selected_fields)

        processed_unique_values = {}
        for field, values in unique_values.items():
            processed_unique_values[field] = {
                'values': values,
            }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': date_filtered_data,  # Pass filtered data to the frontend
            'unique_values': processed_unique_values,
        })

    def get_unique_values_for_fields(self, data, selected_fields):
        unique_values = {field: set() for field in selected_fields}
        # Extract data from the provided content
        for record in data:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        # Convert sets to lists
        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values
       
    @action(detail=False, methods=['post'])
    def filter_document_report(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')

        # Retrieve filtered data from Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        filtered_data = cache.get(cache_key)

        if filtered_data is None:
            return JsonResponse({'status': 'error', 'message': 'No date-filtered data available'}, status=404)

        # Apply additional filtering here if needed
        # For example, based on other fields:
        additional_filters = {key: value for key, value in request.data.items() if key not in ('report_id',)}
        
        # Further filter based on additional criteria
        filtered_data = [
            row for row in filtered_data
            if all(row.get(key) == value for key, value in additional_filters.items())
        ]

        return JsonResponse({
            'filtered_data': filtered_data,
            'report_id': report_id,
        })

    # def match_filter_criteria(self, row_data, filter_criteria):
    #     for field, values in filter_criteria.items():
    #         row_value = row_data.get(field, '').strip() if row_data.get(field) else ''
    #         print(f"Checking field {field} with values {values} against row value {row_value}")  # Debugging statement
    #         if row_value not in values:
    #             return False
    #     return True 
    
    @action(detail=False, methods=['get'])
    def generate_excel_report(self, request, *args, **kwargs):
        report_id = request.GET.get('report_id')
        if not report_id:
            return HttpResponse('Report ID is missing', status=400)

        filtered_data = request.session.get('filtered_data')
        if not filtered_data:
            return HttpResponse('No filtered data available', status=400)

        # Mapping of internal field names to display names 
        field_names_mapping = {
            "emp_id": "Employee Code",
            "emp_first_name": "First Name",
            "emp_active_date": "Active Date",
            "emp_branch_id": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "emp_doc_type": "Document Type",
            "emp_doc_number": "Document Number",
            "emp_doc_issued_date": "Issued Date",
            "emp_doc_expiry_date": "Expiry Date",
            "is_active": "Active",
        }

        try:
            report_instance = Doc_Report.objects.get(id=int(report_id))
        except (Doc_Report.DoesNotExist, ValueError):
            return HttpResponse('Invalid or missing Report ID', status=404)

        # Create an Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Filtered Report'
        
        # Define style for header row
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFF")
        header_style.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

        # Add header row to Excel using display names and apply style
        if filtered_data:
            headers = [field_names_mapping.get(field_name, field_name) for field_name in filtered_data[0].keys()]
            sheet.append(headers)
            for cell in sheet[1]:
                cell.style = header_style

        # Add data rows to Excel using values from filtered_data
        for row in filtered_data:
            row_values = [row.get(field_name, '') for field_name in filtered_data[0].keys()]
            sheet.append(row_values)

        # Autofit column widths
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Save the workbook to a BytesIO stream
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)

        # Prepare the response with Excel file as attachment
        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=filtered_report_{report_id}.xlsx'
        return response
     
class Bulkupload_DocumentViewSet(viewsets.ModelViewSet):
    queryset = Emp_Documents.objects.all()
    serializer_class = DocBulkuploadSerializer
    parser_classes = (MultiPartParser, FormParser)
    
    @action(detail=False, methods=['get'])
    def download_demo_excel(self, request):
        # Only column headers
        columns = ["Employee Code", "Document Type", "Document Number", "Document Issued Date","Document Expiry Date","Active"]

        # Empty DataFrame (only headers, no rows)
        df = pd.DataFrame(columns=columns)

        # Save to Excel in memory
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Employee Documents")

        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="sample_emp_documents_sheet.xlsx"'
        return response

    @action(detail=False, methods=['get'])
    def download_demo_csv(self, request):
        # Only column headers
        columns = ["Employee Code", "Document Type", "Document Number", "Document Issued Date","Document Expiry Date","Active"]

        df = pd.DataFrame(columns=columns)

        buffer = io.StringIO()
        df.to_csv(buffer, index=False)

        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="sample_emp_documents_sheet.csv"'
        return response
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if 'file' not in request.FILES:
            return Response({"error": "Please provide a file."}, status=400)

        upload_file = request.FILES['file']
        filename = upload_file.name.lower()
        all_errors = {"sheet1_errors": [], "sheet2_errors": []}

        try:
            dataset_sheet1 = Dataset()
            dataset_sheet2 = Dataset()
            dataset_sheet1.headers = []
            dataset_sheet2.headers = ['Document Number', 'Field Name', 'Field Value']

            # ---------------- XLS/XLSX ----------------
            if filename.endswith(('.xlsx', '.xls')):
                workbook = load_workbook(upload_file, data_only=True)

                # Sheet 1: DocumentMaster
                if "DocumentMaster" not in workbook.sheetnames:
                    return Response({"error": "Sheet1 (DocumentMaster) missing"}, status=400)
                sheet1 = workbook["DocumentMaster"]
                dataset_sheet1.headers = [cell.value for cell in sheet1[1]]
                for row in sheet1.iter_rows(min_row=2):
                    dataset_sheet1.append([cell.value for cell in row])

                # Sheet 2: UDF
                if "Doc-UDF" in workbook.sheetnames:
                    sheet2 = workbook["Doc-UDF"]
                    if sheet2.max_row > 1:
                        headers = [cell.value for cell in sheet2[1]]

                        # Wide format parsing: all columns after Employee Code are UDF
                        emp_code_index = headers.index("Document Number")
                        for row in sheet2.iter_rows(min_row=2):
                            emp_code = str(row[emp_code_index].value or "")
                            for col_idx, field_name in enumerate(headers):
                                if field_name == "Document Number":
                                    continue
                                field_value = row[col_idx].value
                                if field_value not in (None, ''):
                                    dataset_sheet2.append([emp_code, field_name, str(field_value)])

            # ---------------- CSV ----------------
            elif filename.endswith('.csv'):
                file_data = upload_file.read().decode("utf-8")
                csv_reader = csv.DictReader(io.StringIO(file_data))
                headers = csv_reader.fieldnames

                employee_columns = [f.column_name for f in DocumentResource().fields.values()]
                dataset_sheet1.headers = employee_columns

                for row in csv_reader:
                    emp_row = [row.get(col, '') for col in employee_columns]
                    dataset_sheet1.append(emp_row)

                    emp_code = row.get('Document Number', '')

                    # Wide format UDF columns
                    udf_columns = [h for h in headers if h not in employee_columns]
                    for udf_col in udf_columns:
                        field_value = row.get(udf_col, '')
                        if field_value not in (None, ''):
                            dataset_sheet2.append([emp_code, udf_col, str(field_value)])

            else:
                return Response({"error": "Invalid file format. Only .xlsx, .xls, .csv supported."}, status=400)

            # ---------------- Validation ----------------
            employee_resource = DocumentResource()
            custom_field_resource = EmpDocumentCustomFieldValueResource()

            # DocumentMaster validation
            for row_idx, row in enumerate(dataset_sheet1.dict, start=2):
                try:
                    employee_resource.before_import_row(row, row_idx=row_idx)
                except ValidationError as e:
                    all_errors["sheet1_errors"].append({"row": row_idx, "error": str(e)})

            # UDF validation
            for row_idx, row in enumerate(dataset_sheet2.dict, start=2):
                try:
                    field_name = (row.get('Field Name') or '').strip()
                    field_value = row.get('Field Value', '')
                    emp_code = row.get('Document Number', '')

                    if not field_name:
                        raise ValidationError("Field Name cannot be empty")
                    if not emp_code:
                        raise ValidationError("Employee Code cannot be empty")

                    # Check if custom field exists
                    custom_field = EmpDocuments_CustomField.objects.filter(emp_custom_field=field_name).first()
                    if not custom_field:
                        raise ValidationError(f"Custom field '{field_name}' does not exist.")

                    # Handle date fields
                    if custom_field.data_type == 'date' and field_value:
                        if isinstance(field_value, (datetime, date)):
                            field_value = field_value.strftime('%d-%m-%Y')
                        elif isinstance(field_value, str):
                            field_value = field_value.strip()
                            if ' ' in field_value:
                                field_value = field_value.split(' ')[0]
                            for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
                                try:
                                    parsed_date = datetime.strptime(field_value, fmt)
                                    field_value = parsed_date.strftime('%d-%m-%Y')
                                    break
                                except ValueError:
                                    continue
                            else:
                                raise ValidationError(
                                    f"Invalid date format for field '{field_name}'. Expected DD-MM-YYYY or DD/MM/YYYY."
                                )

                    # Update row with normalized value
                    row['Field Value'] = field_value

                except ValidationError as e:
                    all_errors["sheet2_errors"].append({"row": row_idx, "error": str(e)})

            if all_errors["sheet1_errors"] or all_errors["sheet2_errors"]:
                return Response({"errors": all_errors}, status=400)

            # ---------------- Import ----------------
            with transaction.atomic():
                employee_result = employee_resource.import_data(dataset_sheet1, dry_run=False, raise_errors=True)

            if dataset_sheet2:
                with transaction.atomic():
                    custom_field_result = custom_field_resource.import_data(dataset_sheet2, dry_run=False, raise_errors=True)
                return Response({
                    "message": f"{employee_result.total_rows} Employee Documents records created, "
                               f"{custom_field_result.total_rows} UDF records created successfully"
                })

            return Response({"message": f"{employee_result.total_rows} Employee Documents records created successfully."})

        except Exception as e:
            return Response({"error": str(e)}, status=400)

class EmplistViewSet(BranchAccessMixin,viewsets.ReadOnlyModelViewSet):  # Read-only API (GET only)
    queryset = emp_master.objects.filter(is_active=True)
    serializer_class = EmplistSerializer
    permission_classes = [EmployeePermission]
    
# EmpLeaveRequest
class EmpLeaveRequestViewSet(viewsets.ModelViewSet):
    queryset = EmpLeaveRequest.objects.all()
    serializer_class = EmpLeaveRequestSerializer
    # permission_classes = [IsAuthenticated]
    def get_serializer_context(self):
        return {'request': self.request}
    

class NotificationViewset(viewsets.ModelViewSet):
    queryset = notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [NotificationPermission]
    def get_queryset(self):
        user = self.request.user

        # Admin / staff / superuser → see all notifications
        if user.is_superuser or user.is_staff:
            return notification.objects.all().order_by('-created_at')

        # Normal employee (ESS User) → only notifications related to their documents
        return notification.objects.filter(
            document_id__emp_id__users=user
        ).order_by('-created_at')

class EmpMarketSkillViewSet(viewsets.ModelViewSet):
    queryset = EmployeeMarketingSkill.objects.all()
    serializer_class = EmpMarketSkillSerializer 
    permission_classes = [EmployeeMarketingSkillPermission]
 
class EmpPrgrmSkillViewSet(viewsets.ModelViewSet):
    queryset = EmployeeProgramSkill.objects.all()
    serializer_class = EmpPrgrmSkillSerializer
    permission_classes = [EmployeeProgramSkillPermission]

    
class EmpLangSkillViewSet(viewsets.ModelViewSet):
    queryset = EmployeeLangSkill.objects.all()
    serializer_class = EmpLangSkillSerializer  
    permission_classes = [EmployeeLangSkillPermission]


class RequestTypeViewset(viewsets.ModelViewSet):
    queryset = RequestType.objects.all()
    serializer_class = RequestTypeSerializer
    permission_classes = [RequestTypePermission]

    def perform_create(self, serializer):
        request_type = serializer.save()

        workflow = ApprovalWorkflow.objects.create(
            request_type=request_type,
            approval_type="no_approval"
        )

        workflow.branch.set(request_type.branch.all())

        ApprovalLevel.objects.create(
            workflow=workflow,
            level=1,
            role="Auto Level",
            approver=None
        )

class EmailTemplateViewset(viewsets.ModelViewSet):
    queryset = EmailTemplate.objects.all()
    serializer_class = EmailTemplateSerializer
    # Custom action to get the placeholders dynamically
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            'request': [
                '{{ doc_number }}',
                '{{ request_type }}',
                '{{ reason }}',
                # Add other request-related placeholders here
            ],
            'employee': [
                '{{ doc_number }}',
                '{{ request_type }}',
                '{{ reason }}',
                '{{ recipient_name }}',
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_gender }}',
                '{{ emp_date_of_birth }}',
                '{{ emp_personal_email }}',
                '{{ emp_company_email }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}'
            ]
        }
        return Response(placeholders)

    # Custom action to fetch the available From and To addresses
    @action(detail=False, methods=['get'], url_path='from-to-addresses')
    def from_to_list(self, request):
        # Fetch active email configurations for "From" addresses
        from_addresses = EmailConfiguration.objects.filter(is_active=True).values_list('email_host_user', flat=True)

        # Fetch employee emails for "To" addresses
        to_addresses = emp_master.objects.all().values_list('emp_personal_email', 'emp_company_email')

        to_list = []
        for emp_personal, emp_company in to_addresses:
            if emp_personal:
                to_list.append(emp_personal)
            if emp_company:
                to_list.append(emp_company)

        return Response({
            'from_addresses': from_addresses,
            'to_addresses': to_list
        })

from django.db import transaction
class GeneralRequestViewset(viewsets.ModelViewSet):
    queryset = GeneralRequest.objects.all()
    serializer_class = GeneralRequestSerializer
    permission_classes =[IsSuperUserOrHasGeneralRequestPermission]
    def perform_create(self, serializer):
        with transaction.atomic():
            employee = serializer.validated_data.get('employee')
            document_number = serializer.validated_data.get('document_number')

            # ✅ Check employee
            if not employee:
                raise ValidationError("Employee is required.")

            # ✅ FIX: fallback to work_location if emp_branch_id missing
            branch_id = employee.emp_branch_id or employee.work_location

            if not branch_id:
                raise ValidationError("Employee branch is missing in employee master.")

            try:
                doc_config = DocumentNumbering.objects.get(
                    branch_id=branch_id,
                    type='general_request',
                )
            except DocumentNumbering.DoesNotExist:
                raise NotFound(
                    f"No document numbering configuration found for branch {branch_id} and general request."
                )

            current_date = timezone.now().date()

            # ✅ Manual document number validation
            if document_number:
                if doc_config.start_date and doc_config.end_date:
                    if not (doc_config.start_date <= current_date <= doc_config.end_date):
                        raise ValidationError(
                            "Document number cannot be assigned outside the valid date range."
                        )
            else:
                # ✅ Auto-generate
                document_number = doc_config.get_next_number()

            serializer.save(document_number=document_number)

    @action(detail=False, methods=['get'])
    def employee_request_history(self, request):
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({'error': 'Employee ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        requests = GeneralRequest.get_employee_requests(employee_id)

        # Manually serialize the fields you want
        history_data = []
        for request in requests:
            history_data.append({
                'doc_number': request.doc_number,
                'reason': request.reason,
                'branch': request.branch.branch_name if request.branch else None,
                'request_type': request.request_type.name if request.request_type else None,
                'status': request.status,
                'created_at_date': request.created_at_date,
            })

        return Response(history_data, status=status.HTTP_200_OK)
    

    # @action(detail=False, methods=['get'])
    # def employee_request_history(self, request):
    #     employee_id = request.query_params.get('employee_id')
    #     if not employee_id:
    #         return Response({'error': 'Employee ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        
    #     requests = GeneralRequest.get_employee_requests(employee_id)

    #     # Manually serialize the fields you want
    #     history_data = []
    #     for request in requests:
    #         history_data.append({
    #             'doc_number': request.doc_number,
    #             'reason': request.reason,
    #             'branch': request.branch.branch_name if request.branch else None,
    #             'request_type': request.request_type.name if request.request_type else None,
    #             'status': request.status,
    #             'created_at_date': request.created_at_date,
    #         })

    #     return Response(history_data, status=status.HTTP_200_OK)

class ApprovalLevelViewset(viewsets.ModelViewSet):
    queryset = ApprovalWorkflow.objects.all()
    serializer_class = ApprovalWorkflowSerializer
    permission_classes = [ApprovalLevelPermission]

class CommonWorkflowViewSet(viewsets.ModelViewSet):
        queryset = CommonWorkflow.objects.all()
        serializer_class  = CommonWorkflowSerializer   

class ApprovalViewset(viewsets.ModelViewSet):
    queryset = Approval.objects.all()
    serializer_class = ApprovalSerializer
    lookup_field = 'pk'

    def get_queryset(self):
        user = self.request.user

        if not user.is_authenticated:
            return Approval.objects.none()

        if user.is_superuser:
            return Approval.objects.all()

        return Approval.objects.filter(
        Q(approver=user) |
        Q(deligate_to=user, is_deligate=True)
    ).distinct()

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approval = self.get_object()

        note = request.data.get("note")

        approval.approve(note=note)

        return Response(
            {
                "message": "Request approved successfully.",
                "status": approval.status,
                "note": note,
            },
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()

        note = request.data.get("note")

        approval.reject(note=note)

        return Response(
            {
                "message": "Request rejected successfully.",
                "status": approval.status,
                "note": note,
            },
            status=status.HTTP_200_OK
        )

    
    @action(detail=True, methods=["post"])
    def delegate(self, request, pk=None):
        approval = self.get_object()

        delegate_user_id = request.data.get("deligate_to")

        if not delegate_user_id:
            return Response(
                {"error": "Delegate user is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        delegate_user = get_object_or_404(CustomUser, pk=delegate_user_id)

        if delegate_user == approval.approver:
            return Response(
                {"error": "You cannot delegate to yourself."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if approval.is_deligate:
            return Response(
                {"error": "This approval has already been delegated."},
                status=status.HTTP_400_BAD_REQUEST
            )

        approval.deligate_to = delegate_user
        approval.is_deligate = True
        approval.deligate_response = None
        approval.save()

        if delegate_user.email:

            subject = "Delegation Assigned"

            message = f"""
                Delegation Assigned

                Hello {delegate_user.get_username() or delegate_user.username},

                You have been assigned a new delegation request.

                DELEGATION DETAILS
                ___________________

                Original Approver : {approval.approver.username}
                Delegate To       : {delegate_user.username}
                Delegated At      : {approval.updated_at}

                REQUEST DETAILS
                _________________

                Document Number : {approval.general_request.document_number}
                Employee        : {approval.general_request.employee}
                Request Type    : {approval.general_request.request_type}
                Status          : {approval.general_request.status}

                Please review the request and send your response to the original approver.

                Thank You.
                """

            send_mail(
                subject,
                message,
                None,
                [delegate_user.email],
                fail_silently=False,
            )

        created_notification = send_notification_email(
            user=delegate_user,
            employee=None,
            branch=None,
            title="Delegation Assigned",
            message=f"{approval.approver.username} has delegated request {approval.general_request.document_number} to you.",
            template_type="request_created",
            delegate_user=approval.approver,
        )

        print("Notification Created:", created_notification)

        return Response(
            {
                "message": "Approval delegated successfully.",
                "approval_id": approval.id,
                "approver": approval.approver.username,
                "delegate_to": delegate_user.username,
                "status": approval.status,
            },
            status=status.HTTP_200_OK,
        )
    @action(detail=True, methods=["post"])
    def send_response(self, request, pk=None):
        approval = self.get_object()

        response_text = request.data.get("deligate_response")

        if not response_text:
            return Response({"error": "Response is required"}, status=400)

        approval.deligate_response = response_text
        approval.save()

        # ---------------- EMAIL ----------------
        if approval.approver and approval.approver.email:
            send_mail(
                subject="Delegation Response Received",
                message=response_text,
                from_email=None,
                recipient_list=[approval.approver.email],
                fail_silently=False,
            )

        # ---------------- NOTIFICATION ----------------
        send_notification_email(
            user=approval.approver,
            employee=None,
            branch=None,
            title="Delegation Response Received",
            message=response_text,
            template_type="request_created",
            delegate_user=approval.deligate_to,
        )

        return Response({
            "message": "Response sent successfully",
            "response": response_text
        })
    
class UserNotificationsViewSet(viewsets.ModelViewSet):
    queryset = RequestNotification.objects.all()
    serializer_class = ReqNotifySerializer
    # permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        # Admin / staff / superuser → see all request notifications
        if user.is_superuser or user.is_staff:
            return RequestNotification.objects.all().order_by('-created_at')
        # Normal user → show request notifications assigned directly to them
        qs = RequestNotification.objects.filter(
            Q(recipient_user=user.id, is_deligate=False) |
            Q(recipient_employee__users=user, is_deligate=False) |
            Q(deligate_user=user.id, is_deligate=True)
        ).distinct().order_by('-created_at')

        return qs
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse


class EmailConfigurationViewSet(viewsets.ModelViewSet):
    queryset = EmailConfiguration.objects.all()
    serializer_class = EmailConfigurationSerializer
    lookup_field = 'pk'
    
    def perform_create(self, serializer):
        instance = serializer.save()
        if instance.is_active:
            # Deactivate other configurations
            EmailConfiguration.objects.filter(is_active=True).exclude(pk=instance.pk).update(is_active=False)


    
class GeneralReportViewset(viewsets.ModelViewSet):
    queryset = GeneralRequestReport.objects.all()
    serializer_class = GeneralReportSerializer
    # permission_classes = [GeneralRequestReportPermission]
    
    def __init__(self, *args, **kwargs):
        super(GeneralReportViewset, self).__init__(*args, **kwargs)
        self.general_standard_report_exists()

    def get_available_fields(self):
        excluded_fields = {'id', 'created_by'}
        included_emp_master_fields = { 'emp_first_name', 'emp_dept_id', 'emp_desgntn_id', 'emp_ctgry_id','emp_branch_id'}
        
        display_names = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "emp_active_date": "Active Date",
            "emp_branch_id":"Branches",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "doc_number": "Document Number",
            "reason": "Reason",
            "total":"Total",
            "request_type": "Request Type",
            "approved": "Approved Request",
            "created_at_date":"Request Date",
           
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name in included_emp_master_fields]
        general_request_fields = [field.name for field in GeneralRequest._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + general_request_fields}
        return available_fields
    
    @action(detail=False, methods=['get'])
    def select_generalreport_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
       
    @action(detail=False, methods=['post'])
    def generate_general_report(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'report')
                fields_to_include = request.POST.getlist('fields', [])
                # from_date = parse_date(request.POST.get('from_date'))
                # to_date = parse_date(request.POST.get('to_date'))
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
            
            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())
            
            generalreport = GeneralRequest.objects.all()
            # documents = self.filter_documents_by_date_range(documents)

            report_data = self.generate_report_data(fields_to_include,generalreport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format


            GeneralRequestReport.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({'status': 'success', 'file_path': file_path,'selected_fields_data': fields_to_include,})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

   
    def general_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if GeneralRequestReport.objects.filter(file_name='generalrequest_std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'generalrequest_std_report'
            fields_to_include = self.get_available_fields().keys()
            generalreport = GeneralRequest.objects.all()

            report_data = self.generate_report_data(fields_to_include, generalreport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            GeneralRequestReport.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )
        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = GeneralRequestReport.objects.get(file_name='generalrequest_std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except GeneralRequestReport.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    def generate_report_data(self, fields_to_include,generalreport):
        column_headings = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "branch": "Branch",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "doc_number": "Document Number",
            "reason": "Reason",
            "total":"Total",
            "request_type": "Request Type",
            "approved": "Approved Request",
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        general_request_fields = [field.name for field in GeneralRequest._meta.get_fields() if isinstance(field, Field) and field.name != 'id']

        report_data = []
        for document in generalreport:
            general_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(document.employee, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()
                elif field in general_request_fields:
                    value = getattr(document, field, 'N/A')
                else:
                    value = 'N/A'
                general_data[field] = value
            report_data.append(general_data)
        return report_data            

    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])
        report_id = request.GET.get('report_id')  # Get report_id from query parameters

        
        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        }) 
      
    @action(detail=False, methods=['post'])
    def filter_by_date(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        # Replace slashes with hyphens
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')

        # Parse and validate the date range
        try:
            start_date = datetime.fromisoformat(start_date)
            end_date = datetime.fromisoformat(end_date)
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid date format: {str(e)}'}, status=400)

        # Fetch report data from your database
        try:
            report_instance = GeneralRequestReport.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except GeneralRequestReport.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)

        # Filter data by date range
        date_filtered_data = [
            row for row in report_data
            if 'created_at_date' in row and row['created_at_date'] and
            start_date <= datetime.fromisoformat(row['created_at_date']) <= end_date
        ]

        # Save filtered data to Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        cache.set(cache_key, date_filtered_data, timeout=None)  # Set timeout as needed

        return JsonResponse({
            'date_filtered_data': date_filtered_data,
            'report_id': report_id,
        })
    
    @action(detail=False, methods=['post'])
    def generate_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.data.get('report_id')

        # Fetch previously filtered date data from the `apply_date_filter` method
        date_filtered_data = getattr(self, 'date_filtered_data', [])
        print("previous",date_filtered_data)

        # If no date-filtered data, attempt to fetch full report
        if not date_filtered_data:
            report_instance = get_object_or_404(GeneralRequestReport, id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
            date_filtered_data = report_data

        # Default to all fields if no specific fields selected
        if not selected_fields and date_filtered_data:
            selected_fields = list(date_filtered_data[0].keys())

        # Get unique values for selected_fields from date-filtered data
        unique_values = self.get_unique_values_for_fields(date_filtered_data, selected_fields)

        processed_unique_values = {
            field: {'values': values}
            for field, values in unique_values.items()
        }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': date_filtered_data,
            'unique_values': processed_unique_values,
            # 'column_headings': column_headings
        })

    def get_unique_values_for_fields(self, data, selected_fields):
        unique_values = {field: set() for field in selected_fields}
        for record in data:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values
    @action(detail=False, methods=['post'])
    def general_filter_report(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')

        # Retrieve filtered data from Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        filtered_data = cache.get(cache_key)

        if filtered_data is None:
            return JsonResponse({'status': 'error', 'message': 'No date-filtered data available'}, status=404)
        # Apply additional filtering here if needed
        # For example, based on other fields:
        additional_filters = {key: value for key, value in request.data.items() if key not in ('report_id',)}
        
        # Further filter based on additional criteria
        filtered_data = [
            row for row in filtered_data
            if all(row.get(key) == value for key, value in additional_filters.items())
        ]
        return JsonResponse({
            'filtered_data': filtered_data,
            'report_id': report_id,
        })

class UpdateESSUserView(APIView):
    def post(self, request, *args, **kwargs):
        selected_employee_ids = request.data.get('selected_employee_ids', [])

        if not selected_employee_ids:
            return Response({'detail': 'At least one selected employee ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

        selected_employees = emp_master.objects.filter(id__in=selected_employee_ids)

        if not selected_employees.exists():
            return Response({'detail': 'No valid employees found.'}, status=status.HTTP_404_NOT_FOUND)

        # Update the SelectedEmpNotify record with the selected employees
        preference, created = SelectedEmpNotify.objects.get_or_create(id=1)
        preference.selected_employees.set(selected_employees)
        preference.save()

        return Response({'detail': 'Selected employees updated successfully.'}, status=status.HTTP_200_OK)


class ESSUserListView(APIView):
    def get(self, request, *args, **kwargs):
        ess_users = emp_master.objects.filter(is_ess=True)
        serializer = EmpSerializer(ess_users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class NotificationSettingsViewSet(BranchAccessMixin,viewsets.ModelViewSet):
    queryset = NotificationSettings.objects.all()
    serializer_class =  NotificationSettingsSerializer

class DocExpEmailTemplateViewset(viewsets.ModelViewSet):
    queryset = DocExpEmailTemplate.objects.all()
    serializer_class = DocExpEmailTemplateSerializer
    
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            'employee': [
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}',
                '{{ document_type }}',
                '{{ expiry_date }}',
                '{{ is_active }}',
                '{{ document_number }}',
                '{{ doc_issued_date }}',
                '{{ doc_expiry_date }}',
                '{{ emp_date_of_birth }}',
            ]
        }
        return Response(placeholders)

class EmployeeBankDetailViewset(viewsets.ModelViewSet):
    queryset = EmployeeBankDetail.objects.all()
    serializer_class = EmpBankDetailsSerializer


class EmpBankBulkuploadViewSet(viewsets.ModelViewSet):
    queryset = EmployeeBankDetail.objects.all()
    serializer_class = EmpBankBulkuploadSerializer
    parser_classes = (MultiPartParser, FormParser)
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method == 'POST' and request.FILES.get('file'):
            excel_file = request.FILES['file']
            if excel_file.name.endswith('.xlsx'):
                try:
                    dataset = Dataset()
                    dataset.load(excel_file.read(), format='xlsx')
                    resource = EmpBankDetailsResource()
                    all_errors = []
                    valid_rows = []
                    with transaction.atomic():
                        for row_idx, row in enumerate(dataset.dict, start=2):
                            row_errors = []
                            try:
                                resource.before_import_row(row, row_idx=row_idx)
                            except ValidationError as e:
                                row_errors.extend([f"Row {row_idx}: {error}" for error in e.messages])
                            if row_errors:
                                all_errors.extend(row_errors)
                            else:
                                valid_rows.append(row)

                    if all_errors:
                        return Response({"errors": all_errors}, status=400)

                    with transaction.atomic():
                        result = resource.import_data(dataset, dry_run=False, raise_errors=True)

                    return Response({"message": f"{result.total_rows} records created successfully"})
                except Exception as e:
                    return Response({"error": str(e)}, status=400)
            else:
                return Response({"error": "Invalid file format. Only Excel files (.xlsx) are supported."}, status=400)
        else:
            return Response({"error": "Please provide an Excel file."}, status=400)
class DocRequestTypeViewset(viewsets.ModelViewSet):
    queryset = DocRequestType.objects.all()
    serializer_class = DocRequestTypeSerializer
    def perform_create(self, serializer):
        request_type = serializer.save()

        workflow =  DocumentApprovalWorkflow.objects.create(
            request_type=request_type,
            approval_type="no_approval"
        )

        workflow.branch.set(request_type.branch.all())

        DocumentApprovalLevel.objects.create(
            workflow=workflow,
            level=1,
            role="Auto Level",
            approver=None
        )

class DocumentRequestViewset(viewsets.ModelViewSet):
    queryset = DocumentRequest.objects.all()
    serializer_class = DocRequestSerializer
    def perform_create(self, serializer):
        with transaction.atomic():

            employee = serializer.validated_data.get('employee')
            document_number = serializer.validated_data.get('document_number')

            # ✅ Employee validation
            if not employee:
                raise ValidationError("Employee is required.")

            # ✅ Branch fallback
            branch = employee.emp_branch_id or employee.work_location

            if not branch:
                raise ValidationError(
                    "Employee branch is missing in employee master."
                )

            # ✅ Get document numbering configuration
            try:
                doc_config = DocumentNumbering.objects.get(
                    branch_id=branch.id,
                    type='document_request'
                )

            except DocumentNumbering.DoesNotExist:
                raise NotFound(
                    f"No document numbering configuration found "
                    f"for branch {branch} and document request."
                )

            current_date = timezone.now().date()

            # ✅ Manual document number validation
            if document_number:

                if (
                    doc_config.start_date
                    and doc_config.end_date
                    and not (
                        doc_config.start_date
                        <= current_date
                        <= doc_config.end_date
                    )
                ):
                    raise ValidationError(
                        "Document number cannot be assigned "
                        "outside the valid date range."
                    )

            else:
                # ✅ Auto-generate document number
                document_number = doc_config.get_next_number()

            # ✅ Save document request
            serializer.save(
                document_number=document_number,
                branch=branch,
                created_by=self.request.user
            )
class DocumentApprovalViewset(viewsets.ModelViewSet):
    queryset = DocumentApproval.objects.all()
    serializer_class = DocApprovalSerializer

    def get_queryset(self):
        user = self.request.user

        if not user.is_authenticated:
            return DocumentApproval.objects.none()

        if user.is_superuser:
            return DocumentApproval.objects.all()

        return DocumentApproval.objects.filter(
        Q(approver=user) |
        Q(deligate_to=user, is_deligate=True)
    ).distinct()


    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approval = self.get_object()
        # if request.user != approval.approver:
        #     return Response({'error': 'You are not authorized to approve this request.'}, status=status.HTTP_403_FORBIDDEN)

        note = request.data.get('note')  # Get the note from the request
        approval.approve(note=note)
        return Response({'status': 'approved', 'note': note}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')  # Get the note from the request
        approval.reject(note=note)
        return Response({'status': 'rejected', 'note': note}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=["post"])
    def delegate(self, request, pk=None):
        approval = self.get_object()

        delegate_user_id = request.data.get("deligate_to")

        if not delegate_user_id:
            return Response(
                {"error": "Delegate user is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        delegate_user = get_object_or_404(CustomUser, pk=delegate_user_id)

        if delegate_user == approval.approver:
            return Response(
                {"error": "You cannot delegate to yourself."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if approval.is_deligate:
            return Response(
                {"error": "This approval has already been delegated."},
                status=status.HTTP_400_BAD_REQUEST
            )

        approval.deligate_to = delegate_user
        approval.is_deligate = True
        approval.deligate_response = None
        approval.save()

        if delegate_user.email:

            subject = "Delegation Assigned"

            message = f"""
                Delegation Assigned

                Hello {delegate_user.get_username() or delegate_user.username},

                You have been assigned a new delegation request.

                REQUEST DETAILS
                _________________

                Document Number : {approval.document_request.document_number}
                Employee        : {approval.document_request.employee}
                Request Type    : {approval.document_request.request_type}
                Status          : {approval.document_request.status}

                Please review the request and send your response to the original approver.

                Thank You.
                """

            send_mail(
                subject,
                message,
                None,
                [delegate_user.email],
                fail_silently=False,
            )
            created_notification = send_notification_email(
                user=delegate_user,
                employee=None,
                branch=None,
                title="Delegation Assigned",
                message=f"{approval.approver.username} has delegated request {approval.document_request.document_number} to you.",
                delegate_user=approval.approver,
                template_type="request_created",
                context={
                    **get_employee_context(approval.document_request.employee),
                        "doc_number": approval.document_request.document_number,
                        "request_type": approval.document_request.request_type.type_name,
                    },
                    email_template_model=DocRequestEmailTemplate,
                    notification_model=DocRequestNotification,
            )

            print("Notification Created:", created_notification)

            return Response(
                {
                    "message": "Approval delegated successfully.",
                    "approval_id": approval.id,
                    "approver": approval.approver.username,
                    "delegate_to": delegate_user.username,
                    "status": approval.status,
                },
                status=status.HTTP_200_OK,
            )
    @action(detail=True, methods=["post"])
    def send_response(self, request, pk=None):
        approval = self.get_object()

        response_text = request.data.get("deligate_response")

        if not response_text:
            return Response({"error": "Response is required"}, status=400)

        approval.deligate_response = response_text
        approval.save()

        # ---------------- EMAIL ----------------
        if approval.approver and approval.approver.email:
            send_mail(
                subject="Delegation Response Received",
                message=response_text,
                from_email=None,
                recipient_list=[approval.approver.email],
                fail_silently=False,
            )

        # ---------------- NOTIFICATION ----------------
        send_notification_email(
            user=approval.approver,
            employee=None,
            branch=None,
            title="Delegation Response Received",
            message=response_text,
            delegate_user=approval.deligate_to,
            template_type="request_created",
            context={
                    **get_employee_context(approval.document_request.employee),
                    "doc_number": approval.document_request.document_number,
                    "request_type": approval.document_request.request_type.type_name,
                },
                email_template_model=DocRequestEmailTemplate,
                notification_model=DocRequestNotification,
        )

        return Response({
            "message": "Response sent successfully",
            "response": response_text
        })


class DocumentApprovalLevelViewset(viewsets.ModelViewSet):
    queryset =DocumentApprovalWorkflow.objects.all()
    serializer_class = DocumentApprovalWorkflowSerializer
class DocRequestEmailTemplateViewset(viewsets.ModelViewSet):
    queryset = DocRequestEmailTemplate.objects.all()
    serializer_class = DocRequestEmailTemplateSerializer
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            
            'employee': [
                '{{ doc_number }}',
                '{{ request_type }}',
                '{{ reason }}',
                '{{ recipient_name }}',
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_gender }}',
                '{{ emp_date_of_birth }}',
                '{{ emp_personal_email }}',
                '{{ emp_company_email }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}'
            ]
        }
        return Response(placeholders)
class DocRequestNotificationViewset(viewsets.ModelViewSet):
    queryset = DocRequestNotification.objects.all()
    serializer_class = DocRequestNotificationSerializer
    def get_queryset(self):
        user = self.request.user

        # Admin / staff / superuser → see all request notifications
        if user.is_superuser or user.is_staff:
            return DocRequestNotification.objects.all().order_by('-created_at')

        # Normal user → show request notifications assigned directly to them
        qs = DocRequestNotification.objects.filter(
            Q(recipient_user=user) |
            Q(recipient_employee__users=user)      # employee assigned to this user
        ).order_by('-created_at')

        return qs
class EmployeeResignationViewset(viewsets.ModelViewSet):
    queryset = EmployeeResignation.objects.all()
    serializer_class = EmployeeResignationSerializer
    # permission_classes = [EmployeeResignationPermission]

    def perform_create(self, serializer):

        with transaction.atomic():

            employee = serializer.validated_data.get('employee')
            document_number = serializer.validated_data.get('document_number')

            # ✅ Employee validation
            if not employee:
                raise ValidationError("Employee is required.")

            # ✅ Branch fallback
            branch_id = employee.emp_branch_id or employee.work_location

            if not branch_id:
                raise ValidationError(
                    "Employee branch is missing in employee master."
                )

            # ✅ Get document numbering config
            try:
                doc_config = DocumentNumbering.objects.get(
                    branch_id=branch_id,
                    type='resignation_request'
                )

            except DocumentNumbering.DoesNotExist:

                raise NotFound(
                    f"No document numbering configuration found for branch "
                    f"{branch_id} and resignation request."
                )

            current_date = timezone.now().date()

            # ✅ Manual document number validation
            if document_number:

                if doc_config.start_date and doc_config.end_date:

                    if not (
                        doc_config.start_date <= current_date <= doc_config.end_date
                    ):

                        raise ValidationError(
                            "Document number cannot be assigned outside "
                            "the valid date range."
                        )

            else:
                # ✅ Auto-generate document number
                document_number = doc_config.get_next_number()

            serializer.save(document_number=document_number)

    @action(detail=False, methods=['get'], url_path='approved_resignations',permission_classes=[CanViewApprovedResignations])
    def list_approved_resignations(self, request):
        # Fetch all approved resignations
        approved_resignations = EmployeeResignation.objects.filter(status='Approved')

        # Serialize the employee details only
        data = []
        for resignation in approved_resignations:
            employee = resignation.employee
            data.append({
                'employee_id': employee.id,
                'employee_code': getattr(employee, 'emp_code', None),
                'employee_name': f"{getattr(employee, 'emp_first_name', '')} {getattr(employee, 'emp_last_name', '')}".strip(),
                'resignation_id': resignation.id,
                'resigned_on': resignation.resigned_on,
                'last_working_date': resignation.last_working_date,
                'status': resignation.status,
            })

        return Response(data, status=status.HTTP_200_OK)
    @action(detail=False, methods=['post'], url_path='create_eos_by_employee/(?P<employee_id>[^/.]+)',permission_classes=[CanCreateEOS])
    def create_eos_by_employee(self, request, employee_id=None):
        try:
            # Get the latest approved resignation for this employee
            resignation = EmployeeResignation.objects.filter(
                employee_id=employee_id,
                status='Approved'
            ).order_by('-id').first()

            if not resignation:
                return Response({'detail': 'No approved resignation found for this employee.'},
                                status=status.HTTP_400_BAD_REQUEST)

            # Check if EOS already exists
            if hasattr(resignation, 'eos'):
                return Response({'detail': 'EOS already exists for this resignation.'},
                                status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():
                employee = resignation.employee
                start_date = employee.emp_joined_date
                end_date = resignation.last_working_date

                if not start_date or not end_date:
                    return Response({'detail': 'Invalid joining or last working date.'},
                                    status=status.HTTP_400_BAD_REQUEST)

                # Calculate service time
                total_days = (end_date - start_date).days
                years_of_service = total_days / 365.0

                eos = EndOfService.objects.create(
                    resignation=resignation,
                    date_of_joining=start_date,
                    date_of_resignation_termination=resignation.resigned_on,
                    last_working_date=end_date,
                    years_of_service=years_of_service,
                    total_service_days=total_days,
                )

                calculate_settlement(eos)  # Gratuity logic here

                return Response({'detail': 'EOS created successfully.'}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='employees_with_eos')
    def employees_with_eos(self, request):
        eos_records = EndOfService.objects.select_related('resignation__employee')

        data = []
        for eos in eos_records:
            employee = eos.resignation.employee
            data.append({
                'employee_id': employee.id,
                'employee_code': getattr(employee, 'emp_code', None),
                'employee_name': f"{getattr(employee, 'emp_first_name', '')} {getattr(employee, 'emp_last_name', '')}".strip(),
                'resignation_id': eos.resignation.id,
                'eos_id': eos.id,
                'date_of_joining': eos.date_of_joining,
                'last_working_date': eos.last_working_date,
                'years_of_service': eos.years_of_service,
                'total_service_days': eos.total_service_days,
            })

        return Response(data, status=status.HTTP_200_OK)
class ResignationApprovalLevelViewset(viewsets.ModelViewSet):
    queryset = ResignationApprovalWorkflow.objects.all()
    serializer_class = ResignationApprovalWorkflowSerializer

class ResignationApprovalViewset(viewsets.ModelViewSet):
    queryset = ResignationApproval.objects.all()
    serializer_class = ResignationApprovalSerializer

    def get_queryset(self):
        user = self.request.user

        if not user.is_authenticated:
            return ResignationApproval.objects.none()

        if user.is_superuser:
            return ResignationApproval.objects.all()

        return ResignationApproval.objects.filter(
        Q(approver=user) |
        Q(deligate_to=user, is_deligate=True)
    ).distinct()

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approval = self.get_object()
        # if request.user != approval.approver:
        #     return Response({'error': 'You are not authorized to approve this request.'}, status=status.HTTP_403_FORBIDDEN)

        note = request.data.get('note')  # Get the note from the request
        approval.approve(note=note)
        return Response({'status': 'approved', 'note': note}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')  # Get the note from the request
        approval.reject(note=note)
        return Response({'status': 'rejected', 'note': note}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=["post"])
    def delegate(self, request, pk=None):
        approval = self.get_object()

        delegate_user_id = request.data.get("deligate_to")

        if not delegate_user_id:
            return Response(
                {"error": "Delegate user is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        delegate_user = get_object_or_404(CustomUser, pk=delegate_user_id)

        if delegate_user == approval.approver:
            return Response(
                {"error": "You cannot delegate to yourself."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if approval.is_deligate:
            return Response(
                {"error": "This approval has already been delegated."},
                status=status.HTTP_400_BAD_REQUEST
            )

        approval.deligate_to = delegate_user
        approval.is_deligate = True
        approval.deligate_response = None
        approval.save()

        if delegate_user.email:

            subject = "Delegation Assigned"

            message = f"""
                Delegation Assigned

                Hello {delegate_user.get_username() or delegate_user.username},

                You have been assigned a new delegation request.

                REQUEST DETAILS
                _________________

                Document Number : {approval.resignation_request.document_number}
                Employee        : {approval.resignation_request.employee}
                Request Type    : {approval.resignation_request.termination_type}
                Status          : {approval.resignation_request.status}

                Please review the request and send your response to the original approver.

                Thank You.
                """

            send_mail(
                subject,
                message,
                None,
                [delegate_user.email],
                fail_silently=False,
            )
            created_notification = send_notification_email(
                user=delegate_user,
                employee=None,
                branch=None,
                title="Delegation Assigned",
                message=f"{approval.approver.username} has delegated request {approval.resignation_request.document_number} to you.",
                delegate_user=approval.approver,
                template_type="request_created",
                context={
                    **get_employee_context(approval.resignation_request.employee),
                        "doc_number": approval.resignation_request.document_number,
                        "request_type": approval.resignation_request.termination_type,
                    },
                    email_template_model=ResignationEmailTemplate,
                    notification_model=ResignationRequestNotification
            )

            print("Notification Created:", created_notification)

            return Response(
                {
                    "message": "Approval delegated successfully.",
                    "approval_id": approval.id,
                    "approver": approval.approver.username,
                    "delegate_to": delegate_user.username,
                    "status": approval.status,
                },
                status=status.HTTP_200_OK,
            )
    @action(detail=True, methods=["post"])
    def send_response(self, request, pk=None):
        approval = self.get_object()

        response_text = request.data.get("deligate_response")

        if not response_text:
            return Response({"error": "Response is required"}, status=400)

        approval.deligate_response = response_text
        approval.save()

        # ---------------- EMAIL ----------------
        if approval.approver and approval.approver.email:
            send_mail(
                subject="Delegation Response Received",
                message=response_text,
                from_email=None,
                recipient_list=[approval.approver.email],
                fail_silently=False,
            )

        # ---------------- NOTIFICATION ----------------
        send_notification_email(
            user=approval.approver,
            employee=None,
            branch=None,
            title="Delegation Response Received",
            message=response_text,
            delegate_user=approval.deligate_to,
            template_type="request_created",
            context={
                 **get_employee_context(approval.resignation_request.employee),
                    "doc_number": approval.resignation_request.document_number,
                    "request_type": approval.resignation_request.termination_type,
                },
                email_template_model=ResignationEmailTemplate,
                notification_model=ResignationRequestNotification
                )

        return Response({
            "message": "Response sent successfully",
            "response": response_text
        })


class ResignationEmailTemplateViewset(viewsets.ModelViewSet):
    queryset = ResignationEmailTemplate.objects.all()
    serializer_class = ResignationTemplateSerializer
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            
            'employee': [
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}',
                '{{document_date}}',
                '{{resigned_on}}',
                '{{notice_period}}',
                '{{last_working_date}}',
                '{{location}}',
                '{{termination_type}}',
                '{{reason_for_leaving}}',
                '{{status}}',
            ]
        }
        return Response(placeholders)
    # Custom action to fetch the available From and To addresses
    @action(detail=False, methods=['get'], url_path='from-to-addresses')
    def from_to_list(self, request):
        # Fetch active email configurations for "From" addresses
        from_addresses = EmailConfiguration.objects.filter(is_active=True).values_list('email_host_user', flat=True)

        # Fetch employee emails for "To" addresses
        to_addresses = emp_master.objects.all().values_list('emp_personal_email', 'emp_company_email')

        to_list = []
        for emp_personal, emp_company in to_addresses:
            if emp_personal:
                to_list.append(emp_personal)
            if emp_company:
                to_list.append(emp_company)

        return Response({
            'from_addresses': from_addresses,
            'to_addresses': to_list
        })
class ResignationRequestNotificationViewset(viewsets.ModelViewSet):
    queryset = ResignationRequestNotification.objects.all()
    serializer_class = ResignationRequestNotificationSerializer
    def get_queryset(self):
        user = self.request.user

        # Admin / staff / superuser → see all request notifications
        if user.is_superuser or user.is_staff:
            return ResignationRequestNotification.objects.all().order_by('-created_at')

        # Normal user → show request notifications assigned directly to them
        qs = ResignationRequestNotification.objects.filter(
            Q(recipient_user=user) |
            Q(recipient_employee__users=user)      # employee assigned to this user
        ).order_by('-created_at')

        return qs
    

from rest_framework.response import Response
from django.http import FileResponse
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.shortcuts import get_object_or_404
from decimal import Decimal
class EndOfServiceViewset(viewsets.ModelViewSet):
    queryset = EndOfService.objects.all()
    serializer_class = EndOfServiceSerializer
    @action(detail=False, methods=['get'], url_path='employee/(?P<employee_id>[^/.]+)')
    def get_by_employee(self, request, employee_id=None):
        try:
            eos = EndOfService.objects.get(resignation__employee_id=employee_id)
            serializer = self.get_serializer(eos)
            return Response(serializer.data)
        except EndOfService.DoesNotExist:
            return Response({"detail": "End of service not found for this employee."}, status=status.HTTP_404_NOT_FOUND)
    @action(detail=True, methods=['get'], url_path='final-settlement-data')
    def final_settlement_data(self, request, pk=None):
        eos = get_object_or_404(EndOfService, pk=pk)
        resignation = eos.resignation
        employee = resignation.employee

        # Get latest approved payslip
        payslip = employee.payslips.filter(status='Approved').order_by('-created_at').first()

        if not payslip:
            return Response({"detail": "Approved payslip not found."}, status=status.HTTP_404_NOT_FOUND)

        # Payslip component details
        components = payslip.components.all()
        deductions = []
        additions = []
        total_deductions = 0
        total_additions = 0

        for comp in components:
            item = {
                "component_name": comp.component.name,
                "amount": float(comp.amount)
            }
            if comp.component.component_type == 'deduction':
                deductions.append(item)
                total_deductions += float(comp.amount)
            elif comp.component.component_type == 'addition':
                additions.append(item)
                total_additions += float(comp.amount)

        # Include EOS specific amounts as additions
        if eos.gratuity_amount:
            additions.append({"component_name": "Gratuity", "amount": float(eos.gratuity_amount)})
            total_additions += float(eos.gratuity_amount)
        if eos.notice_pay:
            additions.append({"component_name": "Notice Pay", "amount": float(eos.notice_pay)})
            total_additions += float(eos.notice_pay)
        if eos.final_month_salary:
            additions.append({"component_name": "Final Month Salary", "amount": float(eos.final_month_salary)})
            total_additions += float(eos.final_month_salary)
        if eos.air_ticket:
            additions.append({"component_name": "Air Ticket", "amount": float(eos.air_ticket)})
            total_additions += float(eos.air_ticket)

        net_amount = total_additions - total_deductions

        # Construct response
        data = {
            "employee": {
                "code": employee.emp_code,
                "name": f"{employee.emp_first_name} {employee.emp_last_name}",
                "designation": str(employee.emp_desgntn_id),
                "department": str(employee.emp_dept_id),
            },
            "gratuity_entitlement": {
                "date_of_joining": eos.date_of_joining,
                "work_status": "Resigned",
                "date_of_resignation_termination": eos.date_of_resignation_termination,
                "notice_period_days": eos.notice_period_days,
                "last_working_date": eos.last_working_date,
                "total_service_days": eos.total_service_days,
                "leave_days_without_pay": eos.leave_days_without_pay,
                "net_days_worked": eos.net_number_of_days_worked,
                "basic_salary": float(payslip.gross_salary or 0),
                "gratuity_days": eos.gratuity_days,
                "last_month_salary": eos.last_month_salary,
            },
            "payslip_summary": {
                "deductions": deductions,
                "additions": additions,
                "total_deductions": total_deductions,
                "total_additions": total_additions,
                "net_amount": net_amount
            },
            "status": eos.status,
            "processed_date": eos.processed_date,
        }

        return Response(data)
class EscalationRuleViewSet(viewsets.ModelViewSet):
    """
    API for managing escalation settings on each approval level.
    """
    serializer_class = EscalationRuleSerializer
    queryset = ApprovalLevel.objects.all().order_by('workflow__request_type', 'level')

    def get_queryset(self):
        queryset = super().get_queryset()
        request_type_id = self.request.query_params.get('request_type')
        branch_id = self.request.query_params.get('branch')

        if request_type_id:
            queryset = queryset.filter(request_type_id=request_type_id)
        if branch_id:
            queryset = queryset.filter(branch__id=branch_id)

        return queryset.distinct()

    def update(self, request, *args, **kwargs):
        """
        Update only escalation fields for a level.
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response({
            "message": "Escalation rule updated successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)
    @action(detail=True, methods=['post'])
    def reset(self, request, pk=None):
        instance = self.get_object()
        instance.escalate_to = None
        instance.escalate_after_days = 0
        instance.escalate_after_hours = 0
        instance.escalate_after_minutes = 0
        instance.save()
        return Response({"message": "Escalation rule reset successfully"}, status=200)

class EmployeeByUserViewSet(EmpViewSet):
    serializer_class = EmpSerializer
    def get_queryset(self):
        return emp_master.objects.filter(users=self.request.user, is_ess=False)

class ResignationRequestNotificationViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ResignationRequestNotification.objects.all()
    serializer_class = ResignationRequestNotificationSerializer   
    def get_queryset(self):
        user = self.request.user

        # Admin / staff / superuser → see all request notifications
        if user.is_superuser or user.is_staff:
            return ResignationRequestNotification.objects.all().order_by('-created_at')

        # Normal user → show request notifications assigned directly to them
        qs = ResignationRequestNotification.objects.filter(
            Q(recipient_user=user) |
            Q(recipient_employee__users=user)      # employee assigned to this user
        ).order_by('-created_at')

        return qs