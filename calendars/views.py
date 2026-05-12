from django.shortcuts import render
from .models import( weekend_calendar,assign_weekend,holiday,holiday_calendar,assign_holiday,WeekendDetail,leave_type,leave_entitlement,applicablity_critirea,emp_leave_balance,leave_accrual_transaction,leave_reset_transaction,employee_leave_request,Attendance,Shift,
                     EmployeeMachineMapping,LeaveReport,LeaveApprovalLevels,LeaveApproval,LvEmailTemplate,LvApprovalNotify,LvCommonWorkflow,LvRejectionReason,LeaveApprovalReport,
                     AttendanceReport,lvBalanceReport,EmployeeYearlyCalendar,CompensatoryLeaveRequest,CompensatoryLeaveTransaction,CompensatoryLeaveBalance,ShiftPattern,EmployeeShiftSchedule,ShiftOverride,LeaveResetPolicy,LeaveCarryForwardTransaction,
                     LeaveEncashmentTransaction,EmployeeRejoining,EmployeeOvertime,MonthlyAttendanceSummary,AttendanceRecheck,OvertimePolicy,OvertimeRule,AttendanceLog,AttendancePolicy,LeavePayRule,
                     LatinEarlyoutEmailTemplate,LateinEarlyRequestNotification,LateinEarlyoutRequest,LateinEarlyoutApprovalLevel,LateinEarlyoutApproval,LVApprovalWorkflow,LatinEarlyApprovalWorkflow,AttendanceCalendar
                     )
from . serializer import (WeekendCalendarSerailizer,WeekendAssignSerializer,HolidayAssignSerializer,HolidayCalandarSerializer,HolidaySerializer,WeekendDetailSerializer,LeaveTypeSerializer,LeaveEntitlementSerializer,ApplicableSerializer,EmployeeLeaveBalanceSerializer,AccrualSerializer,ResetSerializer,LeaveRequestSerializer,
                         AttendanceSerializer,ShiftSerializer,ImportAttendanceSerializer,EmployeeMappingSerializer,LeaveReportSerializer,LvApprovalLevelSerializer,EmployeeYearlyCalendarSerializer,
                         LvApprovalSerializer,LvEmailTemplateSerializer,LvApprovalNotifySerializer,LvCommonWorkflowSerializer,LvRejectionReasonSerializer,LvApprovalReportSerializer,AttendanceReportSerializer,lvBalanceReportSerializer,
                         CompensatoryLeaveRequestSerializer,CompensatoryLeaveTransactionSerializer,CompensatoryLeaveBalanceSerializer,ShiftOverrideSerializer,ShiftPatternSerializer,EmployeeShiftScheduleSerializer,LeaveResetPolicySerializer,LeaveCarryForwardTransactionSerializer,
                         LeaveEncashmentTransactionSerializer,EmpOpeningsBlkupldSerializer,EmployeeRejoiningSerializer,EmployeeOvertimeSerializer,MonthlyAttendanceSummarySerializer,LVEscalationRuleSerializer,AttendanceRecheckSerializer,OvertimePolicySerializer,OvertimeRuleSerializer,
                         AttendanceLogSerializer,AttendancePolicySerializer,LeavePayRuleSerializer,
                         LatinEarlyoutEmailTemplateSerializer,LateinEarlyRequestNotificationSerializer,LateinEarlyoutRequestSerializer,LateinEarlyoutApprovalLevelSerializer, LateinEarlyoutApprovalSerializer,LVApprovalWorkflowSerializer,LatinEarlyApprovalWorkflowSerializer,AttendanceCalendarSerializer
                         )
from . import face_utils
from rest_framework import viewsets,filters,status
from rest_framework.response import Response
from rest_framework.decorators import action

from EmpManagement.models import emp_master
from .permissions import( WeekendCalendarPermission, WeekendDetailPermission, AssignWeekendPermission, HolidayPermission, HolidayCalendarPermission, AssignHolidayPermission,LeaveTypePermission,LeaveEntitlementPermission,EmpLeaveBalancePermission,ApplicabilityCriteriaPermission,EmployeeLeaveRequestPermission,LvEmailTemplatePermission,
                            LvCommonWorkflowPermission,LvRejectionReasonPermission,LeaveApprovalLevelsPermission,EmployeeMachineMappingPermission,ShiftPermission,ShiftPatternPermission,AttendancePermission,CompensatoryLeaveRequestPermission,CompensatoryLeaveTransactionPermission,CompensatoryLeaveBalancePermission,CompensatoryLeaveRequestPermission,
                            LeaveReportPermission,LeaveApprovalReportPermission,AttendanceReportPermission,LvBalanceReportPermission,LeaveAccrualTransactionPermission,LeaveResetTransactionPermission,ShiftOverridePermission,WeekPatternAssignmentPermission,EmployeeShiftSchedulePermission,EmployeeYearlyCalendarPermission,
                        )
from rest_framework.parsers import MultiPartParser, FormParser
from EmpManagement.models import emp_master
from .resource import AttendanceResource,EmployeeOpenBalanceResource,MonthlyAttendanceResource
from django.http import HttpResponse,JsonResponse
from tablib import Dataset
from django.core.exceptions import ValidationError
from rest_framework.views import APIView
from rest_framework.decorators import action
from django.utils import timezone
from django.shortcuts import get_object_or_404
from EmpManagement .models import EmailConfiguration
from django.utils.timezone import localtime,now
from django.conf import settings
import os,json
from datetime import date
from collections import defaultdict
from django.core.cache import cache
import redis
from rest_framework import viewsets,filters, status
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
from django.db.models import Field
from django.db import transaction
from django.db.models import Q
from OrganisationManager.models import DocumentNumbering,BranchGeoFence
from OrganisationManager.serializer import DocumentNumberingSerializer
from rest_framework.exceptions import NotFound
from import_export.formats.base_formats import XLSX,CSV
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from datetime import date
from dateutil.relativedelta import relativedelta
from .utils import get_attendance_summary
from .serializer import AttendanceSummarySerializer
from EmpManagement.models import emp_master
import calendar
from django.utils.dateparse import parse_date
from EmpManagement.utils import send_notification_email, get_employee_context
from django.core.mail import EmailMessage
import math
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from Core .mixins import BranchAccessMixin
from .utils import validate_employee_geofence,apply_check_in_policy,apply_check_out_policy
# Create your views here.

class WeekendDetailsViewset(viewsets.ModelViewSet):
    queryset = WeekendDetail.objects.all()
    serializer_class = WeekendDetailSerializer
    permission_classes = [WeekendCalendarPermission]

class WeekendViewset(viewsets.ModelViewSet):
    queryset = weekend_calendar.objects.all()
    serializer_class = WeekendCalendarSerailizer
    permission_classes = [WeekendDetailPermission]

    @action(detail=True, methods=['get'])
    def details(self, request, pk=None):
        weekend_calendar = self.get_object()
        serializer = self.get_serializer(weekend_calendar)
        return Response(serializer.data)
    # @action(detail=False, methods=['post'])
    # @action(detail=False, methods=['post'])
    # def set_monthly_weekends(self, request):
    #     data = request.data
    #     calendar_code = data.get('calendar_code')
    #     year = data.get('year')
    #     weekday = data.get('weekday')
    #     day_type = data.get('day_type')
    #     week_of_month = data.get('week_of_month')

    #     weekend_calendar, created = weekend_calendar.objects.get_or_create(calendar_code=calendar_code, year=year)
    #     for month in range(1, 13):
    #         for week in range(1, 6):  # Assuming up to 5 weeks in a month
    #             if week == week_of_month:
    #                 WeekendDetail.objects.create(
    #                     weekend_calendar=weekend_calendar,
    #                     weekday=weekday,
    #                     day_type=day_type,
    #                     week_of_month=week
    #                 )
        
    #     serializer = self.get_serializer(weekend_calendar)
    #     return Response(serializer.data)


class AssignWeekendViewset(viewsets.ModelViewSet):
    queryset = assign_weekend.objects.all()
    serializer_class = WeekendAssignSerializer
    permission_classes = [AssignWeekendPermission]
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

class HolidayViewset(viewsets.ModelViewSet):
    queryset = holiday.objects.all()
    serializer_class = HolidaySerializer
    permission_classes = [HolidayPermission]

class HolidayCalendarViewset(viewsets.ModelViewSet):
    queryset = holiday_calendar.objects.all()
    serializer_class = HolidayCalandarSerializer
    permission_classes = [HolidayCalendarPermission]

class HolidayAssignViewset(viewsets.ModelViewSet):
    queryset = assign_holiday.objects.all()
    serializer_class = HolidayAssignSerializer
    permission_classes = [AssignHolidayPermission]

#leave
# Create your views here.

class LeaveTypeviewset(BranchAccessMixin,viewsets.ModelViewSet):
    queryset = leave_type.objects.all()
    serializer_class = LeaveTypeSerializer
    permission_classes = [LeaveTypePermission] 
    
class LvEmailTemplateviewset(viewsets.ModelViewSet):
    queryset = LvEmailTemplate.objects.all()
    serializer_class = LvEmailTemplateSerializer
    permission_classes = [LvEmailTemplatePermission] 
    
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            
            'employee': [
                '{{ document_number }}',
                '{{ leave_type }}',
                '{{ reason }}',
                '{{ start_date }}',
                '{{ end_date }}',
                '{{ recipient_name }}',
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_gender }}',
                '{{ emp_date_of_birth }}',
                '{{ emp_personal_email }}',
                '{{ emp_company_email }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}'
            ]
        }
        return Response(placeholders)
    # Custom action to fetch the available From and To addresses
    @action(detail=False, methods=['get'], url_path='from-to-addresses')
    def from_to_list(self, request):
        # Fetch active email configurations for "From" addresses
        from_addresses = EmailConfiguration.objects.filter(is_active=True).values_list('email_host_user', flat=True)

        # Fetch employee emails for "To" addresses
        to_addresses = emp_master.objects.all().values_list('emp_personal_email', 'emp_company_email')

        to_list = []
        for emp_personal, emp_company in to_addresses:
            if emp_personal:
                to_list.append(emp_personal)
            if emp_company:
                to_list.append(emp_company)

        return Response({
            'from_addresses': from_addresses,
            'to_addresses': to_list
        })

class LvApprovalNotifyviewset(viewsets.ModelViewSet):
    queryset = LvApprovalNotify.objects.all()
    serializer_class = LvApprovalNotifySerializer
    def get_queryset(self):
        user = self.request.user

        # Admin / staff / superuser → see all request notifications
        if user.is_superuser or user.is_staff:
            return LvApprovalNotify.objects.all().order_by('-created_at')

        # Normal user → show request notifications assigned directly to them
        qs = LvApprovalNotify.objects.filter(
            Q(recipient_user=user) |
            Q(recipient_employee__users=user)      # employee assigned to this user
        ).order_by('-created_at')

        return qs
class LeaveEntitlementviewset(viewsets.ModelViewSet):
    queryset = leave_entitlement.objects.all()
    serializer_class = LeaveEntitlementSerializer
    permission_classes = [LeaveEntitlementPermission]
    def perform_create(self, serializer):
        instance = serializer.save()
        self.process_accrual(instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        self.process_accrual(instance)

    def process_accrual(self, instance):
        if instance.accrual and instance.accrual_month == timezone.now().strftime('%b') and instance.accrual_day == '1st':
            employees = emp_leave_balance.objects.filter(leave_type=instance.leave_type)
            for emp_balance in employees:
                leave_accrual_transaction.objects.create(
                    employee=emp_balance.employee,
                    leave_type=instance.leave_type,
                    accrual_date=timezone.now().date(),
                    amount=instance.accrual_rate
                )
                emp_balance.balance += instance.accrual_rate
                emp_balance.save()
class LeavePayRuleViewset(viewsets.ModelViewSet):
    queryset = LeavePayRule.objects.all()
    serializer_class = LeavePayRuleSerializer
class LeaveResetPolicyviewset(viewsets.ModelViewSet):
    queryset = LeaveResetPolicy.objects.all()
    serializer_class = LeaveResetPolicySerializer

class LeaveCarryForwardTransactionviewset(viewsets.ModelViewSet):
    queryset = LeaveCarryForwardTransaction.objects.all()
    serializer_class = LeaveCarryForwardTransactionSerializer

class LeaveEncashmentTransactionviewset(viewsets.ModelViewSet):
    queryset = LeaveEncashmentTransaction.objects.all()
    serializer_class = LeaveEncashmentTransactionSerializer

class Applicableviewset(viewsets.ModelViewSet):
    queryset = applicablity_critirea.objects.all()
    serializer_class = ApplicableSerializer
    # permission_classes = [ApplicabilityCriteriaPermission] 



class leave_balance_viewset(viewsets.ModelViewSet):
    queryset = emp_leave_balance.objects.all()
    serializer_class = EmployeeLeaveBalanceSerializer
    # permission_classes = [EmpLeaveBalancePermission] 

class Acrualviewset(viewsets.ModelViewSet):
    queryset = leave_accrual_transaction.objects.all()
    serializer_class = AccrualSerializer
    permission_classes = [LeaveAccrualTransactionPermission] 

class Resetviewset(viewsets.ModelViewSet):
    queryset = leave_reset_transaction.objects.all()
    serializer_class = ResetSerializer
    permission_classes = [LeaveResetTransactionPermission] 

# class Enchashviewset(viewsets.ModelViewSet):
#     queryset = leave_encashment.objects.all()
#     serializer_class = EnchashSerializer


class LeaveRequestviewset(viewsets.ModelViewSet):
    queryset = employee_leave_request.objects.all()
    serializer_class = LeaveRequestSerializer
    permission_classes = [EmployeeLeaveRequestPermission]
    def get_queryset(self):
        # Filter queryset based on user access
        if self.request.user.is_ess:
            # Return only requests related to the ESS user's employee record
            return self.queryset.filter(employee__emp_code=self.request.user.username)
        return super().get_queryset()  # Non-ESS users can access as per their permissions
    @action(detail=False, methods=['get'], url_path='approved-leaves')
    def approved_leaves(self, request):
        approved_queryset = employee_leave_request.objects.filter(status='approved')
        serializer = self.get_serializer(approved_queryset, many=True)
        return Response(serializer.data)
    def perform_create(self, serializer):
        with transaction.atomic():
            employee = serializer.validated_data.get('employee')
            branch_id = employee.emp_branch_id.id  # Ensure this field exists
            leave_type = serializer.validated_data['leave_type']

            try:
                doc_config = DocumentNumbering.objects.get(
                    branch_id=branch_id,
                    type='leave_request',
                )
            except DocumentNumbering.DoesNotExist:
                raise NotFound(f"No document numbering configuration found for branch {branch_id} and leave type {leave_type}.")

            # Check if the user entered a document number manually
            document_number = serializer.validated_data.get('document_number')

            if document_number:
                # Validate the entered document number falls within the allowed date range
                current_date = timezone.now().date()
                if doc_config.start_date and doc_config.end_date:
                    if not (doc_config.start_date <= current_date <= doc_config.end_date):
                        raise ValidationError("The document number cannot be assigned outside the valid date range.")
            else:
                # Generate the document number automatically
                document_number = doc_config.get_next_number()

            serializer.save(document_number=document_number)
            


    @action(detail=False, methods=['get'], url_path='leave-request-history')
    def employee_leave_request(self, request):
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({'error': 'Employee ID is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        requests = employee_leave_request.objects.filter(employee_id=employee_id).order_by('-applied_on')
     
        # Manually serialize the fields you want
        history_data = []
        for request in requests:
            history_data.append({
                'start_date': request.start_date,
                'end_date': request.end_date,
                'leave_type': request.leave_type.name if request.leave_type else None,
                'reason': request.reason ,
                'status': request.status,
                'applied_on': request.applied_on,
                'number_of_days':request.number_of_days
            })

        return Response(history_data, status=status.HTTP_200_OK)
    
#     def get_serializer_class(self):
#         if self.request.method in ['POST', 'PUT']:
#             return EmployeeLeaveSerializer
#         return super().get_serializer_class()

#     def get_serializer_context(self):
#         context = super().get_serializer_context()
#         context['employee_id'] = self.request.data.get('employee_id', None)
#         return context

#filtering for using assigned models for employees
from rest_framework.response import Response
from rest_framework.decorators import action

class LeaveTypeViewSet(viewsets.ViewSet):
    @action(detail=False, methods=['get'])
    def available_leave_types(self, request):
        employee_id = request.query_params.get('employee_id')
        if not employee_id:
            return Response({"error": "employee_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = emp_master.objects.get(id=employee_id)
        except emp_master.DoesNotExist:
            return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)

        leave_types = leave_type.objects.filter(
            id__in=emp_leave_balance.objects.filter(employee=employee).values_list('leave_type_id', flat=True)
        )
        serializer = LeaveTypeSerializer(leave_types, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class EmployeeMachineMappingViewset(viewsets.ModelViewSet):
    queryset =EmployeeMachineMapping.objects.all()
    serializer_class = EmployeeMappingSerializer
    permission_classes = [EmployeeMachineMappingPermission] 
    

class ShiftViewSet(viewsets.ModelViewSet):
    queryset = Shift.objects.all()
    serializer_class = ShiftSerializer
    permission_classes = [ShiftPermission] 

    

class ShiftPatternViewSet(viewsets.ModelViewSet):
    queryset = ShiftPattern.objects.all()
    serializer_class = ShiftPatternSerializer
    permission_classes = [ShiftPatternPermission]

class ShiftOverrideViewSet(viewsets.ModelViewSet):
    queryset = ShiftOverride.objects.all()
    serializer_class = ShiftOverrideSerializer
    permission_classes = [ShiftOverridePermission]

class OvertimePolicyViewSet(viewsets.ModelViewSet):
    queryset = OvertimePolicy.objects.all()
    serializer_class = OvertimePolicySerializer
class OvertimeRuleViewSet(viewsets.ModelViewSet):
    queryset = OvertimeRule.objects.all()
    serializer_class = OvertimeRuleSerializer

class EmployeeShiftScheduleViewSet(viewsets.ModelViewSet):
    queryset = EmployeeShiftSchedule.objects.all()
    serializer_class = EmployeeShiftScheduleSerializer
    # permission_classes = [EmployeeShiftSchedulePermission]
    def get_shift_for_day(self, request, *args, **kwargs):
        """
        Get shift for a given employee and date.
        URL parameters should include employee_id and date (format: YYYY-MM-DD).
        """
        employee_id = request.query_params.get('employee')
        date_str = request.query_params.get('date')
        
        if employee_id and date_str:
            try:
                employee = emp_master.objects.get(id=employee_id)
                date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
                schedule = self.get_object()  # Assume the schedule is retrieved by URL ID
                # Removed extra employee argument:
                shift = schedule.get_shift_for_date(date)
                
                if shift:
                    return Response({"shift": str(shift)}, status=status.HTTP_200_OK)
                else:
                    return Response({"error": "No shift found for the specified date"}, status=status.HTTP_404_NOT_FOUND)
            except emp_master.DoesNotExist:
                return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response({"error": "Invalid parameters"}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def get_shifts_for_year(self, request):
        schedule_id = request.query_params.get('schedule_id')
        year = request.query_params.get('year')
        month = request.query_params.get('month')

        if not schedule_id or not year:
            return Response(
                {"error": "schedule_id and year are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        schedule = get_object_or_404(EmployeeShiftSchedule, id=schedule_id)

        try:
            year = int(year)
            month = int(month) if month else None
        except ValueError:
            return Response(
                {"error": "Invalid year or month"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Date range
        if month:
            start_date = datetime(year, month, 1).date()
            end_date = (
                datetime(year, month + 1, 1) - timedelta(days=1)
                if month < 12 else datetime(year, 12, 31).date()
            )
        else:
            start_date = datetime(year, 1, 1).date()
            end_date = datetime(year, 12, 31).date()

        # ✅ CORRECT EMPLOYEE RESOLUTION
        employees = schedule.get_assigned_employees()

        if not employees.exists():
            return Response(
                {"error": "No employees matched this shift schedule criteria"},
                status=status.HTTP_404_NOT_FOUND
            )

        all_dates = [
            start_date + timedelta(days=i)
            for i in range((end_date - start_date).days + 1)
        ]

        shifts_calendar = {}

        # Pre-calc shifts once
        shift_map = {
            date: schedule.get_shift_for_date(date)
            for date in all_dates
        }

        for employee in employees:
            shifts_calendar[employee.emp_code] = {}

            for date, shift in shift_map.items():
                shifts_calendar[employee.emp_code][
                    date.strftime("%d-%m-%Y")
                ] = str(shift) if shift else "No shift"

        return Response(
            {
                "schedule": schedule.schedule_name,
                "year": year,
                "month": month if month else "Full Year",
                "total_employees": employees.count(),
                "shifts": shifts_calendar
            },
            status=status.HTTP_200_OK
        )
class AttendancePolicyViewset(viewsets.ModelViewSet):
    queryset = AttendancePolicy.objects.all()
    serializer_class = AttendancePolicySerializer
class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    parser_classes = [viewsets.ModelViewSet.parser_classes[0], MultiPartParser, FormParser] # JSON, MultiPart, Form
    # permission_classes = [AttendancePermission]
    from .utils import validate_employee_geofence
    
    @staticmethod
    def parse_date(date_string):
        try:
            # Parse the date from string to a datetime.date object
            return datetime.strptime(date_string, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None
    
    @action(detail=False, methods=['post'])
    def enroll_face(self, request):
        emp_id = request.data.get("employee")
        # Support both file upload (request.FILES) and base64 (request.data)
        face_photo = request.FILES.get("face_photo") or request.data.get("face_photo")
        
        if not emp_id or not face_photo:
            return Response({"detail": "Employee ID and face photo are required. You can upload a file or send a base64 string."}, status=400)
            
        try:
            employee = emp_master.objects.get(id=emp_id)
        except emp_master.DoesNotExist:
            return Response({"detail": "Employee not found"}, status=404)
        
        encoding = face_utils.get_face_encoding(face_photo)
        
        if encoding:
            employee.face_encoding = encoding
            employee.save()
            return Response({"detail": "Face enrolled successfully"})
        else:
            if face_utils.DeepFace is None:
                return Response({"detail": "DeepFace library is not installed on the server. Please run: pip install deepface"}, status=500)
            return Response({"detail": "Face enrollment failed. Please ensure your face is clearly visible and not too far from the camera."}, status=400)

    @action(detail=False, methods=['post'])
    def register_barcode(self, request):
        emp_id = request.data.get("employee")
        barcode = request.data.get("barcode")
        
        if not emp_id or not barcode:
            return Response({"detail": "Employee ID and barcode string are required."}, status=400)
            
        try:
            employee = emp_master.objects.get(id=emp_id)
        except emp_master.DoesNotExist:
            return Response({"detail": "Employee not found"}, status=404)
            
        # Check if barcode is already assigned to someone else
        if emp_master.objects.filter(barcode_number=barcode).exclude(id=emp_id).exists():
            return Response({"detail": "This barcode is already assigned to another employee."}, status=400)
            
        employee.barcode_number = barcode
        employee.save()
        
        return Response({"detail": f"Barcode registered successfully for {employee.emp_code}"})
    @action(detail=False, methods=['post'])
    def check_in(self, request):

        emp_id = request.data.get("employee")
        barcode = request.data.get("barcode")

        lat = request.data.get("check_in_lat")
        lng = request.data.get("check_in_lng")

        face_photo = face_utils.convert_base64_to_file(
            request.FILES.get("face_photo") or request.data.get("face_photo"),
            "face"
        )

        check_in_image = face_utils.convert_base64_to_file(
            request.FILES.get("check_in_image") or request.data.get("check_in_image"),
            "check_in"
        )

        # 🔐 AUTH
        employee = None
        auth_method = "manual"
        is_verified = False

        if barcode:
            try:
                employee = emp_master.objects.get(barcode_number=barcode)
                auth_method = "barcode"
                is_verified = True
            except:
                return Response({"detail": "Invalid barcode"}, status=400)

        elif face_photo:
            if not emp_id:
                return Response({"detail": "Employee ID required for face"}, status=400)

            employee = emp_master.objects.get(id=emp_id)

            current_encoding = face_utils.get_face_encoding(face_photo)
            if not current_encoding:
                return Response({"detail": "No face detected"}, status=400)

            if not face_utils.verify_face(employee.face_encoding, current_encoding):
                return Response({"detail": "Face does not match"}, status=400)

            auth_method = "face"
            is_verified = True

        elif emp_id:
            employee = emp_master.objects.get(id=emp_id)
            is_verified = True

        else:
            return Response({"detail": "Provide employee/face/barcode"}, status=400)

        # 🌍 GEOFENCE
        if not validate_employee_geofence(employee, lat, lng):
            return Response({"detail": "Outside geofence"}, status=400)

        attendance, _ = Attendance.objects.get_or_create(
            employee=employee,
            date=now().date()
        )

        current_time = localtime(now()).time()
        current_time = apply_check_in_policy(employee, current_time)

        if not attendance.check_in_time:
            attendance.check_in_time = current_time
            attendance.check_in_lat = lat
            attendance.check_in_lng = lng

            if check_in_image:
                attendance.check_in_image = check_in_image

        AttendanceLog.objects.create(
            attendance=attendance,
            log_type='check_in',
            lat=lat,
            lng=lng,
            is_face_verified=is_verified,
            auth_method=auth_method
        )

        attendance.save()

        return Response({
            "status": "Check-in successful",
            "face_verified": is_verified,
            "check_in_image": request.build_absolute_uri(
                attendance.check_in_image.url
            ) if attendance.check_in_image else None
        })
    @action(detail=False, methods=['post'])
    def check_out(self, request):

        emp_id = request.data.get("employee")
        barcode = request.data.get("barcode")

        lat = request.data.get("check_out_lat")
        lng = request.data.get("check_out_lng")

        face_photo = face_utils.convert_base64_to_file(
            request.FILES.get("face_photo") or request.data.get("face_photo"),
            "face"
        )

        check_out_image = face_utils.convert_base64_to_file(
            request.FILES.get("check_out_image") or request.data.get("check_out_image"),
            "check_out"
        )

        # 🔐 AUTH
        employee = None
        auth_method = "manual"
        is_verified = False

        if barcode:
            employee = emp_master.objects.get(barcode_number=barcode)
            attendance = Attendance.objects.get(employee=employee, date=now().date())
            auth_method = "barcode"
            is_verified = True

        elif face_photo:
            employee = emp_master.objects.get(id=emp_id)

            current_encoding = face_utils.get_face_encoding(face_photo)
            if not current_encoding:
                return Response({"detail": "No face detected"}, status=400)

            if not face_utils.verify_face(employee.face_encoding, current_encoding):
                return Response({"detail": "Face does not match"}, status=400)

            attendance = Attendance.objects.get(employee=employee, date=now().date())
            auth_method = "face"
            is_verified = True

        elif emp_id:
            attendance = Attendance.objects.get(employee_id=emp_id, date=now().date())
            employee = attendance.employee
            is_verified = True

        else:
            return Response({"detail": "Provide employee/face/barcode"}, status=400)

        # 🌍 GEOFENCE
        if not validate_employee_geofence(employee, lat, lng):
            return Response({"detail": "Outside geofence"}, status=400)

        tenant_time = localtime(now()).time()
        tenant_time = apply_check_out_policy(employee, tenant_time)

        attendance.check_out_time = tenant_time
        attendance.check_out_lat = lat
        attendance.check_out_lng = lng

        if check_out_image:
            attendance.check_out_image = check_out_image

        AttendanceLog.objects.create(
            attendance=attendance,
            log_type='check_out',
            lat=lat,
            lng=lng,
            is_face_verified=is_verified,
            auth_method=auth_method
        )

        attendance.calculate_total_hours()
        attendance.save()

        return Response({
            "status": "Check-out recorded successfully",
                "working_hours": str(attendance.total_hours) if attendance.total_hours else None,
                "check_out_image": request.build_absolute_uri(
                    attendance.check_out_image.url
                ) if attendance.check_out_image else None
            },
            status=200)

    @action(detail=False, methods=['get'])
    def employee_attendance(self, request):
        """
        Multi-mode attendance report:
        - No filters → all employees, all dates
        - employee_id → all dates of that employee
        - date filters → filtered attendance (all employees or single employee)
        - month/year → monthly summary (all employees or single employee)
        """

        emp_id = request.query_params.get("employee_id")
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        from_date = request.query_params.get("from_date")
        to_date = request.query_params.get("to_date")

        qs = Attendance.objects.all().select_related("employee", "shift").prefetch_related("logs")

        # ------------------------------------------------------------
        # 1️⃣ IF EMPLOYEE SELECTED → SHOW ONLY THAT EMPLOYEE
        # ------------------------------------------------------------
        if emp_id:
            qs = qs.filter(employee=emp_id)

        # ------------------------------------------------------------
        # 2️⃣ APPLY DATE FILTERS IF GIVEN
        # ------------------------------------------------------------
        if month and year:
            qs = qs.filter(date__month=month, date__year=year)

        if from_date and to_date:
            qs = qs.filter(date__range=[from_date, to_date])

        if from_date and not to_date:
            qs = qs.filter(date=from_date)

        # ------------------------------------------------------------
        # 3️⃣ NO FILTER AT ALL → LIST ALL EMPLOYEES WITH ALL DATES
        # ------------------------------------------------------------
        qs = qs.order_by("employee__emp_first_name", "-date")

        result = {}

        for att in qs:
            emp_key = f"{att.employee.id} - {att.employee.emp_first_name}"

            if emp_key not in result:
                result[emp_key] = []

            # Convert durations to readable string
            total_hours = str(att.total_hours) if att.total_hours else "00:00:00"
            overtime = getattr(att, "overtime_hours", None)
            overtime_str = str(overtime) if overtime else "00:00:00"

            result[emp_key].append({
                "date": att.date,
                "day": att.date.strftime("%A"),
                "shift": att.shift.name if att.shift else None,
                "check_in": att.check_in_time,
                "check_out": att.check_out_time,
                "check_in_location": att.check_in_location,
                "check_out_location": att.check_out_location,
                "total_hours": total_hours,
                "overtime": overtime_str,
                "is_face_verified": any(log.is_face_verified for log in att.logs.all()),
                "logs": AttendanceLogSerializer(att.logs.all(), many=True).data,
            })

        return Response({
            "filtered_employee": emp_id,
            "attendance": result
        })


    @action(detail=False, methods=['get'])
    def monthly_late_and_early_attendance(self, request):
        from datetime import datetime, timedelta, time

        month = request.query_params.get("month")
        year = request.query_params.get("year")

        if not month or not year:
            return Response(
                {"detail": "month and year are required. Example: ?month=6&year=2025"},
                status=400
            )

        try:
            month = int(month)
            year = int(year)
            start_date = datetime(year, month, 1).date()
            end_date = (
                datetime(year + 1, 1, 1).date()
                if month == 12
                else datetime(year, month + 1, 1).date()
            )
        except ValueError:
            return Response({"detail": "Invalid month or year"}, status=400)

        records = Attendance.objects.filter(
            date__gte=start_date,
            date__lt=end_date,
            check_in_time__isnull=False,
            check_out_time__isnull=False,
            shift__isnull=False
        ).select_related("employee", "shift")

        result = []

        for record in records:
            shift = record.shift

            if not shift.start_time or not shift.end_time:
                continue

            check_in_dt = datetime.combine(record.date, record.check_in_time)
            check_out_dt = datetime.combine(record.date, record.check_out_time)

            shift_start_dt = datetime.combine(record.date, shift.start_time)
            shift_end_dt = datetime.combine(record.date, shift.end_time)

            # 🔥 Handle overnight shifts
            if shift_end_dt <= shift_start_dt:
                shift_end_dt += timedelta(days=1)

            if check_out_dt <= check_in_dt:
                check_out_dt += timedelta(days=1)

            break_duration = shift.break_duration or timedelta()

            expected_work_duration = shift_end_dt - shift_start_dt - break_duration
            actual_work_duration = check_out_dt - check_in_dt

            late_check_in = check_in_dt > shift_start_dt
            early_check_out = check_out_dt < shift_end_dt
            less_working_hours = actual_work_duration < expected_work_duration

            if late_check_in or early_check_out or less_working_hours:
                result.append({
                    "employee_id": record.employee.id,
                    "employee_name": f"{record.employee.emp_first_name} {record.employee.emp_last_name}",
                    "date": record.date,
                    "check_in_time": record.check_in_time,
                    "check_out_time": record.check_out_time,
                    "shift_start_time": shift.start_time,
                    "shift_end_time": shift.end_time,
                    "expected_work_duration": str(expected_work_duration),
                    "actual_work_duration": str(actual_work_duration),
                    "late_check_in": late_check_in,
                    "early_check_out": early_check_out,
                    "less_working_hours": less_working_hours
                })

        return Response(result, status=200)
class AttendanceRecheckViewSet(viewsets.ModelViewSet):
    queryset = AttendanceRecheck.objects.all()
    serializer_class = AttendanceRecheckSerializer
    @staticmethod
    def parse_date(date_string):
        try:
            # Parse the date from string to a datetime.date object
            return datetime.strptime(date_string, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None
    @action(detail=False, methods=['post'])
    def recheck_location(self, request):
        emp_id = request.data.get("employee")
        date_str = request.data.get("date")

        lat = request.data.get("lat")
        lng = request.data.get("lng")
        location_name = request.data.get("location")

        if not all([emp_id, lat, lng, location_name]):
            return Response(
                {"detail": "Employee, latitude, longitude and location are required"},
                status=400
            )

        date = self.parse_date(date_str) if date_str else timezone.now().date()

        try:
            attendance = Attendance.objects.get(employee_id=emp_id, date=date)
        except Attendance.DoesNotExist:
            return Response(
                {"detail": "Attendance record not found"},
                status=404
            )

        # ❌ Recheck allowed only AFTER check-in and BEFORE check-out
        if not attendance.check_in_time:
            return Response(
                {"detail": "Employee has not checked in yet"},
                status=400
            )

        if attendance.check_out_time:
            return Response(
                {"detail": "Employee already checked out"},
                status=400
            )

        recheck = AttendanceRecheck.objects.create(
            attendance=attendance,
            lat=lat,
            lng=lng,
            location=location_name,
            # requested_by=request.user
        )

        return Response(
            {
                "status": "Recheck location captured successfully",
                "checked_at": recheck.checked_at,
                "location": recheck.location
            },
            status=200
        )
    
    def send_recheckin_email(self, attendance):

        employee = attendance.employee
        if not employee or not employee.emp_personal_email:
            print("D")
            return False

        config = EmailConfiguration.objects.filter(is_active=True).first()
        if not config:
            return False

        subject = "Attendance Check-in Confirmation"

        body = f"""
            Dear {employee.emp_first_name} {employee.emp_last_name},

            This is to inform you that your attendance has been marked for recheck.

            Regards,
            HR Team
            """

        EmailMessage(
            subject,
            body,
            config.email_host_user,
            [employee.emp_personal_email],
        ).send(fail_silently=False)

        return True
    @action(detail=False, methods=['post'], url_path='email')
    def send_email(self, request):
        emp_code = request.data.get("emp_code")

        if not emp_code:
            return Response(
                {"error": "emp_code is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        attendance = (
            Attendance.objects
            .select_related("employee")
            .filter(
                employee__emp_code__iexact=emp_code,
                check_in_time__isnull=False
            )
            .order_by("-date")
            .first()
        )

        if not attendance:
            return Response(
                {"error": "Employee has not checked in"},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            email_sent = self.send_recheckin_email(attendance)
        except Exception as e:
            return Response(
                {"error": f"Failed to send email: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        if not email_sent:
            return Response(
                {"error": "Employee email not available"},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response(
            {"message": "Attendance Recheck email sent successfully"},
            status=status.HTTP_200_OK
        )
class ImportAttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class= ImportAttendanceSerializer
    permission_classes = [AttendancePermission]
    resource_class = AttendanceResource
    parser_classes = (MultiPartParser, FormParser)

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        if 'file' not in request.FILES:
            return Response({"error": "Please provide a file."}, status=400)

        upload_file = request.FILES['file']
        filename = upload_file.name.lower()

        try:
            dataset = Dataset()
            resource = MonthlyAttendanceResource()

            # ---------- EXCEL ----------
            if filename.endswith('.xlsx'):
                dataset.load(upload_file.read(), format='xlsx')

            # ---------- CSV ----------
            elif filename.endswith('.csv'):
                file_data = upload_file.read().decode('utf-8')
                dataset.load(file_data, format='csv')

            else:
                return Response(
                    {"error": "Invalid file format. Only .xlsx and .csv are supported."},
                    status=400
                )

            # ---------- IMPORT ----------
            resource.import_data(dataset, dry_run=False, raise_errors=True)

            return Response({"message": "Monthly attendance imported successfully."})

        except Exception as e:
            return Response({"error": str(e)}, status=400)
        

    @action(detail=False, methods=['get'])
    def download_default_attendance_excel_file(self, request):
        resource = MonthlyAttendanceResource()
        headers = ["Identifier Code", "Year", "Month"]
        for day in range(1, 31):
            headers.append(f"{day}_In")
            headers.append(f"{day}_Out")

        wb = Workbook()

        black_font = Font(color="000000", bold=True)
        blue_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def style_header_row(ws, max_cols):
            for col in range(1, max_cols + 1):
                cell = ws.cell(row=1, column=col)
                if not cell.value:
                    cell.value = ""
                cell.fill = blue_fill
                cell.font = black_font
                cell.border = border_style
                ws.column_dimensions[cell.column_letter].width = 25
            ws.freeze_panes = "A2"

        # ======================================================
        # Sheet 1: Attendance
        # ======================================================
        ws1 = wb.active
        ws1.title = "Attendance"

        for col_num, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col_num, value=header)

        style_header_row(ws1, max_cols=len(headers))

        # ======================================================
        # Save response
        # ======================================================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            'attachment; filename="Attendance_BulkUpload_Template.xlsx"'
        )
        return response
    
    @action(detail=False, methods=['get'])
    def download_default_attendance_csv_file(self, request):
        resource = MonthlyAttendanceResource()
        headers = ["Identifier Code", "Year", "Month"]
        for day in range(1, 31):
            headers.append(f"{day}_In")
            headers.append(f"{day}_Out")

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = (
            'attachment; filename="Attendance_BulkUpload_Template.csv"'
        )
        return response

class Leave_ReportViewset(viewsets.ModelViewSet):
    queryset = LeaveReport.objects.all()
    serializer_class = LeaveReportSerializer
    permission_classes = [LeaveReportPermission] 


    def __init__(self, *args, **kwargs):
        super(Leave_ReportViewset, self).__init__(*args, **kwargs)
        self.leave_standard_report_exists()
    def get_available_fields(self):
        excluded_fields = {'id', 'created_by','created_at'}
        included_emp_master_fields = { 'emp_first_name', 'emp_dept_id', 'emp_desgntn_id', 'emp_ctgry_id','emp_branch_id'}
        
        display_names = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "emp_branch_id":"Branches",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "leave_type": "Leave Type",
            "reason": "Reason",
            "status":"Status",
            "approved_by": "Approved Request",
            "applied_on":"Request Date",
           
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name in included_emp_master_fields]
        leave_request_fields = [field.name for field in employee_leave_request._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + leave_request_fields}
        return available_fields

    @action(detail=False, methods=['get'])
    def select_leavereport_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
       
    @action(detail=False, methods=['post'])
    def generate_leave_report(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'report')
                fields_to_include = request.POST.getlist('fields', [])
                # from_date = parse_date(request.POST.get('from_date'))
                # to_date = parse_date(request.POST.get('to_date'))
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
            
            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())
            
            leavereport = employee_leave_request.objects.all()
            # documents = self.filter_documents_by_date_range(documents)

            report_data = self.generate_report_data(fields_to_include,leavereport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format


            LeaveReport.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({'status': 'success', 'file_path': file_path,'selected_fields_data': fields_to_include,})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

   
    def leave_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if LeaveReport.objects.filter(file_name='leave_std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'leave_std_report'
            fields_to_include = self.get_available_fields().keys()
            leavereport = employee_leave_request.objects.all()

            report_data = self.generate_report_data(fields_to_include, leavereport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            LeaveReport.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )

            print("Standard report generated successfully.")

        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = LeaveReport.objects.get(file_name='leave_std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except LeaveReport.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    def generate_report_data(self, fields_to_include,generalreport):
        column_headings = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "emp_branch_id":"Branches",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "leave_type": "Leave Type",
            "reason": "Reason",
            "status":"Status",
            "approved_by": "Approved Request",
            "applied_on":"Request Date",
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        leave_request_fields = [field.name for field in employee_leave_request._meta.get_fields() if isinstance(field, Field) and field.name != 'id']

        report_data = []
        for document in generalreport:
            general_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(document.employee, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()
                elif field in leave_request_fields:
                    value = getattr(document, field, 'N/A')
                else:
                    value = 'N/A'
                general_data[field] = value
            report_data.append(general_data)
        return report_data
    
    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])
        report_id = request.GET.get('report_id')  # Get report_id from query parameters

        
        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        })
    
    @action(detail=False, methods=['post'])
    def filter_by_date(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        # Replace slashes with hyphens
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')

        # Parse and validate the date range
        try:
            start_date = datetime.fromisoformat(start_date)
            end_date = datetime.fromisoformat(end_date)
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid date format: {str(e)}'}, status=400)

        # Fetch report data from your database
        try:
            report_instance = LeaveReport.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except LeaveReport.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)

        # Filter data by date range
        date_filtered_data = [
            row for row in report_data
            if 'applied_on' in row and row['applied_on'] and
            start_date <= datetime.fromisoformat(row['applied_on']) <= end_date
        ]

        # Save filtered data to Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        cache.set(cache_key, date_filtered_data, timeout=None)  # Set timeout as needed

        return JsonResponse({
            'date_filtered_data': date_filtered_data,
            'report_id': report_id,
        })
    
    @action(detail=False, methods=['post'])
    def generate_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.data.get('report_id')

        # Save selected fields to session
        request.session['selected_fields'] = selected_fields

        # Fetch date-filtered report data from session
        date_filtered_data = request.session.get('date_filtered_data', [])
        print("previously date filtered ",date_filtered_data)
        
        # If no date-filtered data, attempt to fetch full report
        if not date_filtered_data:
            try:
                report = LeaveReport.objects.get(id=report_id)
                report_file_path = os.path.join(settings.MEDIA_ROOT, report.report_data.name)
                with open(report_file_path, 'r') as file:
                    report_content = json.load(file)
            except LeaveReport.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Report not found'})

            date_filtered_data = report_content

        # If no fields are selected for filtration, default to all existing fields in the report
        if not selected_fields:
            if date_filtered_data:
                selected_fields = list(date_filtered_data[0].keys())  # Default to all keys in the first record
            else:
                selected_fields = []  # No data in the report
            
        # Define display names for fields
        column_headings = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "emp_branch_id":"Branches",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "leave_type": "Leave Type",
            "reason": "Reason",
            "status":"Status",
            "approved_by": "Approved Request",
            "applied_on":"Request Date",
        }

        # Get unique values for selected_fields from date-filtered data
        unique_values = self.get_unique_values_for_fields(date_filtered_data, selected_fields)

        processed_unique_values = {}
        for field, values in unique_values.items():
            processed_unique_values[field] = {
                'values': values,
            }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': date_filtered_data,  # Pass filtered data to the frontend
            'unique_values': processed_unique_values,
            'column_headings':column_headings
        })
        

    def get_unique_values_for_fields(self, data, selected_fields):
        unique_values = {field: set() for field in selected_fields}
        # Extract data from the provided content
        for record in data:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        # Convert sets to lists
        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values
       
    
    @action(detail=False, methods=['post'])
    def general_filter_report(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')

        # Retrieve filtered data from Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        filtered_data = cache.get(cache_key)

        if filtered_data is None:
            return JsonResponse({'status': 'error', 'message': 'No date-filtered data available'}, status=404)
        # Apply additional filtering here if needed
        # For example, based on other fields:
        additional_filters = {key: value for key, value in request.data.items() if key not in ('report_id',)}
        
        # Further filter based on additional criteria
        filtered_data = [
            row for row in filtered_data
            if all(row.get(key) == value for key, value in additional_filters.items())
        ]
        return JsonResponse({
            'filtered_data': filtered_data,
            'report_id': report_id,
        })

    
class LvApprovalLevelViewset(viewsets.ModelViewSet):
    queryset= LVApprovalWorkflow.objects.all()
    serializer_class= LVApprovalWorkflowSerializer
    permission_classes = [LeaveApprovalLevelsPermission]


class LvCommonWorkflowViewset(viewsets.ModelViewSet):
    queryset=LvCommonWorkflow.objects.all()
    serializer_class=LvCommonWorkflowSerializer
    permission_classes = [LvCommonWorkflowPermission]

class LvRejectionViewset(viewsets.ModelViewSet):
    queryset=LvRejectionReason.objects.all()
    serializer_class=LvRejectionReasonSerializer
    permission_classes = [LvRejectionReasonPermission]

class LvApprovalViewset(viewsets.ModelViewSet):
    queryset=LeaveApproval.objects.all()
    serializer_class=LvApprovalSerializer
    lookup_field = 'pk'
    def get_queryset(self):
        import json

        queryset = LeaveApproval.objects.select_related(
            'leave_request',
            'leave_request__employee',
            'approver'
        ).filter(leave_request__isnull=False)

        branch_ids = self.request.query_params.get('branch_id')

        if branch_ids:
            try:
                branch_ids = json.loads(branch_ids)

                if not isinstance(branch_ids, list):
                    branch_ids = [branch_ids]

            except Exception:
                branch_ids = branch_ids.strip("[]").split(",")

            branch_ids = [
                int(i) for i in branch_ids if str(i).strip().isdigit()
            ]

            if branch_ids:
                queryset = queryset.filter(
                   leave_request__employee__emp_branch_id__in=branch_ids  # ✅ FIXED
                )

        return queryset
                
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approval = self.get_object()

        if not request.user.is_superuser and approval.approver != request.user:
            return Response(
                {"error": "You are not allowed to approve this request."},
                status=status.HTTP_403_FORBIDDEN
            )

        note = request.data.get('note')
        approved_days = request.data.get('approved_days')

        if approved_days not in [None, ""]:
            try:
                approved_days = float(approved_days)
            except ValueError:
                return Response(
                    {"error": "Invalid value for approved_days"},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            approved_days = None

        try:
            approval.approve(note=note, approved_days=approved_days)
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'status': 'approved',
            'note': note,
            'approved_days': approved_days
        }, status=status.HTTP_200_OK)

    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()
        if not request.user.is_superuser and approval.approver != request.user:
            return Response(
                {"error": "You are not allowed to reject this request."},
                status=status.HTTP_403_FORBIDDEN
            )

        note = request.data.get('note')
        rejection_reason = request.data.get('rejection_reason')

        if not rejection_reason:
            raise ValidationError("Rejection reason is required.")

        approval.reject(rejection_reason=rejection_reason, note=note)

        return Response(
            {
                'status': 'rejected',
                'note': note,
                'rejection_reason': rejection_reason
            },
            status=status.HTTP_200_OK
        )
    @action(detail=False, methods=['get'])
    def grouped_approvals(self, request):
        approvals = LeaveApproval.objects.select_related('leave_request', 'compensatory_request', 'approver').order_by('leave_request', 'level')

        # Group approvals by leave_request or compensatory_request
        grouped_approvals = defaultdict(list)
        for approval in approvals:
            if approval.leave_request:
                request_id = f"LeaveRequest-{approval.leave_request.id}"
            elif approval.compensatory_request:
                request_id = f"CompensatoryRequest-{approval.compensatory_request.id}"
            else:
                # Skip approvals without any associated request
                continue

            grouped_approvals[request_id].append({
                'id': approval.id,
                'role': approval.role,
                'level': approval.level,
                'status': approval.status,
                'note': approval.note,
                'created_at': approval.created_at,
                'updated_at': approval.updated_at,
                'approver': approval.approver.username,
                'rejection_reason': approval.rejection_reason.reason_text if approval.rejection_reason else None,
            })

        response_data = [
            {
                'request_id': request_id,
                'approvals': levels
            }
            for request_id, levels in grouped_approvals.items()
        ]

        return Response(response_data, status=status.HTTP_200_OK)
    
class Lv_Approval_ReportViewset(viewsets.ModelViewSet):
    queryset = LeaveApprovalReport.objects.all()
    serializer_class = LvApprovalReportSerializer
    permission_classes = [LeaveApprovalReportPermission]
    
    def __init__(self, *args, **kwargs):
        super(Lv_Approval_ReportViewset, self).__init__(*args, **kwargs)
        self.lv_apprvl_std_report_exists()
    def get_available_fields(self):
        excluded_fields = {'id', 'created_by','created_at'}
        included_emp_master_fields = { 'emp_first_name', 'emp_dept_id', 'emp_desgntn_id', 'emp_ctgry_id','emp_branch_id'}
        
        display_names = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "emp_branch_id":"Branches",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "leave_request": "Leave Request",
            "approver":"Approver",
            "level":"Level",
            "created_at": "Approve/Reject Date",
            "status":"Status",
            "note": "Comments",
            "rejection_reason":"Rejection Reason",
           
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name in included_emp_master_fields]
        leave_approval_fields = [field.name for field in LeaveApproval._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + leave_approval_fields}
        return available_fields

    @action(detail=False, methods=['get'])
    def select_approve_report_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
       
    @action(detail=False, methods=['post'])
    def generate_leave_report(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'report')
                fields_to_include = request.POST.getlist('fields', [])
                # from_date = parse_date(request.POST.get('from_date'))
                # to_date = parse_date(request.POST.get('to_date'))
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
            
            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())
            
            leavereport = LeaveApproval.objects.all()
            # documents = self.filter_documents_by_date_range(documents)

            report_data = self.generate_report_data(fields_to_include,leavereport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format


            LeaveApprovalReport.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({'status': 'success', 'file_path': file_path,'selected_fields_data': fields_to_include,})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

   
    def lv_apprvl_std_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if LeaveReport.objects.filter(file_name='lv_approv_std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'lv_approv_std_report'
            fields_to_include = self.get_available_fields().keys()
            leavereport = LeaveApproval.objects.all()

            report_data = self.generate_report_data(fields_to_include, leavereport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            LeaveApprovalReport.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )

            print("Standard report generated successfully.")

        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = LeaveApprovalReport.objects.get(file_name='lv_approv_std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except LeaveApprovalReport.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    def generate_report_data(self, fields_to_include, generalreport):
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        leave_approval_fields = [field.name for field in LeaveApproval._meta.get_fields() if isinstance(field, Field) and field.name != 'id']

        report_data = {}

        for document in generalreport:
            leave_request = document.leave_request
            compensatory_request = document.compensatory_request
            request_id = leave_request.id if leave_request else (compensatory_request.id if compensatory_request else 'N/A')

            # Shared information container (one per request_id)
            if request_id not in report_data:
                shared_data = {
                    "request_id": request_id,
                    "approvals": []
                }

                employee = leave_request.employee if leave_request else (compensatory_request.employee if compensatory_request else None)

                for field in fields_to_include:
                    if field in emp_master_fields and employee:
                        shared_data[field] = getattr(employee, field, 'N/A')
                    elif field == 'leave_request' and leave_request:
                        shared_data['leave_request'] = str(leave_request)
                    elif field == 'compensatory_request' and compensatory_request:
                        shared_data['compensatory_request'] = str(compensatory_request)
                
                report_data[request_id] = shared_data

            # Approval-specific info
            approval_data = {}
            approver = document.approver if document.approver else None

            for field in fields_to_include:
                if field in leave_approval_fields:
                    approval_data[field] = getattr(document, field, 'N/A')
                elif field == 'approver' and approver:
                    approval_data['approver'] = approver.username
                elif field == 'rejection_reason' and document.rejection_reason:
                    approval_data['rejection_reason'] = document.rejection_reason.reason_text

            report_data[request_id]["approvals"].append(approval_data)

        return list(report_data.values())
    
    
    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])
        report_id = request.GET.get('report_id')  # Get report_id from query parameters      
        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        })
    
    @action(detail=False, methods=['post'])
    def filter_by_date(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        # Replace slashes with hyphens
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')

        # Parse and validate the date range
        try:
            start_date = datetime.fromisoformat(start_date)
            end_date = datetime.fromisoformat(end_date)
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid date format: {str(e)}'}, status=400)

        # Fetch report data from your database
        try:
            report_instance = LeaveApprovalReport.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except LeaveApprovalReport.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)

        # Filter data by date range
        date_filtered_data = [
            row for row in report_data
            if 'created_at' in row and row['created_at'] and
            start_date <= datetime.fromisoformat(row['created_at']) <= end_date
        ]

        # Save filtered data to Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        cache.set(cache_key, date_filtered_data, timeout=None)  # Set timeout as needed

        return JsonResponse({
            'date_filtered_data': date_filtered_data,
            'report_id': report_id,
        })
    

    @action(detail=False, methods=['post'])
    def approval_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.data.get('report_id')

        # Save selected fields to session
        request.session['selected_fields'] = selected_fields

        # Fetch date-filtered report data from session
        date_filtered_data = request.session.get('date_filtered_data', [])
        print("previosly date filtered ",date_filtered_data)
        
        # If no date-filtered data, attempt to fetch full report
        if not date_filtered_data:
            try:
                report = LeaveApprovalReport.objects.get(id=report_id)
                report_file_path = os.path.join(settings.MEDIA_ROOT, report.report_data.name)
                with open(report_file_path, 'r') as file:
                    report_content = json.load(file)
            except LeaveApprovalReport.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Report not found'})

            date_filtered_data = report_content

        # If no fields are selected for filtration, default to all existing fields in the report
        if not selected_fields:
            if date_filtered_data:
                selected_fields = list(date_filtered_data[0].keys())  # Default to all keys in the first record
            else:
                selected_fields = []  # No data in the report
        # Get unique values for selected_fields from date-filtered data
        unique_values = self.get_unique_values_for_fields(date_filtered_data, selected_fields)

        processed_unique_values = {}
        for field, values in unique_values.items():
            processed_unique_values[field] = {
                'values': values,
            }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': date_filtered_data,  # Pass filtered data to the frontend
            'unique_values': processed_unique_values,
        })

    def get_unique_values_for_fields(self, data, selected_fields):
        unique_values = {field: set() for field in selected_fields}
        # Extract data from the provided content
        for record in data:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        # Convert sets to lists
        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values

    @action(detail=False, methods=['post'])
    def approval_filter_report(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')

        # Retrieve filtered data from Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        filtered_data = cache.get(cache_key)

        if filtered_data is None:
            return JsonResponse({'status': 'error', 'message': 'No date-filtered data available'}, status=404)

        # Apply additional filtering here if needed
        # For example, based on other fields:
        additional_filters = {key: value for key, value in request.data.items() if key not in ('report_id',)}
        
        # Further filter based on additional criteria
        filtered_data = [
            row for row in filtered_data
            if all(row.get(key) == value for key, value in additional_filters.items())
        ]

        return JsonResponse({
            'filtered_data': filtered_data,
            'report_id': report_id,
        })

class AttendanceReportViewset(viewsets.ModelViewSet):
    queryset = AttendanceReport.objects.all()
    serializer_class = AttendanceReportSerializer
    permission_classes = [AttendanceReportPermission]

    
    
    def __init__(self, *args, **kwargs):
        super(AttendanceReportViewset, self).__init__(*args, **kwargs)
        self.attendance_standard_report_exists()
    def get_available_fields(self):
        excluded_fields = {'id', 'created_by'}
        included_emp_master_fields = { 'emp_first_name', 'emp_dept_id', 'emp_desgntn_id', 'emp_ctgry_id'}
        
        display_names = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "emp_branch_id":"Branches",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "shift": "Shift",
            "date": "Date",
            "check_in_time":"Check In",
            "check_out_time": "Check Out",
            "total_hours":"Total Hours",
           
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name in included_emp_master_fields]
        attendance_fields = [field.name for field in Attendance._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + attendance_fields}
        return available_fields

    @action(detail=False, methods=['get'])
    def select_attendancereport_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
       
    @action(detail=False, methods=['post'])
    def generate_leave_report(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'report')
                fields_to_include = request.POST.getlist('fields', [])
                # from_date = parse_date(request.POST.get('from_date'))
                # to_date = parse_date(request.POST.get('to_date'))
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
            
            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())
            
            attendancereport = Attendance.objects.all()
            # documents = self.filter_documents_by_date_range(documents)

            report_data = self.generate_report_data(fields_to_include,attendancereport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format


            AttendanceReport.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({'status': 'success', 'file_path': file_path,'selected_fields_data': fields_to_include,})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

   
    def attendance_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if AttendanceReport.objects.filter(file_name='attendance_std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'attendance_std_report'
            fields_to_include = self.get_available_fields().keys()
            attendancereport = Attendance.objects.all()

            report_data = self.generate_report_data(fields_to_include, attendancereport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            AttendanceReport.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )

            print("Standard report generated successfully.")

        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = AttendanceReport.objects.get(file_name='attendance_std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except AttendanceReport.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    def generate_report_data(self, fields_to_include,generalreport):
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        leave_request_fields = [field.name for field in Attendance._meta.get_fields() if isinstance(field, Field) and field.name != 'id']

        report_data = []
        for document in generalreport:
            general_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(document.employee, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()
                elif field in leave_request_fields:
                    value = getattr(document, field, 'N/A')
                else:
                    value = 'N/A'
                general_data[field] = value
            report_data.append(general_data)
        return report_data
   
    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])
        report_id = request.GET.get('report_id')
        
        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        })
    
    @action(detail=False, methods=['post'])
    def filter_by_date(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')

        # Replace slashes with hyphens
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')

        # Parse and validate the date range
        try:
            start_date = datetime.fromisoformat(start_date)
            end_date = datetime.fromisoformat(end_date)
        except ValueError as e:
            return JsonResponse({'status': 'error', 'message': f'Invalid date format: {str(e)}'}, status=400)

        # Fetch report data from your database
        try:
            report_instance = AttendanceReport.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except AttendanceReport.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'}, status=404)

        # Filter data by date range
        date_filtered_data = [
            row for row in report_data
            if 'date' in row and row['date'] and
            start_date <= datetime.fromisoformat(row['date']) <= end_date
        ]

        # Save filtered data to Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        cache.set(cache_key, date_filtered_data, timeout=None)  # Set timeout as needed

        return JsonResponse({
            'date_filtered_data': date_filtered_data,
            'report_id': report_id,
        })
    

    @action(detail=False, methods=['post'])
    def attendance_filter_table(self, request, *args, **kwargs):
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.data.get('report_id')

        # Save selected fields to session
        request.session['selected_fields'] = selected_fields

        # Fetch date-filtered report data from session
        date_filtered_data = request.session.get('date_filtered_data', [])
        print("previosly date filtered ",date_filtered_data)
        
        # If no date-filtered data, attempt to fetch full report
        if not date_filtered_data:
            try:
                report = AttendanceReport.objects.get(id=report_id)
                report_file_path = os.path.join(settings.MEDIA_ROOT, report.report_data.name)
                with open(report_file_path, 'r') as file:
                    report_content = json.load(file)
            except AttendanceReport.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Report not found'})

            date_filtered_data = report_content

        # If no fields are selected for filtration, default to all existing fields in the report
        if not selected_fields:
            if date_filtered_data:
                selected_fields = list(date_filtered_data[0].keys())  # Default to all keys in the first record
            else:
                selected_fields = []  # No data in the report
        # Get unique values for selected_fields from date-filtered data
        unique_values = self.get_unique_values_for_fields(date_filtered_data, selected_fields)

        processed_unique_values = {}
        for field, values in unique_values.items():
            processed_unique_values[field] = {
                'values': values,
            }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': date_filtered_data,  # Pass filtered data to the frontend
            'unique_values': processed_unique_values,
        })

    def get_unique_values_for_fields(self, data, selected_fields):
        unique_values = {field: set() for field in selected_fields}
        # Extract data from the provided content
        for record in data:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        # Convert sets to lists
        for field in unique_values:
            unique_values[field] = list(unique_values[field])
        return unique_values

    @action(detail=False, methods=['post'])
    def approval_filter_report(self, request, *args, **kwargs):
        tenant_id = request.tenant.schema_name
        report_id = request.data.get('report_id')

        # Retrieve filtered data from Redis cache
        cache_key = f"{tenant_id}_{report_id}_date_filtered_data"
        filtered_data = cache.get(cache_key)

        if filtered_data is None:
            return JsonResponse({'status': 'error', 'message': 'No date-filtered data available'}, status=404)

        # Apply additional filtering here if needed
        # For example, based on other fields:
        additional_filters = {key: value for key, value in request.data.items() if key not in ('report_id',)}
        
        # Further filter based on additional criteria
        filtered_data = [
            row for row in filtered_data
            if all(row.get(key) == value for key, value in additional_filters.items())
        ]

        return JsonResponse({
            'filtered_data': filtered_data,
            'report_id': report_id,
        })

class LvBalanceReportViewset(viewsets.ModelViewSet):
    queryset = lvBalanceReport.objects.all()
    serializer_class = lvBalanceReportSerializer
    permission_classes = [LvBalanceReportPermission]

    

    def __init__(self, *args, **kwargs):
        super(LvBalanceReportViewset, self).__init__(*args, **kwargs)
        self.lvbalance_standard_report_exists()
    def get_available_fields(self):
        excluded_fields = {'id', 'created_by','created_at'}
        included_emp_master_fields = { 'emp_first_name', 'emp_dept_id', 'emp_desgntn_id', 'emp_ctgry_id','emp_branch_id'}
        
        display_names = {
            "employee": "Employee Code",
            "emp_first_name": "First Name",
            "emp_branch_id":"Branches",
            "emp_dept_id": "Department",
            "emp_desgntn_id": "Designation",
            "emp_ctgry_id": "Category",
            "leave_type": "Leave Type",
            "balance": "Balance",
            "openings":"Openings",
            
           
        }

        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name in included_emp_master_fields]
        leave_balance = [field.name for field in emp_leave_balance._meta.get_fields() if isinstance(field, Field) and field.name not in excluded_fields]
        
        available_fields = {field: display_names.get(field, field) for field in emp_master_fields + leave_balance}
        return available_fields

    @action(detail=False, methods=['get'])
    def select_attendancereport_fields(self, request, *args, **kwargs):
        available_fields = self.get_available_fields()
        return Response({'available_fields': available_fields})
       
    @action(detail=False, methods=['post'])
    def generate_leave_report(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                file_name = request.POST.get('file_name', 'report')
                fields_to_include = request.POST.getlist('fields', [])
                # from_date = parse_date(request.POST.get('from_date'))
                # to_date = parse_date(request.POST.get('to_date'))
            except Exception as e:
                return JsonResponse({'status': 'error', 'message': str(e)})
            
            if not fields_to_include:
                fields_to_include = list(self.get_available_fields().keys())
            
            attendancereport = emp_leave_balance.objects.all()
            # documents = self.filter_documents_by_date_range(documents)

            report_data = self.generate_report_data(fields_to_include,attendancereport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)  # Serialize dates to string format


            lvBalanceReport.objects.create(file_name=file_name, report_data=file_name + '.json')
            return JsonResponse({'status': 'success', 'file_path': file_path,'selected_fields_data': fields_to_include,})
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

   
    def lvbalance_standard_report_exists(self):
        # Update the standard report if it exists, otherwise create a new one
        if lvBalanceReport.objects.filter(file_name='lvbalance_std_report').exists():
            self.generate_standard_report()
        else:
            self.generate_standard_report()
    
    def generate_standard_report(self):
        try:
            file_name = 'lvbalance_std_report'
            fields_to_include = self.get_available_fields().keys()
            lvbalancereport = emp_leave_balance.objects.all()

            report_data = self.generate_report_data(fields_to_include, lvbalancereport)
            file_path = os.path.join(settings.MEDIA_ROOT, file_name + '.json')

            # Save report data to a file
            with open(file_path, 'w') as file:
                json.dump(report_data, file, default=str)

            # Update or create the standard report entry in the database
            lvBalanceReport.objects.update_or_create(
                file_name=file_name,
                defaults={'report_data': file_name + '.json'}
            )

            print("Standard report generated successfully.")

        except Exception as e:
            print(f"Error generating standard report: {str(e)}")

    @action(detail=False, methods=['get'])
    def std_report(self, request, *args, **kwargs):
        try:
            # Ensure the standard report is up-to-date
            self.generate_standard_report()
            report = lvBalanceReport.objects.get(file_name='lvbalance_std_report')
            serializer = self.get_serializer(report)
            return Response(serializer.data)
        except lvBalanceReport.DoesNotExist:
            return Response({"error": "Standard report not found."}, status=status.HTTP_404_NOT_FOUND)
    
    def generate_report_data(self, fields_to_include,generalreport):
        emp_master_fields = [field.name for field in emp_master._meta.get_fields() if isinstance(field, Field) and field.name != 'id']
        leave_request_fields = [field.name for field in emp_leave_balance._meta.get_fields() if isinstance(field, Field) and field.name != 'id']

        report_data = []
        for document in generalreport:
            general_data = {}
            for field in fields_to_include:
                if field in emp_master_fields:
                    value = getattr(document.employee, field, 'N/A')
                    if isinstance(value, date):
                        value = value.isoformat()
                elif field in leave_request_fields:
                    value = getattr(document, field, 'N/A')
                else:
                    value = 'N/A'
                general_data[field] = value
            report_data.append(general_data)
        return report_data
    
    @action(detail=False, methods=['get'])
    def select_filter_fields(self, request, *args, **kwargs):
       
        available_fields = self.get_available_fields()
        selected_fields = request.session.get('selected_fields', [])
        report_id = request.GET.get('report_id')

        return Response({
            'available_fields': available_fields,
            'selected_fields': selected_fields,
            'report_id': report_id
        })

    @csrf_exempt
    @action(detail=False, methods=['post'])
    def generate_balance_filter_table(self, request, *args, **kwargs):
        
        selected_fields = request.POST.getlist('selected_fields')
        report_id = request.POST.get('report_id')

        try:
            report = lvBalanceReport.objects.get(id=report_id)
            report_file_path = os.path.join(settings.MEDIA_ROOT, report.report_data.name)
            with open(report_file_path, 'r') as file:
                report_content = json.load(file)
        except lvBalanceReport.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Report not found'})

        if not selected_fields:
            selected_fields = list(report_content[0].keys()) if report_content else []

        balances = emp_leave_balance.objects.all()
        unique_values = self.get_unique_values_for_fields(balances, selected_fields, report_content)

        processed_unique_values = {
            field: {'values': values} for field, values in unique_values.items()
        }

        return JsonResponse({
            'selected_fields': selected_fields,
            'report_id': report_id,
            'report_content': report_content,
            'unique_values': processed_unique_values,
        })

    def get_unique_values_for_fields(self, balances, selected_fields, report_content):
        """
        Extract unique values for the selected fields.
        """
        unique_values = {field: set() for field in selected_fields}

        for record in report_content:
            for field in selected_fields:
                if field in record:
                    unique_values[field].add(record[field])

        for balance in balances:
            for field in selected_fields:
                if hasattr(balance, field):
                    value = getattr(balance, field, None)
                    if value is not None:
                        unique_values[field].add(value)

        return {field: list(values) for field, values in unique_values.items()}

    @csrf_exempt
    @action(detail=False, methods=['post'])
    def filter_existing_report(self, request, *args, **kwargs):
        
        report_id = request.data.get('report_id')
        if not report_id:
            return HttpResponse('Report ID is missing', status=400)

        try:
            report_instance = lvBalanceReport.objects.get(id=report_id)
            report_data = json.loads(report_instance.report_data.read().decode('utf-8'))
        except (lvBalanceReport.DoesNotExist, json.JSONDecodeError) as e:
            return HttpResponse(f'Report not found or invalid JSON: {str(e)}', status=404)

        selected_fields = [key for key in request.data.keys() if key != 'report_id']
        filter_criteria = {field: request.data.getlist(field) for field in selected_fields if request.data.getlist(field)}

        filtered_data = [
            row for row in report_data
            if self.match_filter_criteria(row, filter_criteria)
        ]

        request.session['filtered_data'] = filtered_data
        request.session.modified = True

        return JsonResponse({
            'filtered_data': filtered_data,
            'report_id': report_id,
        })

    def match_filter_criteria(self, row_data, filter_criteria):
        
        for field, values in filter_criteria.items():
            row_value = row_data.get(field)
            if row_value is None or str(row_value).strip() not in values:
                return False
        return True

class CompensatoryLeaveRequestviewset(viewsets.ModelViewSet):
    queryset = CompensatoryLeaveRequest.objects.all()
    serializer_class = CompensatoryLeaveRequestSerializer
    permission_classes = [CompensatoryLeaveRequestPermission]

class CompensatoryLeaveBalancetviewset(viewsets.ModelViewSet):
    queryset = CompensatoryLeaveBalance.objects.all()
    serializer_class = CompensatoryLeaveBalanceSerializer
    permission_classes = [CompensatoryLeaveBalancePermission]

class CompensatoryLeaveTransactionviewset(viewsets.ModelViewSet):
    queryset = CompensatoryLeaveTransaction.objects.all()
    serializer_class = CompensatoryLeaveTransactionSerializer 
    permission_classes = [CompensatoryLeaveTransactionPermission]

class EmployeeYearlyCalendarViewset(viewsets.ModelViewSet):
    queryset = EmployeeYearlyCalendar.objects.all()
    serializer_class = EmployeeYearlyCalendarSerializer
    permission_classes = [EmployeeYearlyCalendarPermission]

class EmpOpeningsBlkupldViewSet(viewsets.ModelViewSet):
    queryset = emp_leave_balance.objects.all()
    serializer_class = EmpOpeningsBlkupldSerializer
    
    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        if 'file' not in request.FILES:
            return Response({"error": "Please provide a file."}, status=400)

        upload_file = request.FILES['file']
        file_name = upload_file.name.lower()

        try:
            if file_name.endswith('.xlsx'):
                dataset = XLSX().create_dataset(upload_file.read())

            elif file_name.endswith('.csv'):
                dataset = CSV().create_dataset(
                    upload_file.read().decode('utf-8')
                )

            else:
                return Response(
                    {"error": "Invalid file format. Only .xlsx and .csv are supported."},
                    status=400
                )

            resource = EmployeeOpenBalanceResource()
            all_errors = []

            with transaction.atomic():
                for row_idx, row in enumerate(dataset.dict, start=2):
                    try:
                        resource.before_import_row(row, row_idx=row_idx)
                        resource.import_row(row, None)
                    except ValidationError as e:
                        all_errors.extend(
                            [f"Row {row_idx}: {msg}" for msg in e.messages]
                        )

            if all_errors:
                return Response({"errors": all_errors}, status=400)

            return Response({"message": "Records updated successfully."})

        except Exception as e:
            return Response({"error": str(e)}, status=400)
    @action(detail=False, methods=['get'])
    def download_default_excel_file(self, request):
        resource =EmployeeOpenBalanceResource()
        headers = [field.column_name for field in resource.fields.values()]
        wb = Workbook()

        # ======== Common Styles ========
        black_font = Font(color="000000", bold=True)
        blue_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")  # light cream/yellow
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Helper function to style header row
        def style_header_row(ws, max_cols=10):
            """Style header row with blue fill and black bold text across full width."""
            for col in range(1, max_cols + 1):
                cell = ws.cell(row=1, column=col)
                if not cell.value:
                    cell.value = ""
                cell.fill = blue_fill
                cell.font = black_font
                cell.border = border_style
                ws.column_dimensions[cell.column_letter].width = 25
            ws.freeze_panes = "A2"  # freeze header
        # ======================================================
        # Sheet 1: SalaryComponent
        # ======================================================
        ws1 = wb.active
        ws1.title = "Leave Balance Openings"
        for col_num, header in enumerate(headers, 1):
            ws1.cell(row=1, column=col_num, value=header)

        style_header_row(ws1, max_cols=len(headers))
         # ======================================================
        # Save response
        # ======================================================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            'attachment; filename="Leavebalance_BulkUpload_Template.xlsx"'
        )
        return response
    
    @action(detail=False, methods=['get'])
    def download_default_csv_file(self, request):
        resource = EmployeeOpenBalanceResource()
        headers = [field.column_name for field in resource.fields.values()]
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)  # only headers, no data

        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="Leavebalance_Template.csv"'
        return response

class ApplyOpeningsAPIView(APIView):
    def post(self, request, *args, **kwargs):
        leave_balance_id = request.data.get('leave_balance_id')
        openings = request.data.get('openings')

        if leave_balance_id is None or openings is None:
            return Response({'error': 'Missing required fields.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            leave_balance = emp_leave_balance.objects.get(id=leave_balance_id)
        except emp_leave_balance.DoesNotExist:
            return Response({'error': 'Leave balance not found for given ID.'}, status=status.HTTP_404_NOT_FOUND)

        # Update the openings and balance
        leave_balance.openings = openings
        leave_balance.balance = (leave_balance.balance or 0) + openings  # in case balance is None
        leave_balance.save()

        return Response({'success': 'Openings applied successfully.'}, status=status.HTTP_200_OK)
class EmployeeRejoiningViewset(viewsets.ModelViewSet):
    queryset = EmployeeRejoining.objects.all()
    serializer_class = EmployeeRejoiningSerializer
    @action(detail=True, methods=['post'], url_path='deduct-leave-balance')
    def deduct_leave_balance(self, request, pk=None):
        """
        Custom action to choose from which leave type unpaid days should be deducted.
        """
        try:
            rejoining = self.get_object()
            deduct_from_leave_type = request.data.get('deduct_from_leave_type')

            if not deduct_from_leave_type:
                return Response({"error": "leave_type_id is required."}, status=status.HTTP_400_BAD_REQUEST)

            if rejoining.deducted:
                return Response({"error": "Leave balance already deducted for this rejoining."}, status=status.HTTP_400_BAD_REQUEST)

            # Get employee leave balance
            leave_balance, _ = emp_leave_balance.objects.get_or_create(
                employee=rejoining.employee,
                leave_type_id=deduct_from_leave_type
            )

            unpaid_days = rejoining.unpaid_leave_days

            if leave_balance.balance < unpaid_days:
                return Response({"error": "Not enough balance in the selected leave type."}, status=status.HTTP_400_BAD_REQUEST)

            # Deduct unpaid days
            old_balance = leave_balance.balance
            leave_balance.balance -= unpaid_days
            leave_balance.save()

            # Update rejoining record
            rejoining.deduct_from_leave_type_id = deduct_from_leave_type
            rejoining.deducted = True
            rejoining.save()

            return Response({
                "message": "Leave balance deducted successfully.",
                "employee": rejoining.employee.emp_first_name,
                "unpaid_days": unpaid_days,
                "old_balance": old_balance,
                "new_balance": leave_balance.balance
            }, status=status.HTTP_200_OK)

        except EmployeeRejoining.DoesNotExist:
            return Response({"error": "Rejoining record not found."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
class ImmediateRejectAPIView(APIView):
    """
    API for users with special permissions to immediately reject an approved leave request by document number.
    """
    # permission_classes = [IsAuthenticated]  # Later you can add custom permission here

    def post(self, request, *args, **kwargs):
        document_number = request.data.get('document_number')
        rejection_reason = request.data.get('rejection_reason')
        note = request.data.get('note')

        if not document_number:
            return Response({'error': 'document_number is required.'}, status=status.HTTP_400_BAD_REQUEST)

        if not rejection_reason:
            return Response({'error': 'rejection_reason is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            leave_request = employee_leave_request.objects.get(document_number=document_number)
        except employee_leave_request.DoesNotExist:
            return Response({'error': 'Leave request not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Find the approved LeaveApproval for this leave request
        approval = LeaveApproval.objects.filter(
            leave_request=leave_request,
            status=LeaveApproval.APPROVED
        ).first()

        if not approval:
            return Response({'error': 'No approved leave request found for this document number.'}, status=status.HTTP_400_BAD_REQUEST)

        # # Optional: Check if user has permission to immediately reject
        # if not request.user.has_perm('yourapp.immediate_reject_leave'):
        #     raise PermissionDenied('You do not have permission to immediately reject leave requests.')

        # Perform rejection
        approval.status = LeaveApproval.REJECTED
        approval.rejection_reason = rejection_reason
        approval.note = note
        approval.save()

        # Also change the employee_leave_request status to REJECTED
        # leave_request.status = employee_leave_request.status='rejected'
        leave_request.status = 'rejected'
        leave_request.save()

        # Restore leave balance
        leave_request.restore_leave_balance()

        return Response(
            {
                'status': 'rejected_immediately',
                'document_number': document_number,
                'leave_approval_id': approval.id,
                'leave_request_id': leave_request.id
            },
            status=status.HTTP_200_OK
        )
class EmployeeAttendanceSummaryAPIView(APIView):
    def get(self, request):
        month = int(request.query_params.get("month", date.today().month))
        year = int(request.query_params.get("year", date.today().year))

        start_date = date(year, month, 1)
        end_date = start_date + relativedelta(months=1) - relativedelta(days=1)

        summaries = []

        # You can apply filters here if needed (e.g., by branch, department, etc.)
        all_employees = emp_master.objects.all()

        # for employee in all_employees:
        #     summary_data = get_attendance_summary(employee, start_date, end_date)
        #     serializer = AttendanceSummarySerializer(summary_data)

        #     # Include employee info with the summary
        #     summaries.append({
        #         # "employee_id": employee.id,
        #         "employee_name": f"{employee.emp_code}",
        #         "attendance_summary": serializer.data
        #     })
        #     return Response(serializer.data)
        for employee in all_employees:
            summary_data = get_attendance_summary(employee, start_date, end_date)
            
            if summary_data is None:
                continue  # or handle the case differently

            serializer = AttendanceSummarySerializer(summary_data)

            summaries.append({
                "employee_id": employee.id,
                "employee_name": f"{employee.emp_first_name}",
                "attendance_summary": serializer.data
            })
        return Response(summaries)

class MonthwiseAccrualSimulationView(APIView):
    def get(self, request, format=None):
        employee_id = request.query_params.get("employee_id")
        year = int(request.query_params.get("year", datetime.now().year))

        if not employee_id:
            return Response({"error": "employee_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = emp_master.objects.get(id=employee_id)
        except emp_master.DoesNotExist:
            return Response({"error": "Employee not found"}, status=status.HTTP_404_NOT_FOUND)

        today = timezone.now().date()
        start_date = datetime(year, 1, 1).date()

        entitlements = leave_entitlement.objects.filter(accrual=True).order_by("leave_type", "min_experience")
        all_leave_types = leave_type.objects.all()

        # Select the best entitlement per leave type
        leave_type_entitlements = {}
        for entitlement in entitlements:
            lt = entitlement.leave_type

            base_date = (
                employee.emp_joined_date
                if entitlement.effective_after_from == "date_of_joining"
                else employee.emp_date_of_confirmation
            )
            if not base_date:
                continue

            experience = relativedelta(today, base_date)
            experience_months = experience.years * 12 + experience.months

            min_exp_months = (
                entitlement.min_experience * 12
                if entitlement.effective_after_unit == "years"
                else entitlement.min_experience
            )

            if experience_months >= min_exp_months and (
                lt not in leave_type_entitlements or
                min_exp_months > leave_type_entitlements[lt]["experience"]
            ):
                leave_type_entitlements[lt] = {
                    "entitlement": entitlement,
                    "experience": min_exp_months
                }

        accrual_summary = []

        for lt in all_leave_types:
            entitlement_info = leave_type_entitlements.get(lt)
            accrual_rate = 0
            frequency = None

            if entitlement_info:
                best_entitlement = entitlement_info["entitlement"]
                accrual_rate = best_entitlement.accrual_rate
                frequency = best_entitlement.accrual_frequency

            try:
                balance_obj = emp_leave_balance.objects.get(employee=employee, leave_type=lt)
                existing_balance = balance_obj.balance
            except emp_leave_balance.DoesNotExist:
                existing_balance = 0

            monthly_accruals = []
            accrued_sum = 0

            if frequency == "months" and accrual_rate > 0:
                for i in range(12):
                    accrual_month = start_date + relativedelta(months=i)
                    accrued_sum += accrual_rate
                    monthly_accruals.append({
                        "month": accrual_month.strftime("%b %Y"),
                        "accrued_balance": round(accrued_sum, 2),
                        "total_with_existing_balance": round(accrued_sum + existing_balance, 2)
                    })

            accrual_summary.append({
                "leave_type": lt.name,
                "accrual_frequency": frequency,
                "accrual_rate": accrual_rate,
                "existing_balance": round(existing_balance, 2),
                "monthly_accruals": monthly_accruals
            })

        response_data = {
            "employee_id": employee.id,
            "employee_name": f"{employee.emp_first_name} {employee.emp_last_name}",
            "year": year,
            "leave_accrual_summary": accrual_summary
        }

        return Response(response_data, status=status.HTTP_200_OK)
    
class LatinEarlyoutEmailTemplateViewset(viewsets.ModelViewSet):
    queryset = LatinEarlyoutEmailTemplate.objects.all()
    serializer_class = LatinEarlyoutEmailTemplateSerializer
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            
            'employee': [
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}',
                '{{request_type}}',
                '{{reason}}',
                '{{ status}}',
                '{{date}}',
            ]
        }
        return Response(placeholders)
    # Custom action to fetch the available From and To addresses
    @action(detail=False, methods=['get'], url_path='from-to-addresses')
    def from_to_list(self, request):
        # Fetch active email configurations for "From" addresses
        from_addresses = EmailConfiguration.objects.filter(is_active=True).values_list('email_host_user', flat=True)

        # Fetch employee emails for "To" addresses
        to_addresses = emp_master.objects.all().values_list('emp_personal_email', 'emp_company_email')

        to_list = []
        for emp_personal, emp_company in to_addresses:
            if emp_personal:
                to_list.append(emp_personal)
            if emp_company:
                to_list.append(emp_company)

        return Response({
            'from_addresses': from_addresses,
            'to_addresses': to_list
        })
    
class LateinEarlyRequestNotificationViewset(viewsets.ReadOnlyModelViewSet):
    queryset =LateinEarlyRequestNotification.objects.all()
    serializer_class = LateinEarlyRequestNotificationSerializer   
    # def get_queryset(self):
    #     user = self.request.user

    #     # Admin / staff / superuser → see all request notifications
    #     if user.is_superuser or user.is_staff:
    #         return LateinEarlyRequestNotification.objects.all().order_by('-created_at')

    #     # Normal user → show request notifications assigned directly to them
    #     qs =LateinEarlyRequestNotification.objects.filter(
    #         Q(recipient_user=user) |
    #         Q(recipient_employee__users=user)      # employee assigned to this user
    #     ).order_by('-created_at')

    #     return qs
    
class LateinEarlyoutRequestViewset(viewsets.ModelViewSet):
    queryset = LateinEarlyoutRequest.objects.all()
    serializer_class = LateinEarlyoutRequestSerializer

    # ---------------- APPROVED REQUESTS ----------------
    @action(detail=False,methods=['get'],url_path='approved_requests' )
    def list_approved_requests(self, request):

        approved_requests = LateinEarlyoutRequest.objects.filter(status='APPROVED')

        data = []
        for req in approved_requests:
            employee = req.employee

            data.append({
                'request_id': req.id,
                'employee_id': employee.id,
                'employee_code': getattr(employee, 'emp_code', None),
                'employee_name': f"{getattr(employee, 'emp_first_name', '')} {getattr(employee, 'emp_last_name', '')}".strip(),
                'request_type': req.request_type,
                'reason': req.reason,
                'status': req.status,
                'created_at': req.created_at,
            })

        return Response(data, status=status.HTTP_200_OK)

    # ---------------- PENDING REQUESTS ----------------
    @action(detail=False,methods=['get'],url_path='pending_requests')
    def list_pending_requests(self, request):

        pending_requests = LateinEarlyoutRequest.objects.filter(status='PENDING')

        data = []
        for req in pending_requests:
            employee = req.employee

            data.append({
                'request_id': req.id,
                'employee_id': employee.id,
                'employee_code': getattr(employee, 'emp_code', None),
                'employee_name': f"{getattr(employee, 'emp_first_name', '')} {getattr(employee, 'emp_last_name', '')}".strip(),
                'request_type': req.request_type,
                'reason': req.reason,
                'status': req.status,
                'created_at': req.created_at,
            })

        return Response(data, status=status.HTTP_200_OK)

    # ---------------- EMPLOYEE REQUESTS ----------------
    @action(detail=False,methods=['get'],url_path='employee_requests/(?P<employee_id>[^/.]+)')
    def employee_requests(self, request, employee_id=None):

        requests = LateinEarlyoutRequest.objects.filter(employee_id=employee_id)

        data = []
        for req in requests:
            data.append({
                'request_id': req.id,
                'request_type': req.request_type,
                'reason': req.reason,
                'status': req.status,
                'created_at': req.created_at,
            })

        return Response(data, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'], url_path='approve')
    def approve_request(self, request, pk=None):
        req = self.get_object()
        approval = req.lateinearlyout_approvals.filter(status='PENDING').first()

        if not approval:
            return Response({'detail': 'No pending approval found'}, status=400)

        approval.approve(note=request.data.get('note'))
        return Response({'detail': 'Approved successfully'})
    
class LateinEarlyoutApprovalLevelViewset(viewsets.ModelViewSet):
    queryset = LatinEarlyApprovalWorkflow.objects.all()
    serializer_class =LatinEarlyApprovalWorkflowSerializer

class LateinEarlyoutApprovalViewset(viewsets.ModelViewSet):
    queryset = LateinEarlyoutApproval.objects.all()
    serializer_class = LateinEarlyoutApprovalSerializer
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approval = self.get_object()
        # if request.user != approval.approver:
        #     return Response({'error': 'You are not authorized to approve this request.'}, status=status.HTTP_403_FORBIDDEN)

        note = request.data.get('note')  # Get the note from the request
        approval.approve(note=note)
        return Response({'status': 'approved', 'note': note}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')  # Get the note from the request
        approval.reject(note=note)
        return Response({'status': 'rejected', 'note': note}, status=status.HTTP_200_OK)


class LeaveResetPreviewAPIView(APIView):
    """
    API to preview leave reset logic without saving.
    """

    def post(self, request):
        # Accept custom reset_date or default to today
        reset_date_str = request.data.get("reset_date")
        if reset_date_str:
            try:
                reset_date = parse_date(reset_date_str)
                if not reset_date:
                    raise ValueError
            except ValueError:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"error": "reset_date is required."}, status=status.HTTP_400_BAD_REQUEST)

        response_data = []

        resets = LeaveResetPolicy.objects.filter(reset=True)
        for reset in resets:
            reset_month = month_name_to_number(reset.month)
            reset_day = 1 if reset.day == "1st" else calendar.monthrange(reset_date.year, reset_month)[1]

            if reset.frequency == "years":
                if reset_date.month != reset_month or reset_date.day != reset_day:
                    continue
            elif reset.frequency == "months":
                if reset_date.day != reset_day:
                    continue

            leave_type = reset.leave_type
            employees = emp_master.objects.all()

            for emp in employees:
                leave_balance = emp_leave_balance.objects.filter(employee=emp, leave_type=leave_type).first()
                if not leave_balance:
                    continue

                initial_balance = leave_balance.balance if leave_balance.balance is not None else 0
                carry_forward_amount = 0
                encashment_amount = 0

                if reset.allow_cf and initial_balance > 0:
                    if reset.cf_unit_or_percentage == 'percentage':
                        calculated_cf = (initial_balance * reset.cf_value / 100)
                        carry_forward_amount = min(calculated_cf, reset.cf_max_limit or calculated_cf)
                    else:
                        carry_forward_amount = min(initial_balance, reset.cf_value)

                remaining_balance = initial_balance - carry_forward_amount

                if reset.allow_encashment and remaining_balance > 0:
                    if reset.encashment_unit_or_percentage == 'percentage':
                        encashment_amount = min((remaining_balance * reset.encashment_value / 100),
                                                reset.encashment_max_limit or remaining_balance)
                    else:
                        encashment_amount = min(remaining_balance, reset.encashment_value)

                response_data.append({
                    "employee": emp.emp_code,
                    "leave_type": leave_type.name,
                    "reset_date": reset_date,
                    "initial_balance": float(initial_balance),
                    "carry_forward": float(carry_forward_amount),
                    "encashment": float(encashment_amount),
                    "final_balance": float(carry_forward_amount),
                })

        return Response(response_data, status=status.HTTP_200_OK)


def month_name_to_number(month_name):
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    return month_map.get(month_name[:3], datetime.now().month)

class EmployeeOvertimeViewset(viewsets.ModelViewSet):
    queryset = EmployeeOvertime.objects.all()
    serializer_class = EmployeeOvertimeSerializer

import calendar

def get_month_number(month):
    if isinstance(month, int):
        return month
    if isinstance(month, str):
        month = month.strip().lower()
        month_map = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}
        return month_map.get(month)
    return None
class MonthlyAttendanceSummaryViewSet(viewsets.ModelViewSet):
    queryset = MonthlyAttendanceSummary.objects.all()
    serializer_class = MonthlyAttendanceSummarySerializer

    @action(detail=False, methods=['post'])
    def generate(self, request):

        year = int(request.data.get("year", date.today().year))

        month_input = request.data.get("month", date.today().month)
        month = get_month_number(month_input)

        if not month:
            return Response(
                {"error": "Invalid month. Use month name or number."},
                status=400
            )

        start_date = date(year, month, 1)
        end_date = start_date + relativedelta(months=1) - relativedelta(days=1)

        # 🔹 Multiple-selection filters
        employee_ids    = request.data.get("employee_ids", [])
        branch_ids      = request.data.get("branch_ids", [])
        department_ids  = request.data.get("department_ids", [])
        category_ids    = request.data.get("category_ids", [])
        designation_ids = request.data.get("designation_ids", [])

        employees = emp_master.objects.all()

        if employee_ids:
            employees = employees.filter(id__in=employee_ids)

        if branch_ids:
            employees = employees.filter(emp_branch_id__in=branch_ids)

        if department_ids:
            employees = employees.filter(emp_dept_id__in=department_ids)

        if category_ids:
            employees = employees.filter(emp_ctgry_id__in=category_ids)

        if designation_ids:
            employees = employees.filter(emp_desgntn_id__in=designation_ids)

        result = []

        for employee in employees:
            summary_data = get_attendance_summary(employee, start_date, end_date)

            if not summary_data:
                continue

            serializer = AttendanceSummarySerializer(summary_data)

            summary_obj, _ = MonthlyAttendanceSummary.objects.update_or_create(
                employee=employee,
                month=month,
                year=year,
                defaults={
                    "summary_data": serializer.data["summary"],
                    "total_present": serializer.data["total_present"],
                    "total_absent": serializer.data["total_absent"],
                }
            )

            result.append({
                "employee_id": employee.id,
                "employee_code": employee.emp_code,
                "employee_name": f"{employee.emp_first_name} {employee.emp_last_name}",
                "branch": employee.emp_branch_id.branch_name if employee.emp_branch_id else None,
                "department": employee.emp_dept_id.dept_name if employee.emp_dept_id else None,
                "category": employee.emp_ctgry_id.ctgry_title if employee.emp_ctgry_id else None,
                "designation": employee.emp_desgntn_id.desgntn_job_title if employee.emp_desgntn_id else None,
                "month": calendar.month_name[month],
                "year": year,
                "attendance": MonthlyAttendanceSummarySerializer(summary_obj).data
            })

        return Response(result)


class BulkuploadAttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = ImportAttendanceSerializer
    parser_classes = (MultiPartParser, FormParser)

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        if 'file' not in request.FILES:
            return Response({"error": "Please provide an Excel file."}, status=400)

        excel_file = request.FILES['file']
        if not excel_file.name.endswith('.xlsx'):
            return Response({"error": "Invalid file format. Only .xlsx is supported."}, status=400)

        try:
            dataset = Dataset()
            dataset.load(excel_file.read(), format='xlsx')
            resource = MonthlyAttendanceResource()

            result = resource.import_data(dataset, dry_run=False, raise_errors=True)

            return Response({"message": "Monthly attendance imported successfully."})
        except Exception as e:
            return Response({"error": str(e)}, status=400)

class LVEscalationRuleViewSet(viewsets.ModelViewSet):
    """
    API for managing escalation settings on each approval level.
    """
    serializer_class = LVEscalationRuleSerializer

    queryset = LeaveApprovalLevels.objects.all().order_by(
        'workflow__request_type',
        'level'
    )

    def get_queryset(self):
        queryset = super().get_queryset()

        request_type_id = self.request.query_params.get('request_type')
        branch_ids = self.request.query_params.get('branch')

        # ---------------- REQUEST TYPE FILTER ---------------- #
        if request_type_id:
            queryset = queryset.filter(
                workflow__request_type_id=request_type_id
            )

        # ---------------- BRANCH FILTER (FIXED) ---------------- #
        if branch_ids:
            try:
                # supports "1,2,3" OR "[1,2]" formats
                if isinstance(branch_ids, str):
                    branch_ids = branch_ids.replace('[', '').replace(']', '')
                    branch_ids = [int(x) for x in branch_ids.split(',') if x]

                queryset = queryset.filter(
                    workflow__branch__in=branch_ids
                )
            except Exception:
                pass  # fail-safe, don't break API

        return queryset.distinct()

    def update(self, request, *args, **kwargs):
        """
        Update only escalation fields for a level.
        """
        instance = self.get_object()

        serializer = self.get_serializer(
            instance,
            data=request.data,
            partial=True
        )
        serializer.is_valid(raise_exception=True)

        # ---------------- OPTIONAL SAFETY VALIDATION ---------------- #
        days = request.data.get('escalate_after_days')
        hours = request.data.get('escalate_after_hours')
        minutes = request.data.get('escalate_after_minutes')

        if days is not None and int(days) < 0:
            return Response(
                {"error": "escalate_after_days cannot be negative"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if hours is not None and int(hours) < 0:
            return Response(
                {"error": "escalate_after_hours cannot be negative"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if minutes is not None and int(minutes) < 0:
            return Response(
                {"error": "escalate_after_minutes cannot be negative"},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer.save()

        return Response(
            {
                "message": "Escalation rule updated successfully",
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'])
    def reset(self, request, pk=None):
        """
        Reset escalation configuration for a level.
        """
        instance = self.get_object()

        instance.escalate_to = None
        instance.escalate_after_days = 0
        instance.escalate_after_hours = 0
        instance.escalate_after_minutes = 0

        instance.save()

        return Response(
            {"message": "Escalation rule reset successfully"},
            status=status.HTTP_200_OK
        )

class AttendanceCalendarViewSet(viewsets.ModelViewSet):
    queryset = AttendanceCalendar.objects.all()
    serializer_class = AttendanceCalendarSerializer
    filterset_fields = ['employee', 'date', 'status']

    @action(detail=False, methods=['post'])
    def sync_range(self, request):
        from .utils import sync_attendance_calendar
        employee_id = request.data.get('employee')
        start_date_str = request.data.get('start_date')
        end_date_str = request.data.get('end_date')

        if not all([employee_id, start_date_str, end_date_str]):
            return Response({"error": "Missing parameters"}, status=400)

        try:
            employee = emp_master.objects.get(id=employee_id)
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            sync_attendance_calendar(employee, start_date, end_date)
            return Response({"message": "Calendar synced successfully"})
        except Exception as e:
            return Response({"error": str(e)}, status=400)
    @action(detail=False, methods=['get'])
    def calendar_view(self, request):
        employee_id = request.query_params.get('employee_id')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')

        if not all([employee_id, start_date_str, end_date_str]):
            return Response({"error": "employee_id, start_date, and end_date are required"}, status=400)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400)

        employee = get_object_or_404(emp_master, id=employee_id)
        
        # Get existing records
        records = AttendanceCalendar.objects.filter(
            employee=employee,
            date__range=(start_date, end_date)
        ).select_related('leave_type')
        
        record_map = {r.date: r for r in records}
        
        from .utils import get_employee_weekend_days, get_employee_holidays
        weekend_days = get_employee_weekend_days(employee)
        holiday_dates = get_employee_holidays(employee, start_date, end_date)

        calendar_data = []
        curr_date = start_date
        while curr_date <= end_date:
            record = record_map.get(curr_date)
            
            if record:
                status = record.status
                leave_name = record.leave_type.name if record.leave_type else None
                is_half_day = record.is_half_day
                remarks = record.remarks
            else:
                # Determine status on the fly if no record exists
                weekday = curr_date.strftime("%A")
                if curr_date in holiday_dates:
                    status = "Holiday"
                elif weekday in weekend_days:
                    status = "Weekend"
                else:
                    status = "Present" # Default
                
                leave_name = None
                is_half_day = False
                remarks = "Auto-determined"

            # Format for display
            display_status = status
            if status == "Leave" and leave_name:
                display_status = f"Leave ({leave_name})"
                if is_half_day:
                    display_status += " (Half Day)"
            
            calendar_data.append({
                "date": curr_date.strftime('%Y-%m-%d'),
                "day": curr_date.strftime('%A'),
                "status": status,
                "display_status": display_status,
                "leave_type": leave_name,
                "is_half_day": is_half_day,
                "remarks": remarks
            })
            curr_date += timedelta(days=1)

        return Response({
            "employee_id": employee.id,
            "employee_code": employee.emp_code,
            "employee_name": f"{employee.emp_first_name} {employee.emp_last_name}",
            "start_date": start_date_str,
            "end_date": end_date_str,
            "calendar": calendar_data
        })