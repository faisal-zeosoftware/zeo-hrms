from django.shortcuts import render
from .models import (SalaryComponent,EmployeeSalaryStructure,PayslipComponent,Payslip,PayrollRun,LoanType,LoanApplication,
                     LoanRepayment,LoanApprovalLevels,LoanApproval,PayslipApproval,PayslipCommonWorkflow,AdvanceSalaryRequest,AdvanceSalaryApproval,AdvanceCommonWorkflow,AirTicketPolicy,AirTicketAllocation,AirTicketRequest,
                     LoanEmailTemplate,LoanNotification,AdvanceSalaryEmailTemplate,AdvanceSalaryNotification,AirTicketRule,AirticketApproval,AirticketEmailTemplate,AirticketWorkflow,PayStructure,PayslipLeave,AirticketApprovalWorkflow,AdvanceApprovalWorkflow,LoanApprovalWorkflow,PayslipApprovalWorkflow,AirticketNotification,
                     SalaryRevisionHistory,SalaryStructure,LeaveEncashment)

from .serializer import (SalaryComponentSerializer,EmpBulkuploadSalaryStructureSerializer,EmployeeSalaryStructureSerializer,PayslipSerializer,PaySlipComponentSerializer,LoanTypeSerializer,LoanApplicationSerializer,LoanRepaymentSerializer,
                         LoanApprovalSerializer,LoanApprovalLevelsSerializer,PayrollRunSerializer,PayslipConfirmedSerializer,SIFSerializer,AdvanceSalaryRequestSerializer,AdvanceSalaryApprovalSerializer,AdvanceCommonWorkflowSerializer,PayslipCommonWorkflowSerializer,PayslipApprovalSerializer,AirTicketPolicySerializer,AirTicketAllocationSerializer
                         ,AirTicketRequestSerializer,LoanEmailTemplateSerializer,LoanNotificationSerializer,AdvSalaryEmailTemplateSerializer,AdvSalaryNotificationSerializer,AirTicketRuleSerializer,AirticketEmailTemplateSerializer,AirticketEscalationRuleSerializer,AirticketWorkflowSerializer,AirtcketApprovalSerializer,LoanEscalationRuleSerializer,
                         AdvSalaryEscalationRuleSerializer,PayStructureSerializer,PayslipLeaveSerializer,AirticketApprovalWorkflowSerializer,AdvanceApprovalWorkflowSerializer,LoanApprovalWorkflowSerializer,PayslipApprovalWorkflowSerializer,AirticketNotifySerializer,SalaryRevisionHistorySerializer,SalaryStructureSerializer,BenefitLiabilitySerializer,
                         DetailedPayslipSerializer,LeaveEncashmentSerializer
                         )

from rest_framework import status,generics,viewsets,permissions
from .permissions import(SalaryComponentPermission,EmployeeSalaryStructurePermission,PayrollRunPermission,PayslipComponentPermission,PayslipPermission)
from .resource import EmployeeSalaryStructureResource
from EmpManagement.models import emp_master
from rest_framework.decorators import action
from OrganisationManager.models import DocumentNumbering
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation,ROUND_HALF_UP
from rest_framework.response import Response
from django.utils import timezone
from rest_framework.exceptions import NotFound
from tablib import Dataset
from rest_framework.parsers import MultiPartParser, FormParser
from django.db import transaction
from .utils import generate_payslip_pdf,get_employee_benefit_liability,evaluate_formula
from .models import (SalaryComponent,EmployeeSalaryStructure,PayrollRun,Payslip,PayslipComponent,LoanType,LoanApplication,
                     LoanRepayment,LoanApprovalLevels,LoanApproval)
from .serializer import (SalaryComponentSerializer,EmployeeSalaryStructureSerializer,PayslipSerializer,PaySlipComponentSerializer,LoanTypeSerializer,LoanApplicationSerializer,LoanRepaymentSerializer,
                         LoanApprovalSerializer,LoanApprovalLevelsSerializer,PayrollRunSerializer)
from rest_framework import status,generics,viewsets,permissions
from datetime import datetime
import logging
from django_tenants.utils import get_tenant_model
from django.http import HttpResponse
from rest_framework.views import APIView
import csv
from rest_framework import serializers
import pytz
from .tasks import send_payslip_email_task,accrue_air_tickets
from django.db import connection
from django.db.models import Q
from openpyxl import Workbook
from tablib import Dataset
from openpyxl.styles import PatternFill,Alignment,Font,NamedStyle,Border, Side
from django.http import HttpResponse
import io
from UserManagement.models import CustomUser
from django.core.mail import send_mail
from EmpManagement.utils import send_notification_email,get_employee_context
from django.shortcuts import get_object_or_404, redirect
from calendars .models import leave_type,emp_leave_balance




# Set up logging
logger = logging.getLogger(__name__)
# Create your views here.


class SalaryComponentViewSet(viewsets.ModelViewSet):
    queryset = SalaryComponent.objects.all()
    serializer_class = SalaryComponentSerializer
    @action(detail=False, methods=['get'])
    def fixed_components(self, request):
        queryset = self.get_queryset().filter(component_value_type='fixed')
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    @action(detail=False, methods=['get'])
    def variable_components(self, request):
        queryset = self.get_queryset().filter(component_value_type='variable')
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class EmployeeSalaryStructureViewSet(viewsets.ModelViewSet):
    queryset = EmployeeSalaryStructure.objects.all()
    serializer_class = EmployeeSalaryStructureSerializer

class BenefitLiabilityAPIView(APIView):

    def get(self, request):

        as_of_date = request.GET.get(
            "as_of_date"
        )

        if not as_of_date:

            as_of_date = datetime.today().date()

        else:

            as_of_date = datetime.strptime(
                as_of_date,
                "%Y-%m-%d"
            ).date()

        employees = emp_master.objects.filter(
            is_active=True
        )

        results = []

        for employee in employees:

            results.append(
                get_employee_benefit_liability(
                    employee,
                    as_of_date
                )
            )

        serializer = (
            BenefitLiabilitySerializer(
                results,
                many=True
            )
        )

        return Response(serializer.data)

class SalaryStructureViewSet(viewsets.ModelViewSet):
    queryset = SalaryStructure.objects.all()
    serializer_class = SalaryStructureSerializer

class SalaryRevisionHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SalaryRevisionHistory.objects.select_related('employee', 'component')
    serializer_class = SalaryRevisionHistorySerializer

    def get_queryset(self):
        qs = super().get_queryset()
        employee_id = self.request.query_params.get('employee_id')
        component_id = self.request.query_params.get('component_id')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        if component_id:
            qs = qs.filter(component_id=component_id)
        return qs

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')

        data = self.get_serializer(queryset, many=True).data

        if from_date or to_date:
            for row in data:
                row['revisions'] = [
                    r for r in row['revisions']
                    if self._in_range(r['revised_on'], from_date, to_date)
                ]
            data = [row for row in data if row['revisions']]  # drop rows with no matching revisions in range

        return Response(data)

    @staticmethod
    def _in_range(revised_on, from_date, to_date):
        d = revised_on[:10]  # 'YYYY-MM-DD'
        if from_date and d < from_date:
            return False
        if to_date and d > to_date:
            return False
        return True
class PayslipViewSet(viewsets.ModelViewSet):
    queryset = Payslip.objects.all()
    serializer_class = PayslipSerializer
    @action(detail=False, methods=['get'])
    def aproved_payslips(self, request):
        aproved_payslips = self.queryset.filter(status='Approved')
        serializer = self.get_serializer(aproved_payslips, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    @action(detail=False, methods=['get'], url_path='employee/(?P<emp_code>[^/.]+)/download/(?P<year>\d{4})/(?P<month>\d{1,2})')
    def download_employee_payslip_by_month(self, request, emp_code=None, year=None, month=None):
        """Download a payslip for a specific employee for a given month and year."""
        try:
            # Ensure month and year are integers
            month = int(month)
            year = int(year)
            if not 1 <= month <= 12:
                return Response({"error": "Month must be between 1 and 12"}, status=status.HTTP_400_BAD_REQUEST)

            # Fetch the employee by emp_code
            try:
                employee = emp_master.objects.get(emp_code=emp_code)
            except emp_master.DoesNotExist:
                return Response(
                    {"error": f"No employee found with emp_code {emp_code}"}, 
                    status=status.HTTP_404_NOT_FOUND
                )

            # Fetch the payslip for the employee, month, and year
            payslip = Payslip.objects.get(
                employee=employee,
                payroll_run__month=month,
                payroll_run__year=year
            )
            return generate_payslip_pdf(request, payslip)
        except Payslip.DoesNotExist:
            return Response(
                {"error": f"No payslip found for employee {emp_code} for {month}/{year}"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except ValueError:
            return Response({"error": "Invalid year or month format"}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='employee/(?P<employee_id>\d+)/filter/(?P<year>\d{4})/(?P<month>\d{1,2})')
    def filter_employee_payslip_by_month(self, request, employee_id=None, year=None, month=None):
        """Retrieve payslip data for a specific employee for a given month and year."""
        try:
            # Ensure month and year are integers
            month = int(month)
            year = int(year)
            if not 1 <= month <= 12:
                return Response({"error": "Month must be between 1 and 12"}, status=status.HTTP_400_BAD_REQUEST)

            # Fetch the payslip for the employee, month, and year
            payslip = Payslip.objects.get(
                employee_id=employee_id,
                payroll_run__month=month,
                payroll_run__year=year
            )
            serializer = self.get_serializer(payslip)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Payslip.DoesNotExist:
            return Response(
                {"error": f"No payslip found for employee {employee_id} for {month}/{year}"}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except ValueError:
            return Response({"error": "Invalid year or month format"}, status=status.HTTP_400_BAD_REQUEST)
    @action(detail=True, methods=['post'], url_path='upload-pdf')
    def upload_pdf(self, request, pk=None):
        payslip = self.get_object()
        pdf_file = request.FILES.get('payslip_pdf')
        send_email = request.data.get('send_email', False)

        # Convert string to boolean
        if isinstance(send_email, str):
            send_email = send_email.lower() in ['true', '1', 'yes']

        if not pdf_file:
            return Response({'error': 'PDF file is required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Save PDF and flag
        payslip.payslip_pdf.save(pdf_file.name, pdf_file, save=True)
        payslip.send_email = send_email
        payslip.save(update_fields=['send_email'])

        # Trigger celery task immediately if requested
        if send_email:
            schema_name = connection.schema_name
            send_payslip_email_task.delay(payslip.id, schema_name)

        message = "PDF uploaded."
        if send_email:
            message += " Email will be sent shortly (triggered immediately)."

        return Response({'message': message}, status=status.HTTP_200_OK)

class PayslipComponentViewSet(viewsets.ModelViewSet):
    queryset = PayslipComponent.objects.all()
    serializer_class = PaySlipComponentSerializer


class PayrollRunViewSet(viewsets.ModelViewSet):
    queryset = PayrollRun.objects.all()
    serializer_class = PayrollRunSerializer
    def perform_create(self, serializer):
        with transaction.atomic():

            employees = serializer.validated_data.get('employees')
            branch = serializer.validated_data.get('branch')
            document_number = serializer.validated_data.get('document_number')

            # ✅ Branch check
            if not branch:
                # Try getting branch from selected employees
                if employees and employees.exists():
                    first_employee = employees.first()
                    branch = first_employee.emp_branch_id or first_employee.work_location

            if not branch:
                raise ValidationError("Branch is required or employee branch is missing.")

            try:
                doc_config = DocumentNumbering.objects.get(
                    branch_id=branch.id,
                    type='payroll_run',
                )
            except DocumentNumbering.DoesNotExist:
                raise NotFound(
                    f"No document numbering configuration found for branch {branch} and payslip request."
                )

            current_date = timezone.now().date()

            # ✅ Manual document validation
            if document_number:
                if doc_config.start_date and doc_config.end_date:
                    if not (doc_config.start_date <= current_date <= doc_config.end_date):
                        raise ValidationError(
                            "Document number cannot be assigned outside the valid date range."
                        )
            else:
                # ✅ Auto-generate document number
                document_number = doc_config.get_next_number()

            serializer.save(
                document_number=document_number,
                branch=branch
            )
    @action(detail=True,methods=['get'],url_path='detailed-payslips')
    def detailed_payslips(self, request, pk=None):
        # pk here IS the payroll_run id — no query param needed
        payroll_run = self.get_object()

        payslips = (
            Payslip.objects
            .filter(payroll_run=payroll_run)
            .select_related('employee', 'payroll_run')
            .prefetch_related(
                'components__component',
                'leave_details__leave_type',
            )
        )

        serializer = DetailedPayslipSerializer(
            payslips,
            many=True,
            context={'request': request}
        )

        return Response(serializer.data)
# class EmpBulkuploadSalaryStructureViewSet(viewsets.ModelViewSet):
#     queryset = EmployeeSalaryStructure.objects.all()
#     serializer_class = EmpBulkuploadSalaryStructureSerializer
#     @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
#     def bulk_upload(self, request):
#         if request.method == 'POST' and request.FILES.get('file'):
#             excel_file = request.FILES['file']
#             dataset = Dataset()
#             file_name = excel_file.name.lower()

#             try:
#                 if file_name.endswith('.xlsx'):
#                     dataset.load(excel_file.read(), format='xlsx')

#                 elif file_name.endswith('.csv'):
#                     dataset.load(
#                         excel_file.read().decode('utf-8'),
#                         format='csv'
#                     )

#                 else:
#                     return Response(
#                         {"error": "Invalid file format. Upload .xlsx or .csv only."},
#                         status=400
#                     )

#                 resource = EmployeeSalaryStructureResource()
#                 all_errors = []
#                 valid_rows = []

#                 with transaction.atomic():
#                     for row_idx, row in enumerate(dataset.dict, start=2):
#                         try:
#                             resource.before_import_row(row, row_idx=row_idx)
#                         except ValidationError as e:
#                             all_errors.extend(
#                                 [f"Row {row_idx}: {error}" for error in e.messages]
#                             )

#                 if all_errors:
#                     return Response({"errors": all_errors}, status=400)

#                 with transaction.atomic():
#                     result = resource.import_data(
#                         dataset,
#                         dry_run=False,
#                         raise_errors=True
#                     )

#                 return Response(
#                     {"message": f"{result.total_rows} records created successfully"}
#                 )

#             except Exception as e:
#                 return Response({"error": str(e)}, status=400)

#         return Response({"error": "Please provide an Excel or CSV file."}, status=400)
    
#     @action(detail=False, methods=['get'])
#     def download_default_excel_file(self, request):
#         resource = EmployeeSalaryStructureResource()
#         headers = [field.column_name for field in resource.fields.values()]
#         wb = Workbook()

#         # ======== Common Styles ========
#         black_font = Font(color="000000", bold=True)
#         blue_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
#         yellow_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")  # light cream/yellow
#         border_style = Border(
#             left=Side(style='thin'),
#             right=Side(style='thin'),
#             top=Side(style='thin'),
#             bottom=Side(style='thin')
#         )

#         # Helper function to style header row
#         def style_header_row(ws, max_cols=10):
#             """Style header row with blue fill and black bold text across full width."""
#             for col in range(1, max_cols + 1):
#                 cell = ws.cell(row=1, column=col)
#                 if not cell.value:
#                     cell.value = ""
#                 cell.fill = blue_fill
#                 cell.font = black_font
#                 cell.border = border_style
#                 ws.column_dimensions[cell.column_letter].width = 25
#             ws.freeze_panes = "A2"  # freeze header
#         # ======================================================
#         # Sheet 1: SalaryComponent
#         # ======================================================
#         ws1 = wb.active
#         ws1.title = "Salary Component"
#         for col_num, header in enumerate(headers, 1):
#             ws1.cell(row=1, column=col_num, value=header)

#         style_header_row(ws1, max_cols=len(headers))
#          # ======================================================
#         # Save response
#         # ======================================================
#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         response = HttpResponse(
#             output,
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = (
#             'attachment; filename="SalaryComponent_BulkUpload_Template.xlsx"'
#         )
#         return response
    
#     @action(detail=False, methods=['get'])
#     def download_default_csv_file(self, request):
#         resource = EmployeeSalaryStructureResource()
#         headers = [field.column_name for field in resource.fields.values()]
        
#         output = io.StringIO()
#         writer = csv.writer(output)
#         writer.writerow(headers)  # only headers, no data

#         response = HttpResponse(output.getvalue(), content_type='text/csv')
#         response['Content-Disposition'] = 'attachment; filename="Employee_SalaryComponent_Template.csv"'
#         return response
class EmpBulkuploadSalaryStructureViewSet(viewsets.ModelViewSet):
    queryset = EmployeeSalaryStructure.objects.all()
    serializer_class = EmpBulkuploadSalaryStructureSerializer
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def bulk_upload(self, request):
        if request.method != 'POST' or not request.FILES.get('file'):
            return Response({"error": "Please provide an Excel or CSV file."}, status=400)
 
        excel_file = request.FILES['file']
        file_name = excel_file.name.lower()
        dataset = Dataset()
 
        try:
            if file_name.endswith('.xlsx'):
                dataset.load(excel_file.read(), format='xlsx')
            elif file_name.endswith('.csv'):
                dataset.load(excel_file.read().decode('utf-8'), format='csv')
            else:
                return Response(
                    {"error": "Invalid file format. Upload .xlsx or .csv only."},
                    status=400
                )
        except Exception as e:
            return Response({"error": f"Could not read file: {str(e)}"}, status=400)
 
        headers = [h.strip() if h else h for h in (dataset.headers or [])]
        if not headers or headers[0] != 'Employee Code':
            return Response(
                {"error": "First column header must be exactly 'Employee Code'."},
                status=400
            )
 
        component_columns = [h for h in headers[1:] if h]
        if not component_columns:
            return Response(
                {"error": "No salary component columns found in the file."},
                status=400
            )
 
        # Validate every column header maps to a real SalaryComponent
        existing_components = {
            c.name: c for c in SalaryComponent.objects.filter(name__in=component_columns)
        }
        unknown_components = [c for c in component_columns if c not in existing_components]
        if unknown_components:
            return Response(
                {"error": f"Unknown salary component column(s): {', '.join(unknown_components)}"},
                status=400
            )
 
        all_errors = []
        valid_entries = []  # (employee, component_obj, amount)
        seen_emp_codes = set()
 
        for row_idx, row in enumerate(dataset.dict, start=2):
            emp_code = str(row.get('Employee Code') or '').strip()
            if not emp_code:
                all_errors.append(f"Row {row_idx}: Employee Code cannot be empty")
                continue
 
            if emp_code in seen_emp_codes:
                all_errors.append(f"Row {row_idx}: Duplicate Employee Code '{emp_code}' in file")
            seen_emp_codes.add(emp_code)
 
            employee = emp_master.objects.filter(emp_code=emp_code).first()
            if not employee:
                all_errors.append(
                    f"Row {row_idx}: emp_master matching query does not exist for ID: {emp_code}"
                )
                continue
 
            # Components actually assigned to this employee (i.e. an
            # EmployeeSalaryStructure row already exists for them — created
            # when they were added to a SalaryStructure via the m2m signal).
            # A component NOT in this set cannot receive an amount here.
            assigned_component_ids = set(
                EmployeeSalaryStructure.objects.filter(employee=employee)
                .values_list('component_id', flat=True)
            )
 
            row_has_component = False
            for comp_name in component_columns:
                raw_val = row.get(comp_name)
                if raw_val is None:
                    continue
                val = str(raw_val).strip()
                if val == '':
                    continue  # blank cell = no amount entered for this component, skip
 
                component_obj = existing_components[comp_name]
 
                if component_obj.id not in assigned_component_ids:
                    all_errors.append(
                        f"Row {row_idx}: Salary component '{comp_name}' is not assigned "
                        f"to employee {emp_code}"
                    )
                    continue
 
                row_has_component = True
                try:
                    amount = Decimal(val)
                except (InvalidOperation, ValueError, TypeError):
                    all_errors.append(
                        f"Row {row_idx}: Amount for '{comp_name}' must be a valid number"
                    )
                    continue
 
                if amount < 0:
                    all_errors.append(
                        f"Row {row_idx}: Amount for '{comp_name}' must be at least 0.00"
                    )
                    continue
 
                valid_entries.append((employee, component_obj, amount))
 
            if not row_has_component:
                all_errors.append(
                    f"Row {row_idx}: No valid (assigned) salary component amounts provided for {emp_code}"
                )
 
        if all_errors:
            return Response({"errors": all_errors}, status=400)
 
        updated_count = 0
        with transaction.atomic():
            for employee, component, amount in valid_entries:
                # These pairs were already confirmed to exist during
                # validation (assigned_component_ids check above), so this
                # is always an update, never a create.
                rows = EmployeeSalaryStructure.objects.filter(
                    employee=employee, component=component
                ).update(amount=amount, is_active=True)
                updated_count += rows
 
        return Response({
            "message": f"{updated_count} salary component amount(s) updated successfully"
        })
 
    # ------------------------------------------------------------------
    # DOWNLOAD TEMPLATE - EXCEL
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'])
    def download_default_excel_file(self, request):
        component_names = list(
            SalaryComponent.objects.order_by('name').values_list('name', flat=True)
        )
        headers = ['Employee Code'] + component_names
 
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Salary Structure"
 
        black_font = Font(color="000000", bold=True)
        blue_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
 
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = blue_fill
            cell.font = black_font
            cell.border = border_style
            ws.column_dimensions[cell.column_letter].width = 25
 
        ws.freeze_panes = "A2"
 
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
 
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            'attachment; filename="EmployeeSalaryStructure_BulkUpload_Template.xlsx"'
        )
        return response
 
    # ------------------------------------------------------------------
    # DOWNLOAD TEMPLATE - CSV
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'])
    def download_default_csv_file(self, request):
        component_names = list(
            SalaryComponent.objects.order_by('name').values_list('name', flat=True)
        )
        headers = ['Employee Code'] + component_names
 
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
 
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = (
            'attachment; filename="EmployeeSalaryStructure_BulkUpload_Template.csv"'
        )
        return response
class PayslipConfirmedViewSet(viewsets.ModelViewSet):
    queryset = Payslip.objects.all()
    serializer_class = PayslipConfirmedSerializer
    def get_queryset(self):
        return Payslip.objects.filter(status='processed')

class LoanTypeviewset(viewsets.ModelViewSet):
    queryset = LoanType.objects.all()
    serializer_class = LoanTypeSerializer

class LoanApplicationviewset(viewsets.ModelViewSet):
    queryset = LoanApplication.objects.all()
    serializer_class = LoanApplicationSerializer

    def perform_create(self, serializer):
        with transaction.atomic():
            employee = serializer.validated_data.get('employee')
            document_number = serializer.validated_data.get('document_number')

            # ✅ Validate employee
            if not employee:
                raise ValidationError("Employee is required.")

            # ✅ Get branch
            branch_id = employee.emp_branch_id or employee.work_location

            if not branch_id:
                raise ValidationError("Employee branch is missing in employee master.")

            try:
                # 🔥 FIX: changed type to match DB
                doc_config = DocumentNumbering.objects.get(
                    branch_id=branch_id,
                    type='loan_request'
                )
            except DocumentNumbering.DoesNotExist:
                raise NotFound(
                    f"No document numbering configuration found for branch {branch_id} and loan request."
                )

            current_date = timezone.now().date()

            # ✅ Manual document number validation
            if document_number:
                if doc_config.start_date and doc_config.end_date:
                    if not (doc_config.start_date <= current_date <= doc_config.end_date):
                        raise ValidationError(
                            "Document number cannot be assigned outside the valid date range."
                        )

                # ✅ Prevent duplicate numbers
                if LoanApplication.objects.filter(document_number=document_number).exists():
                    raise ValidationError("Document number already exists.")

            else:
                # ✅ Auto-generate
                document_number = doc_config.get_next_number()

            serializer.save(document_number=document_number)


    @action(detail=False, methods=['get'], url_path='paused-loans')
    def paused_loans(self, request):
        paused_loans = self.queryset.filter(status='Paused')
        serializer = self.get_serializer(paused_loans, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    @action(detail=True, methods=['post'])
    def pause(self, request, pk=None):
        """Pause loan repayments with a reason."""
        loan = self.get_object()
        pause_date = request.data.get('pause_start_date')
        reason = request.data.get('pause_reason')

        if not pause_date:
            return Response({"error": "Pause start date is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            loan.pause(start_date=pause_date, reason=reason)
            return Response({"status": "paused", "pause_date": pause_date, "reason": reason}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def resume(self, request, pk=None):
        """Resume loan repayments with a reason."""
        loan = self.get_object()
        resume_date = request.data.get('resume_date')
        reason = request.data.get('resume_reason')

        if not resume_date:
            return Response({"error": "Resume date is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            loan.resume(resume_date=resume_date, reason=reason)
            return Response({"status": "resumed", "resume_date": resume_date, "reason": reason}, status=status.HTTP_200_OK)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
class PayslipConfirmedViewSet(viewsets.ModelViewSet):
    queryset = Payslip.objects.all()
    serializer_class = PayslipConfirmedSerializer
    def get_queryset(self):
        return Payslip.objects.filter(status='processed')
class LoanRepaymentviewset(viewsets.ModelViewSet):
    queryset = LoanRepayment.objects.all()
    serializer_class = LoanRepaymentSerializer

class LoanApprovalLevelsviewset(viewsets.ModelViewSet):
    queryset =  LoanApprovalWorkflow.objects.all()
    serializer_class =  LoanApprovalWorkflowSerializer

class LoanApprovalviewset(viewsets.ModelViewSet):
    queryset = LoanApproval.objects.all()
    serializer_class = LoanApprovalSerializer
    lookup_field = 'pk'
    def get_queryset(self):
        """
        Filter approvals based on the authenticated user.
        """
        user = self.request.user  # Get the logged-in user
        if user.is_superuser:
            return LoanApproval.objects.all()
        return LoanApproval.objects.filter(
        Q(approver=user) |
        Q(deligate_to=user, is_deligate=True)
    ).distinct() # Filter approvals assigned to the user


    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approvals = self.get_object()
        note = request.data.get('note')  # Get the note from the request
        approvals.approve(note=note)
        return Response({'status': 'approved', 'note': note}, status=status.HTTP_200_OK)

    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')
        rejection_reason_id = request.data.get('rejection_reason')

        if not rejection_reason_id:
            raise ValidationError("Rejection reason is required.")

        # try:
        #     rejection_reason = LvRejectionReason.objects.get(id=rejection_reason_id)
        # except LvRejectionReason.DoesNotExist:
        #     raise ValidationError("Invalid rejection reason.")

        approval.reject(rejection_reason=rejection_reason_id, note=note)
        return Response({'status': 'rejected', 'note': note, 'rejection_reason': rejection_reason_id}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=["post"])
    def delegate(self, request, pk=None):
        approval = self.get_object()

        delegate_user_id = request.data.get("deligate_to")

        if not delegate_user_id:
            return Response(
                {"error": "Delegate user is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        delegate_user = get_object_or_404(CustomUser, pk=delegate_user_id)

        if delegate_user == approval.approver:
            return Response(
                {"error": "You cannot delegate to yourself."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if approval.is_deligate:
            return Response(
                {"error": "This approval has already been delegated."},
                status=status.HTTP_400_BAD_REQUEST
            )

        approval.deligate_to = delegate_user
        approval.is_deligate = True
        approval.deligate_response = None
        approval.save()

        if delegate_user.email:

            subject = "Delegation Assigned"

            message = f"""
                Delegation Assigned

                Hello {delegate_user.get_username() or delegate_user.username},

                You have been assigned a new delegation request.

                REQUEST DETAILS
                _________________

                Document Number : {approval.loan_request.document_number}
                Employee        : {approval.loan_request.employee}
                Request Type    : {approval.loan_request.loan_type}
                Status          : {approval.loan_request.status}

                Please review the request and send your response to the original approver.

                Thank You.
                """

            send_mail(
                subject,
                message,
                None,
                [delegate_user.email],
                fail_silently=False,
            )
            created_notification = send_notification_email(
                user=delegate_user,
                employee=None,
                branch=None,
                title="Delegation Assigned",
                message=f"{approval.approver.username} has delegated request {approval.loan_request.document_number} to you.",
                delegate_user=approval.approver,
                template_type="request_created",
                context={
                    **get_employee_context(approval.loan_request.employee),
                        "doc_number": approval.loan_request.document_number,
                        "request_type": approval.loan_request.loan_type,
                    },
                    email_template_model=LoanEmailTemplate,
                    notification_model=LoanNotification,
            )

            print("Notification Created:", created_notification)

            return Response(
                {
                    "message": "Approval delegated successfully.",
                    "approval_id": approval.id,
                    "approver": approval.approver.username,
                    "delegate_to": delegate_user.username,
                    "status": approval.status,
                },
                status=status.HTTP_200_OK,
            )
    @action(detail=True, methods=["post"])
    def send_response(self, request, pk=None):
        approval = self.get_object()

        response_text = request.data.get("deligate_response")

        if not response_text:
            return Response({"error": "Response is required"}, status=400)

        approval.deligate_response = response_text
        approval.save()

        # ---------------- EMAIL ----------------
        if approval.approver and approval.approver.email:
            send_mail(
                subject="Delegation Response Received",
                message=response_text,
                from_email=None,
                recipient_list=[approval.approver.email],
                fail_silently=False,
            )

        # ---------------- NOTIFICATION ----------------
        send_notification_email(
            user=approval.approver,
            employee=None,
            branch=None,
            title="Delegation Response Received",
            message=response_text,
            delegate_user=approval.deligate_to,
            template_type="request_created",
            context={
                 **get_employee_context(approval.loan_request.employee),
                    "doc_number": approval.loan_request.document_number,
                    "request_type": approval.loan_request.loan_type,
                },
                email_template_model=LoanEmailTemplate,
                notification_model=LoanNotification,
        )

        return Response({
            "message": "Response sent successfully",
            "response": response_text
        })
class SIFDataView(APIView):
    def post(self, request):
        serializer = SIFSerializer(data=request.data)
        if serializer.is_valid():
            try:
                # sif_data, total_salary = serializer.generate_sif_data()
                sif_data, total_salary, skipped_employees = serializer.generate_sif_data()
            except serializers.ValidationError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

            # Log total_salary for debugging
            logger.debug(f"total_salary type: {type(total_salary)}, value: {total_salary}")

            # Get tenant for HDR and SCR
            tenant = get_tenant_model().objects.filter(schema_name=request.tenant.schema_name).first()

            # HDR row
            hdr_row = {
                'Type': 'HDR',
                'Person ID': tenant.name if tenant and tenant.name else 'UNKNOWN_COMPANY',
                'Routing Code': '',
                'IBAN Number': '',
                'Pay Start Date': '',
                'Pay End Date': '',
                'Number of Days': '',
                'Fixed Income': '',
                'Variable Income': '',
                'Days on Leave': ''
            }

            # SCR row
            payroll_run = PayrollRun.objects.get(id=serializer.validated_data['payroll_run_id'])
            month, year = payroll_run.month, payroll_run.year

            # Get tenant's timezone dynamically
            tenant_timezone = 'UTC'
            if tenant and tenant.country and tenant.country.timezone:
                tenant_timezone = tenant.country.timezone
            try:
                tz = pytz.timezone(tenant_timezone)
            except pytz.UnknownTimeZoneError:
                tz = pytz.UTC
                logger.warning(f"Invalid timezone '{tenant_timezone}' for tenant {tenant.schema_name}. Falling back to UTC.")

            current_time = datetime.now(tz=tz)

            employer_unique_id = tenant.employer_unique_id.zfill(13) if tenant and tenant.employer_unique_id else '0' * 13
            bank_routing_code = tenant.bank_routing_code if tenant and tenant.bank_routing_code else '0' * 9

            # Handle total_salary
            salary_value = total_salary
            if isinstance(total_salary, dict):
                # Adjust 'amount' to the correct key based on your dictionary structure
                salary_value = total_salary.get('amount', 0.0)
                logger.debug(f"Extracted salary_value: {salary_value}")
            else:
                try:
                    salary_value = float(total_salary)
                except (TypeError, ValueError):
                    logger.error(f"Invalid total_salary value: {total_salary}")
                    return Response({"error": "Invalid total_salary value"}, status=status.HTTP_400_BAD_REQUEST)

            scr_row = {
                'Type': 'SCR',
                'Person ID': employer_unique_id,
                'Routing Code': bank_routing_code,
                'IBAN Number': current_time.strftime('%Y-%m-%d'),
                'Pay Start Date': current_time.strftime('%H%M'),
                'Pay End Date': f"{month:02d}{year}",
                'Number of Days': len(sif_data),
                'Fixed Income': f"{salary_value:.2f}",  # Use the numeric value
                'Variable Income': 'AED',
                'Days on Leave': ''
            }

            # Combine all rows
            response_data = {
                'status': 'success',
                'data': {
                    'hdr': hdr_row,
                    'edr': sif_data,
                    'scr': scr_row,
                    "skipped_employees": skipped_employees,  # 👈 new field
                }
            }

            return Response(response_data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class PayslipCommonWorkflowViewSet(viewsets.ModelViewSet):
    queryset = PayslipApprovalWorkflow.objects.all()
    serializer_class = PayslipApprovalWorkflowSerializer

class PayslipApprovalViewSet(viewsets.ModelViewSet):
    queryset = PayslipApproval.objects.all()
    serializer_class = PayslipApprovalSerializer
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')
        if approval.status != 'Pending':
            raise ValidationError("This approval has already been processed.")
        approval.approve(note=note)
        return Response({'status': 'approved'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')
        rejection_reason = request.data.get('rejection_reason')

        if not rejection_reason:
            raise ValidationError("Rejection reason is required.")

        if approval.status != 'Pending':
            raise ValidationError("This approval has already been processed.")

        approval.reject(rejection_reason=rejection_reason, note=note)
        return Response({'status': 'rejected'}, status=status.HTTP_200_OK)
    @action(detail=False, methods=['post'])
    def bulk_approve(self, request):
        ids = request.data.get('approval_ids', [])
        note = request.data.get('note', '')

        if not ids:
            raise ValidationError("approval_ids list is required.")

        approvals = PayslipApproval.objects.filter(id__in=ids, status='Pending')
        for approval in approvals:
            approval.approve(note=note)

        return Response({'status': f"{approvals.count()} requests approved."}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def bulk_reject(self, request):
        ids = request.data.get('approval_ids', [])
        note = request.data.get('note', '')
        rejection_reason = request.data.get('rejection_reason', '')

        if not ids:
            raise ValidationError("approval_ids list is required.")
        if not rejection_reason:
            raise ValidationError("rejection_reason is required.")

        approvals = PayslipApproval.objects.filter(id__in=ids, status='Pending')
        for approval in approvals:
            approval.reject(rejection_reason=rejection_reason, note=note)

        return Response({'status': f"{approvals.count()} requests rejected."}, status=status.HTTP_200_OK)

class AdvanceSalaryRequestViewset(viewsets.ModelViewSet):
    queryset = AdvanceSalaryRequest.objects.all()
    serializer_class = AdvanceSalaryRequestSerializer
    def perform_create(self, serializer):
        with transaction.atomic():
            employee = serializer.validated_data.get('employee')
            document_number = serializer.validated_data.get('document_number')  # Get manually entered document number

            branch_id = employee.emp_branch_id.id  

            try:
                doc_config = DocumentNumbering.objects.get(
                    branch_id=branch_id,
                    type='advance_salary_request',
                    # leave_type__isnull=True
                )
            except DocumentNumbering.DoesNotExist:
                raise NotFound(f"No document numbering configuration found for branch {branch_id} and Advance Salary request.")

            current_date = timezone.now().date()

            # Validate if the manually entered document number is within the date range
            if document_number:
                if doc_config.start_date and doc_config.end_date:
                    if not (doc_config.start_date <= current_date <= doc_config.end_date):
                        raise ValidationError("Document number cannot be assigned outside the valid date range.")
            else:
                # If no document number is entered, generate one automatically
                document_number = doc_config.get_next_number()

            serializer.save(document_number=document_number)
    @action(detail=True, methods=['post'])
    def pause(self, request, pk=None):
        """Pause advance salary request with a reason."""
        loan = self.get_object()
        pause_date = request.data.get('pause_start_date')
        reason = request.data.get('pause_reason')

        if not pause_date:
            return Response({"error": "Pause start date is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            loan.pause(start_date=pause_date, reason=reason)
            return Response({"status": "paused", "pause_date": pause_date, "reason": reason}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def resume(self, request, pk=None):
        """Resume loan repayments with a reason."""
        loan = self.get_object()
        resume_date = request.data.get('resume_date')
        reason = request.data.get('resume_reason')

        if not resume_date:
            return Response({"error": "Resume date is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            loan.resume(resume_date=resume_date, reason=reason)
            return Response({"status": "resumed", "resume_date": resume_date, "reason": reason}, status=status.HTTP_200_OK)
        except ValidationError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class AdvanceCommonWorkflowViewSet(viewsets.ModelViewSet):
    queryset = AdvanceApprovalWorkflow.objects.all()
    serializer_class = AdvanceApprovalWorkflowSerializer

class AdvanceSalaryApprovalViewSet(viewsets.ModelViewSet):
    queryset =AdvanceSalaryApproval.objects.all()
    serializer_class = AdvanceSalaryApprovalSerializer
    def get_queryset(self):
        """
        Filter approvals based on the authenticated user.
        """
        user = self.request.user  # Get the logged-in user
        if user.is_superuser:
            return AdvanceSalaryApproval.objects.all()
        return AdvanceSalaryApproval.objects.filter(
        Q(approver=user) |
        Q(deligate_to=user, is_deligate=True)
    ).distinct()  # Filter approvals assigned to the user

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')
        if approval.status != 'Pending':
            raise ValidationError("This approval has already been processed.")
        approval.approve(note=note)
        return Response({'status': 'approved'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')
        rejection_reason = request.data.get('rejection_reason')

        if not rejection_reason:
            raise ValidationError("Rejection reason is required.")

        # if approval.status != 'Pending':
        #     raise ValidationError("This approval has already been processed.")

        approval.reject(rejection_reason=rejection_reason, note=note)
        return Response({'status': 'rejected'}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=["post"])
    def delegate(self, request, pk=None):
        approval = self.get_object()

        delegate_user_id = request.data.get("deligate_to")

        if not delegate_user_id:
            return Response(
                {"error": "Delegate user is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        delegate_user = get_object_or_404(CustomUser, pk=delegate_user_id)

        if delegate_user == approval.approver:
            return Response(
                {"error": "You cannot delegate to yourself."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if approval.is_deligate:
            return Response(
                {"error": "This approval has already been delegated."},
                status=status.HTTP_400_BAD_REQUEST
            )
        approval.deligate_to = delegate_user
        approval.is_deligate = True
        approval.deligate_response = None
        approval.save()

        if delegate_user.email:

            subject = "Delegation Assigned"

            message = f"""
                Delegation Assigned

                Hello {delegate_user.get_username() or delegate_user.username},

                You have been assigned a new delegation request.

                REQUEST DETAILS
                _________________

                Document Number : {approval.request.document_number}
                Employee        : {approval.request.employee}
                Status          : {approval.request.status}

                Please review the request and send your response to the original approver.

                Thank You.
                """

            send_mail(
                subject,
                message,
                None,
                [delegate_user.email],
                fail_silently=False,
            )
            created_notification = send_notification_email(
                user=delegate_user,
                employee=None,
                branch=None,
                title="Delegation Assigned",
                message=f"{approval.approver.username} has delegated request {approval.request.document_number} to you.",
                delegate_user=approval.approver,
                template_type="request_created",
                context={
                    **get_employee_context(approval.request.employee),
                    "doc_number": approval.request.document_number,
                },
                email_template_model=AdvanceSalaryEmailTemplate,
                notification_model=AdvanceSalaryNotification
            )
            print("Notification Created:", created_notification)
            
            return Response(
            {
                "message": "Approval delegated successfully.",
                "approval_id": approval.id,
                "approver": approval.approver.username,
                "delegate_to": delegate_user.username,
                "status": approval.status,
            },
            status=status.HTTP_200_OK,
        )
    @action(detail=True, methods=["post"])
    def send_response(self, request, pk=None):
        approval = self.get_object()

        response_text = request.data.get("deligate_response")

        if not response_text:
            return Response({"error": "Response is required"}, status=400)

        approval.deligate_response = response_text
        approval.save()

        # ---------------- EMAIL ----------------
        if approval.approver and approval.approver.email:
            send_mail(
                subject="Delegation Response Received",
                message=response_text,
                from_email=None,
                recipient_list=[approval.approver.email],
                fail_silently=False,
            )

        # ---------------- NOTIFICATION ----------------
        send_notification_email(
            user=approval.approver,
            employee=None,
            branch=None,
            title="Delegation Response Received",
            message=response_text,
            delegate_user=approval.deligate_to,
            template_type="request_created",
            context={
                    **get_employee_context(approval.request.employee),
                    "doc_number": approval.request.document_number,
                },
                email_template_model=AdvanceSalaryEmailTemplate,
                notification_model=AdvanceSalaryNotification
        )

        return Response({
            "message": "Response sent successfully",
            "response": response_text
        })
    
class AirTicketRuleViewSet(viewsets.ModelViewSet):
    queryset = AirTicketRule.objects.all()
    serializer_class = AirTicketRuleSerializer
    
class AirTicketPolicyViewSet(viewsets.ModelViewSet):
    queryset = AirTicketPolicy.objects.all()
    serializer_class = AirTicketPolicySerializer
    # permission_classes = [IsAuthenticated]

class AirTicketAllocationViewSet(viewsets.ModelViewSet):
    queryset = AirTicketAllocation.objects.all()
    serializer_class = AirTicketAllocationSerializer
    # permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def trigger_auto_allocation(self, request):
        # Trigger automatic allocation task
        try:
            accrue_air_tickets.delay()
            logger.info("Automatic air ticket allocation task triggered")
            return Response({"message": "Automatic allocation task triggered"}, status=status.HTTP_202_ACCEPTED)
        except Exception as e:
            logger.error(f"Error triggering automatic allocation: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class AirTicketRequestViewSet(viewsets.ModelViewSet):
    queryset = AirTicketRequest.objects.all()
    serializer_class = AirTicketRequestSerializer
    # permission_classes = [IsAuthenticated]

    # def get_queryset(self):
    #     return AirTicketRequest.objects.filter(employee__users=self.request.user)


    def perform_create(self, serializer):
        with transaction.atomic():

            employee = serializer.validated_data.get('employee')
            document_number = serializer.validated_data.get('document_number')
            branch_id = employee.emp_branch_id

            try:
                doc_config = DocumentNumbering.objects.get(
                    branch_id=branch_id,
                    type='air_ticket_request',
                )
            except DocumentNumbering.DoesNotExist:
                raise NotFound(
                    f"No document numbering configuration found for branch {branch_id}"
                )

            current_date = timezone.now().date()

            # Validate manual document number
            if document_number:
                if doc_config.start_date and doc_config.end_date:
                    if not (doc_config.start_date <= current_date <= doc_config.end_date):
                        raise ValidationError(
                            "Document number cannot be assigned outside the valid date range."
                        )
            else:
                document_number = doc_config.get_next_number()

            serializer.save(
                document_number=document_number,
                created_by=self.request.user   # ✅ FIX HERE
            )
class AirticketWorkflowViewSet(viewsets.ModelViewSet):
    queryset = AirticketApprovalWorkflow.objects.all()
    serializer_class = AirticketApprovalWorkflowSerializer

class AirticketApprovalViewSet(viewsets.ModelViewSet):
    queryset = AirticketApproval.objects.all()
    serializer_class = AirtcketApprovalSerializer
    def get_queryset(self):
        """
        Filter approvals based on the authenticated user.
        """
        user = self.request.user  # Get the logged-in user
        if user.is_superuser:
            return AirticketApproval.objects.all()
        return AirticketApproval.objects.filter(
        Q(approver=user) |
        Q(deligate_to=user, is_deligate=True)
     ).distinct()
  # Filter approvals assigned to the user

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')
        if approval.status != 'Pending':
            raise ValidationError("This approval has already been processed.")
        approval.approve(note=note)
        return Response({'status': 'approved'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        approval = self.get_object()
        note = request.data.get('note')
        rejection_reason = request.data.get('rejection_reason')

        if not rejection_reason:
            raise ValidationError("Rejection reason is required.")

        # if approval.status != 'Pending':
        #     raise ValidationError("This approval has already been processed.")

        approval.reject(rejection_reason=rejection_reason, note=note)
        return Response({'status': 'rejected'}, status=status.HTTP_200_OK)
      
    
    @action(detail=True, methods=["post"])
    def delegate(self, request, pk=None):
        approval = self.get_object()

        delegate_user_id = request.data.get("deligate_to")

        if not delegate_user_id:
            return Response(
                {"error": "Delegate user is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        delegate_user = get_object_or_404(CustomUser, pk=delegate_user_id)

        if delegate_user == approval.approver:
            return Response(
                {"error": "You cannot delegate to yourself."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if approval.is_deligate:
            return Response(
                {"error": "This approval has already been delegated."},
                status=status.HTTP_400_BAD_REQUEST
            )

        approval.deligate_to = delegate_user
        approval.is_deligate = True
        approval.deligate_response = None
        approval.save()

        if delegate_user.email:

            subject = "Delegation Assigned"

            message = f"""
                Delegation Assigned

                Hello {delegate_user.get_username() or delegate_user.username},

                You have been assigned a new delegation request.
            
                REQUEST DETAILS
                _________________

                Document Number : {approval.request.document_number}
                Employee        : {approval.request.employee}
                Request Type    : {approval.request.request_type}
                Status          : {approval.request.status}

                Please review the request and send your response to the original approver.

                Thank You.
                """

            send_mail(
                subject,
                message,
                None,
                [delegate_user.email],
                fail_silently=False,
            )
            created_notification = send_notification_email(
                user=delegate_user,
                employee=None,
                branch=None,
                title="Delegation Assigned",
                message=f"{approval.approver.username} has delegated request {approval.request.document_number} to you.",
                delegate_user=approval.approver,
                template_type="request_created",
                context={
                    **get_employee_context(approval.request.employee),
                        "doc_number": approval.request.document_number,
                        "request_type": approval.request.request_type,
                    },
                    email_template_model=AirticketEmailTemplate,
                    notification_model=AirticketNotification,
            )

            print("Notification Created:", created_notification)

            return Response(
                {
                    "message": "Approval delegated successfully.",
                    "approval_id": approval.id,
                    "approver": approval.approver.username,
                    "delegate_to": delegate_user.username,
                    "status": approval.status,
                },
                status=status.HTTP_200_OK,
            )
    @action(detail=True, methods=["post"])
    def send_response(self, request, pk=None):
        approval = self.get_object()

        response_text = request.data.get("deligate_response")

        if not response_text:
            return Response({"error": "Response is required"}, status=400)

        approval.deligate_response = response_text
        approval.save()

        # ---------------- EMAIL ----------------
        if approval.approver and approval.approver.email:
            send_mail(
                subject="Delegation Response Received",
                message=response_text,
                from_email=None,
                recipient_list=[approval.approver.email],
                fail_silently=False,
            )

        # ---------------- NOTIFICATION ----------------
        send_notification_email(
            user=approval.approver,
            employee=None,
            branch=None,
            title="Delegation Response Received",
            message=response_text,
            delegate_user=approval.deligate_to,
            template_type="request_created",
            context={
                    **get_employee_context(approval.request.employee),
                    "doc_number": approval.request.document_number,
                    "request_type": approval.request.request_type,
                },
                email_template_model=AirticketEmailTemplate,
                notification_model=AirticketNotification,
        )

        return Response({
            "message": "Response sent successfully",
            "response": response_text
        })
 
class AirticketEmailTemplateViewSet(viewsets.ModelViewSet):
    queryset = AirticketEmailTemplate.objects.all()
    serializer_class = AirticketEmailTemplateSerializer
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            'employee': [
                '{{ document_number }}',
                '{{ recipient_name }}',
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_gender }}',
                '{{ emp_date_of_birth }}',
                '{{ emp_personal_email }}',
                '{{ emp_company_email }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}',
                '{{ requested_amount }}',
                '{{ reason }}',
                '{{ remarks }}',
                '{{ rejection_reason }}',

            ]
        }
        return Response(placeholders)
    
class AirticketNotificationsViewSet(viewsets.ModelViewSet):
    queryset = AirticketNotification.objects.all()
    serializer_class = AirticketNotifySerializer

    def get_queryset(self):
        user = self.request.user
        # Admin / staff / superuser → see all request notifications
        if user.is_superuser or user.is_staff:
            return AirticketNotification.objects.all().order_by('-created_at')
        # Normal user → show request notifications assigned directly to them
        qs = AirticketNotification.objects.filter(
            Q(recipient_user=user.id, is_deligate=False) |
            Q(recipient_employee__users=user,is_deligate=False) |
            Q(deligate_user=user.id,is_deligate=True)
        ).distinct().order_by('-created_at')

        return qs
    
class AirticketEscalationRuleViewSet(viewsets.ModelViewSet):
    """
    API for managing escalation settings on each approval level.
    """
    serializer_class = AirticketEscalationRuleSerializer
    queryset = AirticketWorkflow.objects.all().order_by('level')

    def update(self, request, *args, **kwargs):
        """
        Update only escalation fields for a level.
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response({
            "message": "Escalation rule updated successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)
    @action(detail=True, methods=['post'])
    def reset(self, request, pk=None):
        instance = self.get_object()
        instance.escalate_to = None
        instance.escalate_after_days = 0
        instance.escalate_after_hours = 0
        instance.escalate_after_minutes = 0
        instance.save()
        return Response({"message": "Escalation rule reset successfully"}, status=200)
    
class LoanEmailTemplateViewSet(viewsets.ModelViewSet):
    queryset = LoanEmailTemplate.objects.all()
    serializer_class = LoanEmailTemplateSerializer
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            'employee': [
                '{{ document_number }}',
                '{{ loan_type }}',
                '{{ reason }}',
                '{{ recipient_name }}',
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_gender }}',
                '{{ emp_date_of_birth }}',
                '{{ emp_personal_email }}',
                '{{ emp_company_email }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}',
                '{{ amount_requested }}',
                '{{ repayment_period }}',
                '{{ emi_amount }}',
                '{{ remaining_balance }}',
                '{{ status }}',
                '{{ rejection_reason }}',
            ]
        }
        return Response(placeholders)

class LoanNotificationViewSet(viewsets.ModelViewSet):
    queryset = LoanNotification.objects.all()
    serializer_class = LoanNotificationSerializer
    def get_queryset(self):
        user = self.request.user

        # Admin / staff / superuser → see all request notifications
        if user.is_superuser or user.is_staff:
            return LoanNotification.objects.all().order_by('-created_at')

        # Normal user → show request notifications assigned directly to them
        qs = LoanNotification.objects.filter(
            Q(recipient_user=user) |
            Q(recipient_employee__users=user)      # employee assigned to this user
        ).order_by('-created_at')

        return qs

class AdvSalaryEmailTemplateViewSet(viewsets.ModelViewSet):
    queryset = AdvanceSalaryEmailTemplate.objects.all()
    serializer_class = AdvSalaryEmailTemplateSerializer
    @action(detail=False, methods=['get'], url_path='placeholders')
    def placeholder_list(self, request):
        placeholders = {
            'employee': [
                '{{ document_number }}',
                '{{ reason }}',
                '{{ recipient_name }}',
                '{{ emp_first_name }}',
                '{{ emp_last_name }}',
                '{{ emp_gender }}',
                '{{ emp_date_of_birth }}',
                '{{ emp_personal_email }}',
                '{{ emp_company_email }}',
                '{{ emp_branch_name }}',
                '{{ emp_department_name }}',
                '{{ emp_designation_name }}',
                '{{ requested_amount }}',
                '{{ reason }}',
                '{{ remarks }}',
                '{{ rejection_reason }}'
            ]
        }
        return Response(placeholders)
class AdvSalaryNotificationViewSet(viewsets.ModelViewSet):
    queryset = AdvanceSalaryNotification.objects.all()
    serializer_class = AdvSalaryNotificationSerializer
    def get_queryset(self):
        user = self.request.user

        # Admin / staff / superuser → see all request notifications
        if user.is_superuser or user.is_staff:
            return AdvanceSalaryNotification.objects.all().order_by('-created_at')

        # Normal user → show request notifications assigned directly to them
        qs = AdvanceSalaryNotification.objects.filter(
            Q(recipient_user=user) |
            Q(recipient_employee__users=user)      # employee assigned to this user
        ).order_by('-created_at')

        return qs

class AdvSalaryEscalationRuleViewSet(viewsets.ModelViewSet):
    """
    API for managing escalation settings on each approval level.
    """
    serializer_class = AdvSalaryEscalationRuleSerializer
    queryset = AdvanceCommonWorkflow.objects.all().order_by('level')

    def update(self, request, *args, **kwargs):
        """
        Update only escalation fields for a level.
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response({
            "message": "Escalation rule updated successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reset(self, request, pk=None):
        instance = self.get_object()
        instance.escalate_to = None
        instance.escalate_after_days = 0
        instance.escalate_after_hours = 0
        instance.escalate_after_minutes = 0
        instance.save()

        return Response(
            {"message": "Escalation rule reset successfully"},
            status=status.HTTP_200_OK
        )

class LoanEscalationRuleViewSet(viewsets.ModelViewSet):
    """
    API for managing escalation settings on each approval level.
    """
    serializer_class = LoanEscalationRuleSerializer
    queryset = LoanApprovalLevels.objects.all().order_by('workflow__loan_type', 'level')

    def get_queryset(self):
        queryset = super().get_queryset()
        request_type_id = self.request.query_params.get('loan_type')

        if request_type_id:
            queryset = queryset.filter(loan_type_id=request_type_id)
        
        return queryset.distinct()

    def update(self, request, *args, **kwargs):
        """
        Update only escalation fields for a level.
        """
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response({
            "message": "Escalation rule updated successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)
    @action(detail=True, methods=['post'])
    def reset(self, request, pk=None):
        instance = self.get_object()
        instance.escalate_to = None
        instance.escalate_after_days = 0
        instance.escalate_after_hours = 0
        instance.escalate_after_minutes = 0
        instance.save()
        return Response({"message": "Escalation rule reset successfully"}, status=200)

class PayStructureViewSet(viewsets.ModelViewSet):
    queryset = PayStructure.objects.all()
    serializer_class = PayStructureSerializer
class PayslipLeaveViewSet(viewsets.ModelViewSet):
    queryset = PayslipLeave.objects.all()
    serializer_class = PayslipLeaveSerializer
class CalculateLeaveEncashmentAPIView(APIView):
    def post(self, request):
        employee_id = request.data.get('employee_id')
        leave_type_id = request.data.get('leave_type_id')
        encashment_days = request.data.get('encashment_days')

        if not employee_id or not leave_type_id:
            return Response({"error": "employee_id and leave_type_id are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = emp_master.objects.get(id=employee_id)
        except emp_master.DoesNotExist:
            return Response({"error": "Employee not found."}, status=status.HTTP_404_NOT_FOUND)

        from calendars.models import leave_type, emp_leave_balance
        try:
            leave_type_obj = leave_type.objects.get(id=leave_type_id)
        except leave_type.DoesNotExist:
            return Response({"error": "Leave type not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get balance
        balance_obj = emp_leave_balance.objects.filter(employee=employee, leave_type=leave_type_obj).first()
        available_balance = Decimal(str(balance_obj.balance)) if balance_obj and balance_obj.balance else Decimal("0.00")

        if encashment_days is not None:
            try:
                encashment_days = Decimal(str(encashment_days))
            except Exception:
                return Response({"error": "Invalid value for encashment_days."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            encashment_days = available_balance

        from .models import SalaryComponent, EmployeeSalaryStructure, PayStructure
        from .utils import evaluate_formula
        component = SalaryComponent.objects.filter(payroll_category='leave_encashment', branch=employee.emp_branch_id).first()

        basic_struct = EmployeeSalaryStructure.objects.filter(
            employee=employee,
            component__payroll_category='basic',
            is_active=True
        ).first()
        basic_salary = Decimal(str(basic_struct.amount)) if basic_struct and basic_struct.amount else Decimal("0.00")

        all_structs = EmployeeSalaryStructure.objects.filter(employee=employee, is_active=True)
        total_salary = sum(Decimal(str(s.amount)) for s in all_structs if s.amount)

        pay_structure = PayStructure.objects.filter(branch=employee.emp_branch_id).first()
        fixed_days = Decimal(str(pay_structure.fixed_working_days)) if pay_structure and pay_structure.fixed_working_days else Decimal("30.00")
        calendar_days = Decimal("30.00")

        if component and component.formula:
            variables = {
                "basic_salary": basic_salary,
                "total_salary": total_salary,
                "encashment_days": encashment_days,
                "leave_balance": available_balance,
                "fixed_days": fixed_days,
                "calendar_days": calendar_days,
            }
            try:
                encashment_amount = evaluate_formula(component.formula, variables, employee, component)
            except Exception as e:
                return Response({"error": f"Error evaluating formula: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            encashment_amount = (basic_salary / Decimal("30.00")) * encashment_days

        return Response({
            "employee_id": employee.id,
            "employee_name": f"{employee.emp_first_name} {employee.emp_last_name or ''}",
            "leave_type_id": leave_type_obj.id,
            "leave_type_name": leave_type_obj.name,
            "available_balance": available_balance,
            "encashment_days": encashment_days,
            "basic_salary": basic_salary,
            "formula_used": component.formula if component else "basic_salary / 30 * encashment_days",
            "encashment_amount": encashment_amount.quantize(Decimal("0.01"))
        }, status=status.HTTP_200_OK)

class LeaveEncashmentViewSet(viewsets.ModelViewSet):

    serializer_class = LeaveEncashmentSerializer

    def get_queryset(self):

        return (
            LeaveEncashment.objects
            .select_related(
                "employee",
                "leave_type",
                "approved_by",
                "payroll_run",
            )
            .all()
        )
    def create(self, request, *args, **kwargs):

        employee_id = request.data.get("employee")
        leave_type_id = request.data.get("leave_type")
        encashment_days = request.data.get("encashment_days")

        # ---------------------------------------------
        # Validate required fields
        # ---------------------------------------------

        if not employee_id:
            return Response(
                {"error": "employee is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not leave_type_id:
            return Response(
                {"error": "leave_type is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if encashment_days is None:
            return Response(
                {"error": "encashment_days is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ---------------------------------------------
        # Get employee
        # ---------------------------------------------

        try:
            employee = emp_master.objects.get(
                id=employee_id
            )
        except emp_master.DoesNotExist:
            return Response(
                {"error": "Employee not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # ---------------------------------------------
        # Get leave type
        # ---------------------------------------------

        try:
            leave_type_obj = leave_type.objects.get(
                id=leave_type_id
            )
        except leave_type.DoesNotExist:
            return Response(
                {"error": "Leave type not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # ---------------------------------------------
        # Convert days
        # ---------------------------------------------

        try:
            encashment_days = Decimal(
                str(encashment_days)
            )
        except Exception:
            return Response(
                {"error": "Invalid encashment_days."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if encashment_days <= 0:
            return Response(
                {
                    "error": (
                        "Encashment days must be "
                        "greater than zero."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # ---------------------------------------------
        # Get employee leave balance
        # ---------------------------------------------

        balance_obj = (
            emp_leave_balance.objects
            .filter(
                employee=employee,
                leave_type=leave_type_obj
            )
            .first()
        )

        available_balance = Decimal(
            str(balance_obj.balance or 0)
        ) if balance_obj else Decimal("0.00")

        # ---------------------------------------------
        # Check requested days against balance
        # ---------------------------------------------

        if encashment_days > available_balance:

            return Response(
                {
                    "error": (
                        f"Employee has only "
                        f"{available_balance} days available."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # ---------------------------------------------
        # Get leave encashment salary component
        # ---------------------------------------------

        component = (
            SalaryComponent.objects
            .filter(
                payroll_category="leave_encashment",
                branch=employee.emp_branch_id,
            )
            .first()
        )

        if not component:

            return Response(
                {
                    "error": (
                        "Leave encashment salary component "
                        "is not configured for this employee's branch."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if not component.formula:

            return Response(
                {
                    "error": (
                        "Leave encashment salary component "
                        "does not have a formula."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # ---------------------------------------------
        # Get basic salary
        # ---------------------------------------------

        basic_struct = (
            EmployeeSalaryStructure.objects
            .filter(
                employee=employee,
                component__payroll_category="basic",
                is_active=True
            )
            .first()
        )

        basic_salary = (
            Decimal(str(basic_struct.amount))
            if basic_struct and basic_struct.amount
            else Decimal("0.00")
        )

        # ---------------------------------------------
        # Get total salary
        # ---------------------------------------------

        all_structs = (
            EmployeeSalaryStructure.objects
            .filter(
                employee=employee,
                is_active=True
            )
        )

        total_salary = Decimal("0.00")

        for structure in all_structs:

            if structure.amount:
                total_salary += Decimal(
                    str(structure.amount)
                )

        # ---------------------------------------------
        # Get PayStructure
        # ---------------------------------------------

        pay_structure = (
            PayStructure.objects
            .filter(
                branch=employee.emp_branch_id
            )
            .first()
        )

        if pay_structure and pay_structure.fixed_working_days:

            fixed_days = Decimal(
                str(pay_structure.fixed_working_days)
            )

        else:

            fixed_days = Decimal("30.00")

        # ---------------------------------------------
        # Calendar days
        # ---------------------------------------------

        calendar_days = Decimal("30.00")

        # ---------------------------------------------
        # Formula variables
        # ---------------------------------------------

        variables = {

            "basic_salary": basic_salary,

            "total_salary": total_salary,

            "encashment_days": encashment_days,

            "leave_balance": available_balance,

            "fixed_days": fixed_days,

            "calendar_days": calendar_days,
        }

        # ---------------------------------------------
        # Evaluate formula
        # ---------------------------------------------

        try:

            encashment_amount = evaluate_formula(
                component.formula,
                variables,
                employee,
                component
            )

        except Exception as exc:

            return Response(
                {
                    "error": (
                        f"Error evaluating formula: {str(exc)}"
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # ---------------------------------------------
        # Decimal conversion
        # ---------------------------------------------

        encashment_amount = Decimal(
            str(encashment_amount)
        ).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP
        )

        # ---------------------------------------------
        # Create Leave Encashment
        # ---------------------------------------------

        encashment = LeaveEncashment.objects.create(

            employee=employee,

            leave_type=leave_type_obj,

            leave_balance=available_balance,

            encashment_days=encashment_days,

            basic_salary=basic_salary,

            total_salary=total_salary,

            fixed_days=fixed_days,

            calendar_days=calendar_days,

            formula_used=component.formula,

            encashment_amount=encashment_amount,

            status="pending",

            remarks=request.data.get(
                "remarks",
                ""
            ),
        )

        serializer = self.get_serializer(
            encashment
        )

        return Response(
            serializer.data,
            status=status.HTTP_201_CREATED
        )